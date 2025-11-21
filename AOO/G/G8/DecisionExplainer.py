"""
G8决策解释器
实现决策过程透明化、原因分析、影响评估、建议生成等核心功能
"""

import time
import json
import uuid
import csv
import threading
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union, Callable
from dataclasses import dataclass, field, asdict
from collections import deque, defaultdict
from enum import Enum
import logging
import numpy as np
import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity
import networkx as nx
import weakref


class DecisionType(Enum):
    """决策类型枚举"""
    STRATEGIC = "strategic"          # 战略决策
    TACTICAL = "tactical"           # 战术决策
    OPERATIONAL = "operational"     # 运营决策
    EMERGENCY = "emergency"         # 紧急决策
    ROUTINE = "routine"             # 常规决策
    COMPLEX = "complex"             # 复杂决策
    SIMPLE = "simple"               # 简单决策


class DecisionPriority(Enum):
    """决策优先级枚举"""
    CRITICAL = "critical"           # 关键
    HIGH = "high"                   # 高
    MEDIUM = "medium"               # 中
    LOW = "low"                     # 低
    BACKGROUND = "background"       # 后台


class DecisionStatus(Enum):
    """决策状态枚举"""
    PENDING = "pending"             # 待处理
    ANALYZING = "analyzing"         # 分析中
    DECIDED = "decided"             # 已决策
    EXECUTING = "executing"         # 执行中
    COMPLETED = "completed"         # 已完成
    CANCELLED = "cancelled"         # 已取消
    FAILED = "failed"               # 失败


@dataclass
class DecisionContext:
    """决策上下文"""
    decision_id: str
    timestamp: datetime
    decision_type: DecisionType
    priority: DecisionPriority
    title: str
    description: str
    context_data: Dict[str, Any] = field(default_factory=dict)
    stakeholders: List[str] = field(default_factory=list)
    constraints: List[str] = field(default_factory=list)
    objectives: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['timestamp'] = self.timestamp.isoformat()
        data['decision_type'] = self.decision_type.value
        data['priority'] = self.priority.value
        return data


@dataclass
class DecisionReason:
    """决策原因"""
    reason_id: str
    category: str
    description: str
    evidence: List[str] = field(default_factory=list)
    weight: float = 0.0
    confidence: float = 0.0
    source: str = ""
    timestamp: datetime = field(default_factory=datetime.now)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['timestamp'] = self.timestamp.isoformat()
        return data


@dataclass
class DecisionImpact:
    """决策影响"""
    impact_id: str
    category: str
    description: str
    magnitude: float  # 影响程度 (-1.0 到 1.0)
    probability: float  # 发生概率 (0.0 到 1.0)
    timeframe: str  # 影响时间范围
    affected_entities: List[str] = field(default_factory=list)
    measurement: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        return data


@dataclass
class DecisionExplanation:
    """决策解释"""
    explanation_id: str
    decision_id: str
    explanation_text: str
    reasoning_chain: List[str] = field(default_factory=list)
    key_factors: List[str] = field(default_factory=list)
    alternative_analysis: Dict[str, str] = field(default_factory=dict)
    confidence_score: float = 0.0
    clarity_score: float = 0.0
    generated_at: datetime = field(default_factory=datetime.now)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['generated_at'] = self.generated_at.isoformat()
        return data


@dataclass
class DecisionRecommendation:
    """决策建议"""
    recommendation_id: str
    decision_id: str
    title: str
    description: str
    action_items: List[str] = field(default_factory=list)
    priority: DecisionPriority = DecisionPriority.MEDIUM
    expected_outcome: str = ""
    risk_assessment: str = ""
    resource_requirements: List[str] = field(default_factory=list)
    timeline: str = ""
    success_criteria: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['priority'] = self.priority.value
        return data


@dataclass
class DecisionFeedback:
    """决策反馈"""
    feedback_id: str
    decision_id: str
    feedback_type: str  # "user", "system", "external"
    rating: float  # 1-5分
    comments: str = ""
    outcome_data: Dict[str, Any] = field(default_factory=dict)
    timestamp: datetime = field(default_factory=datetime.now)
    follow_up_actions: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['timestamp'] = self.timestamp.isoformat()
        return data


@dataclass
class DecisionNode:
    """决策节点"""
    node_id: str
    decision_id: str
    node_type: str  # "context", "reason", "impact", "alternative"
    content: str
    metadata: Dict[str, Any] = field(default_factory=dict)
    timestamp: datetime = field(default_factory=datetime.now)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['timestamp'] = self.timestamp.isoformat()
        return data


@dataclass
class DecisionEdge:
    """决策边"""
    edge_id: str
    source_node_id: str
    target_node_id: str
    relationship_type: str
    weight: float = 1.0
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        return data


@dataclass
class DecisionTrace:
    """决策轨迹"""
    trace_id: str
    decision_id: str
    nodes: List[DecisionNode] = field(default_factory=list)
    edges: List[DecisionEdge] = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['created_at'] = self.created_at.isoformat()
        data['nodes'] = [node.to_dict() for node in self.nodes]
        data['edges'] = [edge.to_dict() for edge in self.edges]
        return data


@dataclass
class DecisionReport:
    """决策报告"""
    report_id: str
    decision_id: str
    generated_at: datetime
    report_type: str  # "comprehensive", "summary", "analysis"
    executive_summary: str
    decision_context: Dict[str, Any]
    analysis_results: Dict[str, Any]
    explanations: List[Dict[str, Any]]
    recommendations: List[Dict[str, Any]]
    impact_assessment: Dict[str, Any]
    learning_insights: List[str]
    charts: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['generated_at'] = self.generated_at.isoformat()
        return data


class DecisionExplainer:
    """决策解释器核心类"""
    
    def __init__(self, 
                 max_history_size: int = 10000,
                 enable_learning: bool = True,
                 explanation_depth: str = "detailed"):
        """
        初始化决策解释器
        
        Args:
            max_history_size: 最大历史记录数量
            enable_learning: 是否启用学习功能
            explanation_depth: 解释深度 ("simple", "detailed", "comprehensive")
        """
        self.max_history_size = max_history_size
        self.enable_learning = enable_learning
        self.explanation_depth = explanation_depth
        
        # 数据存储
        self.decision_history: deque = deque(maxlen=max_history_size)
        self.active_decisions: Dict[str, Dict[str, Any]] = {}
        self.decision_traces: Dict[str, DecisionTrace] = {}
        self.feedback_history: List[DecisionFeedback] = []
        
        # 分析模型
        self.reason_classifier = TfidfVectorizer(max_features=1000, stop_words='english')
        self.impact_predictor = None
        self.confidence_calculator = None
        
        # 学习数据
        self.learning_patterns: Dict[str, Any] = {}
        self.success_patterns: Dict[str, List[str]] = defaultdict(list)
        self.failure_patterns: Dict[str, List[str]] = defaultdict(list)
        
        # 线程锁
        self._lock = threading.RLock()
        
        # 日志
        self.logger = logging.getLogger(__name__)
        
        # 初始化模型
        self._initialize_models()
        
        self.logger.info("决策解释器初始化完成")
    
    def _initialize_models(self):
        """初始化分析模型"""
        try:
            # 初始化原因分类器（使用模拟数据进行训练）
            sample_reasons = [
                "基于历史数据分析",
                "遵循公司政策",
                "响应市场需求",
                "降低成本考虑",
                "提高效率需求",
                "风险规避",
                "合规要求",
                "技术限制"
            ]
            sample_categories = ["数据驱动", "政策", "市场", "成本", "效率", "风险", "合规", "技术"]
            
            if len(sample_reasons) > 1:
                self.reason_classifier.fit(sample_reasons)
            
            self.logger.info("分析模型初始化完成")
        except Exception as e:
            self.logger.warning(f"模型初始化失败: {e}")
    
    def create_decision_context(self, 
                              decision_type: DecisionType,
                              priority: DecisionPriority,
                              title: str,
                              description: str,
                              context_data: Optional[Dict[str, Any]] = None,
                              stakeholders: Optional[List[str]] = None,
                              constraints: Optional[List[str]] = None,
                              objectives: Optional[List[str]] = None) -> str:
        """
        创建决策上下文
        
        Args:
            decision_type: 决策类型
            priority: 决策优先级
            title: 决策标题
            description: 决策描述
            context_data: 上下文数据
            stakeholders: 利益相关者
            constraints: 约束条件
            objectives: 目标
            
        Returns:
            决策ID
        """
        decision_id = str(uuid.uuid4())
        
        context = DecisionContext(
            decision_id=decision_id,
            timestamp=datetime.now(),
            decision_type=decision_type,
            priority=priority,
            title=title,
            description=description,
            context_data=context_data or {},
            stakeholders=stakeholders or [],
            constraints=constraints or [],
            objectives=objectives or []
        )
        
        with self._lock:
            self.active_decisions[decision_id] = {
                'context': context,
                'status': DecisionStatus.PENDING,
                'reasons': [],
                'impacts': [],
                'explanations': [],
                'recommendations': [],
                'created_at': datetime.now()
            }
            
            # 创建决策轨迹
            trace = DecisionTrace(
                trace_id=str(uuid.uuid4()),
                decision_id=decision_id
            )
            self.decision_traces[decision_id] = trace
            
            # 添加上下文节点
            context_node = DecisionNode(
                node_id=str(uuid.uuid4()),
                decision_id=decision_id,
                node_type="context",
                content=f"决策创建: {title}",
                metadata=context.to_dict()
            )
            trace.nodes.append(context_node)
        
        self.logger.info(f"创建决策上下文: {decision_id} - {title}")
        return decision_id
    
    def analyze_decision_reasons(self, 
                               decision_id: str,
                               reasons: List[str],
                               evidence: Optional[Dict[str, List[str]]] = None) -> List[DecisionReason]:
        """
        分析决策原因
        
        Args:
            decision_id: 决策ID
            reasons: 原因列表
            evidence: 证据数据
            
        Returns:
            分析后的决策原因列表
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        analyzed_reasons = []
        evidence = evidence or {}
        
        with self._lock:
            for i, reason_text in enumerate(reasons):
                # 分类原因
                category = self._classify_reason(reason_text)
                
                # 计算权重和置信度
                weight = self._calculate_reason_weight(reason_text, category)
                confidence = self._calculate_confidence(reason_text, evidence.get(f"reason_{i}", []))
                
                reason = DecisionReason(
                    reason_id=str(uuid.uuid4()),
                    category=category,
                    description=reason_text,
                    evidence=evidence.get(f"reason_{i}", []),
                    weight=weight,
                    confidence=confidence,
                    source="analysis"
                )
                
                analyzed_reasons.append(reason)
                self.active_decisions[decision_id]['reasons'].append(reason)
                
                # 添加到决策轨迹
                if decision_id in self.decision_traces:
                    reason_node = DecisionNode(
                        node_id=str(uuid.uuid4()),
                        decision_id=decision_id,
                        node_type="reason",
                        content=f"原因分析: {reason_text}",
                        metadata=reason.to_dict()
                    )
                    self.decision_traces[decision_id].nodes.append(reason_node)
        
        self.logger.info(f"分析决策原因完成: {decision_id}, {len(analyzed_reasons)}个原因")
        return analyzed_reasons
    
    def assess_decision_impacts(self, 
                              decision_id: str,
                              impacts: List[Dict[str, Any]]) -> List[DecisionImpact]:
        """
        评估决策影响
        
        Args:
            decision_id: 决策ID
            impacts: 影响描述列表
            
        Returns:
            评估后的决策影响列表
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        assessed_impacts = []
        
        with self._lock:
            for impact_data in impacts:
                # 计算影响程度和概率
                magnitude = self._calculate_impact_magnitude(impact_data)
                probability = self._calculate_impact_probability(impact_data)
                
                impact = DecisionImpact(
                    impact_id=str(uuid.uuid4()),
                    category=impact_data.get('category', 'unknown'),
                    description=impact_data.get('description', ''),
                    magnitude=magnitude,
                    probability=probability,
                    timeframe=impact_data.get('timeframe', 'unknown'),
                    affected_entities=impact_data.get('affected_entities', []),
                    measurement=impact_data.get('measurement', '')
                )
                
                assessed_impacts.append(impact)
                self.active_decisions[decision_id]['impacts'].append(impact)
                
                # 添加到决策轨迹
                if decision_id in self.decision_traces:
                    impact_node = DecisionNode(
                        node_id=str(uuid.uuid4()),
                        decision_id=decision_id,
                        node_type="impact",
                        content=f"影响评估: {impact.description}",
                        metadata=impact.to_dict()
                    )
                    self.decision_traces[decision_id].nodes.append(impact_node)
        
        self.logger.info(f"评估决策影响完成: {decision_id}, {len(assessed_impacts)}个影响")
        return assessed_impacts
    
    def generate_decision_explanation(self, 
                                    decision_id: str,
                                    explanation_type: str = "comprehensive") -> DecisionExplanation:
        """
        生成决策解释
        
        Args:
            decision_id: 决策ID
            explanation_type: 解释类型
            
        Returns:
            决策解释对象
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        decision_data = self.active_decisions[decision_id]
        context = decision_data['context']
        reasons = decision_data['reasons']
        impacts = decision_data['impacts']
        
        # 构建解释文本
        explanation_text = self._build_explanation_text(context, reasons, impacts, explanation_type)
        
        # 构建推理链
        reasoning_chain = self._build_reasoning_chain(reasons, impacts)
        
        # 识别关键因素
        key_factors = self._identify_key_factors(reasons, impacts)
        
        # 分析替代方案
        alternative_analysis = self._analyze_alternatives(context, reasons)
        
        # 计算置信度和清晰度分数
        confidence_score = self._calculate_explanation_confidence(reasons, impacts)
        clarity_score = self._calculate_explanation_clarity(explanation_text)
        
        explanation = DecisionExplanation(
            explanation_id=str(uuid.uuid4()),
            decision_id=decision_id,
            explanation_text=explanation_text,
            reasoning_chain=reasoning_chain,
            key_factors=key_factors,
            alternative_analysis=alternative_analysis,
            confidence_score=confidence_score,
            clarity_score=clarity_score
        )
        
        with self._lock:
            decision_data['explanations'].append(explanation)
            
            # 添加到决策轨迹
            if decision_id in self.decision_traces:
                explanation_node = DecisionNode(
                    node_id=str(uuid.uuid4()),
                    decision_id=decision_id,
                    node_type="explanation",
                    content=f"解释生成: {explanation_type}",
                    metadata=explanation.to_dict()
                )
                self.decision_traces[decision_id].nodes.append(explanation_node)
        
        self.logger.info(f"生成决策解释完成: {decision_id}")
        return explanation
    
    def generate_recommendations(self, 
                               decision_id: str,
                               recommendation_type: str = "action") -> List[DecisionRecommendation]:
        """
        生成决策建议
        
        Args:
            decision_id: 决策ID
            recommendation_type: 建议类型
            
        Returns:
            决策建议列表
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        decision_data = self.active_decisions[decision_id]
        context = decision_data['context']
        reasons = decision_data['reasons']
        impacts = decision_data['impacts']
        
        recommendations = []
        
        with self._lock:
            # 基于原因生成建议
            reason_based_recs = self._generate_reason_based_recommendations(reasons)
            recommendations.extend(reason_based_recs)
            
            # 基于影响生成建议
            impact_based_recs = self._generate_impact_based_recommendations(impacts)
            recommendations.extend(impact_based_recs)
            
            # 基于历史模式生成建议
            if self.enable_learning:
                pattern_based_recs = self._generate_pattern_based_recommendations(context, reasons, impacts)
                recommendations.extend(pattern_based_recs)
            
            # 去重和排序
            unique_recs = self._deduplicate_recommendations(recommendations)
            sorted_recs = self._sort_recommendations_by_priority(unique_recs)
            
            # 添加到决策数据
            for rec in sorted_recs:
                decision_data['recommendations'].append(rec)
                
                # 添加到决策轨迹
                if decision_id in self.decision_traces:
                    rec_node = DecisionNode(
                        node_id=str(uuid.uuid4()),
                        decision_id=decision_id,
                        node_type="recommendation",
                        content=f"建议生成: {rec.title}",
                        metadata=rec.to_dict()
                    )
                    self.decision_traces[decision_id].nodes.append(rec_node)
        
        self.logger.info(f"生成决策建议完成: {decision_id}, {len(sorted_recs)}条建议")
        return sorted_recs
    
    def collect_feedback(self, 
                        decision_id: str,
                        feedback_type: str,
                        rating: float,
                        comments: str = "",
                        outcome_data: Optional[Dict[str, Any]] = None) -> str:
        """
        收集决策反馈
        
        Args:
            decision_id: 决策ID
            feedback_type: 反馈类型
            rating: 评分 (1-5)
            comments: 评论
            outcome_data: 结果数据
            
        Returns:
            反馈ID
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        feedback = DecisionFeedback(
            feedback_id=str(uuid.uuid4()),
            decision_id=decision_id,
            feedback_type=feedback_type,
            rating=rating,
            comments=comments,
            outcome_data=outcome_data or {}
        )
        
        with self._lock:
            self.feedback_history.append(feedback)
            self.active_decisions[decision_id]['feedback'] = feedback
            
            # 学习反馈
            if self.enable_learning:
                self._learn_from_feedback(feedback)
            
            # 添加到决策轨迹
            if decision_id in self.decision_traces:
                feedback_node = DecisionNode(
                    node_id=str(uuid.uuid4()),
                    decision_id=decision_id,
                    node_type="feedback",
                    content=f"反馈收集: {feedback_type}",
                    metadata=feedback.to_dict()
                )
                self.decision_traces[decision_id].nodes.append(feedback_node)
        
        self.logger.info(f"收集决策反馈完成: {decision_id}, 评分: {rating}")
        return feedback.feedback_id
    
    def generate_decision_report(self, 
                               decision_id: str,
                               report_type: str = "comprehensive") -> DecisionReport:
        """
        生成决策报告
        
        Args:
            decision_id: 决策ID
            report_type: 报告类型
            
        Returns:
            决策报告
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        decision_data = self.active_decisions[decision_id]
        context = decision_data['context']
        reasons = decision_data['reasons']
        impacts = decision_data['impacts']
        explanations = decision_data['explanations']
        recommendations = decision_data['recommendations']
        
        # 生成执行摘要
        executive_summary = self._generate_executive_summary(context, reasons, impacts)
        
        # 分析结果
        analysis_results = self._perform_comprehensive_analysis(decision_id)
        
        # 影响评估
        impact_assessment = self._comprehensive_impact_assessment(impacts)
        
        # 学习洞察
        learning_insights = self._extract_learning_insights(decision_id)
        
        # 生成图表
        charts = self._generate_decision_charts(decision_id, report_type)
        
        report = DecisionReport(
            report_id=str(uuid.uuid4()),
            decision_id=decision_id,
            generated_at=datetime.now(),
            report_type=report_type,
            executive_summary=executive_summary,
            decision_context=context.to_dict(),
            analysis_results=analysis_results,
            explanations=[exp.to_dict() for exp in explanations],
            recommendations=[rec.to_dict() for rec in recommendations],
            impact_assessment=impact_assessment,
            learning_insights=learning_insights,
            charts=charts
        )
        
        self.logger.info(f"生成决策报告完成: {decision_id}")
        return report
    
    def get_decision_transparency_score(self, decision_id: str) -> float:
        """
        计算决策透明度分数
        
        Args:
            decision_id: 决策ID
            
        Returns:
            透明度分数 (0.0 - 1.0)
        """
        if decision_id not in self.active_decisions:
            return 0.0
        
        decision_data = self.active_decisions[decision_id]
        
        # 评估各维度透明度
        context_completeness = self._assess_context_completeness(decision_data['context'])
        reason_clarity = self._assess_reason_clarity(decision_data['reasons'])
        impact_coverage = self._assess_impact_coverage(decision_data['impacts'])
        explanation_quality = self._assess_explanation_quality(decision_data['explanations'])
        
        # 综合计算透明度分数
        transparency_score = (
            context_completeness * 0.25 +
            reason_clarity * 0.25 +
            impact_coverage * 0.25 +
            explanation_quality * 0.25
        )
        
        return min(1.0, max(0.0, transparency_score))
    
    def export_decision_data(self, 
                           decision_id: str,
                           format: str = "json",
                           include_charts: bool = True) -> str:
        """
        导出决策数据
        
        Args:
            decision_id: 决策ID
            format: 导出格式 ("json", "csv", "xlsx")
            include_charts: 是否包含图表
            
        Returns:
            导出文件路径
        """
        if decision_id not in self.active_decisions:
            raise ValueError(f"决策ID不存在: {decision_id}")
        
        decision_data = self.active_decisions[decision_id]
        trace = self.decision_traces.get(decision_id)
        
        # 准备导出数据
        export_data = {
            'decision_id': decision_id,
            'context': decision_data['context'].to_dict(),
            'reasons': [r.to_dict() for r in decision_data['reasons']],
            'impacts': [i.to_dict() for i in decision_data['impacts']],
            'explanations': [e.to_dict() for e in decision_data['explanations']],
            'recommendations': [r.to_dict() for r in decision_data['recommendations']],
            'trace': trace.to_dict() if trace else None,
            'exported_at': datetime.now().isoformat()
        }
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"decision_{decision_id[:8]}_{timestamp}"
        
        if format == "json":
            filepath = f"/tmp/{filename}.json"
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        elif format == "csv":
            filepath = f"/tmp/{filename}.csv"
            self._export_to_csv(export_data, filepath)
        
        elif format == "xlsx":
            filepath = f"/tmp/{filename}.xlsx"
            self._export_to_excel(export_data, filepath)
        
        else:
            raise ValueError(f"不支持的导出格式: {format}")
        
        self.logger.info(f"导出决策数据完成: {filepath}")
        return filepath
    
    def _classify_reason(self, reason_text: str) -> str:
        """分类决策原因"""
        try:
            # 使用TF-IDF进行简单分类
            categories = {
                "数据驱动": ["数据", "分析", "统计", "证据"],
                "政策": ["政策", "规定", "制度", "标准"],
                "市场": ["市场", "竞争", "客户", "需求"],
                "成本": ["成本", "预算", "费用", "经济"],
                "效率": ["效率", "优化", "改进", "提升"],
                "风险": ["风险", "安全", "控制", "防范"],
                "合规": ["合规", "法律", "监管", "审计"],
                "技术": ["技术", "系统", "平台", "工具"]
            }
            
            reason_lower = reason_text.lower()
            scores = {}
            
            for category, keywords in categories.items():
                score = sum(1 for keyword in keywords if keyword in reason_lower)
                scores[category] = score
            
            # 返回得分最高的类别
            if scores and max(scores.values()) > 0:
                return max(scores, key=scores.get)
            else:
                return "其他"
                
        except Exception as e:
            self.logger.warning(f"原因分类失败: {e}")
            return "其他"
    
    def _calculate_reason_weight(self, reason_text: str, category: str) -> float:
        """计算原因权重"""
        # 基于文本长度和关键词计算权重
        base_weight = min(1.0, len(reason_text) / 100.0)
        
        # 根据类别调整权重
        category_weights = {
            "数据驱动": 1.2,
            "政策": 1.1,
            "风险": 1.3,
            "合规": 1.2,
            "市场": 1.0,
            "成本": 1.0,
            "效率": 0.9,
            "技术": 0.8,
            "其他": 0.7
        }
        
        category_multiplier = category_weights.get(category, 0.7)
        return min(1.0, base_weight * category_multiplier)
    
    def _calculate_confidence(self, reason_text: str, evidence: List[str]) -> float:
        """计算置信度"""
        # 基于证据数量和质量计算置信度
        evidence_score = min(1.0, len(evidence) / 3.0)
        
        # 基于文本质量计算置信度
        quality_score = 1.0 if len(reason_text) > 20 else 0.5
        
        return (evidence_score + quality_score) / 2.0
    
    def _calculate_impact_magnitude(self, impact_data: Dict[str, Any]) -> float:
        """计算影响程度"""
        description = impact_data.get('description', '')
        
        # 基于描述中的关键词评估影响程度
        positive_keywords = ['提升', '改善', '增长', '优化', '增强']
        negative_keywords = ['降低', '减少', '恶化', '风险', '损失']
        
        positive_score = sum(1 for kw in positive_keywords if kw in description)
        negative_score = sum(1 for kw in negative_keywords if kw in description)
        
        magnitude = (positive_score - negative_score) / max(1, positive_score + negative_score)
        return max(-1.0, min(1.0, magnitude))
    
    def _calculate_impact_probability(self, impact_data: Dict[str, Any]) -> float:
        """计算影响概率"""
        timeframe = impact_data.get('timeframe', '')
        
        # 基于时间范围评估概率
        if '立即' in timeframe or '马上' in timeframe:
            return 0.9
        elif '短期' in timeframe or '近期' in timeframe:
            return 0.7
        elif '中期' in timeframe:
            return 0.5
        elif '长期' in timeframe:
            return 0.3
        else:
            return 0.5
    
    def _build_explanation_text(self, context: DecisionContext, reasons: List[DecisionReason], 
                              impacts: List[DecisionImpact], explanation_type: str) -> str:
        """构建解释文本"""
        explanation_parts = []
        
        # 决策概述
        explanation_parts.append(f"决策概述：{context.title}")
        explanation_parts.append(f"决策类型：{context.decision_type.value}")
        explanation_parts.append(f"决策描述：{context.description}")
        
        # 决策原因
        if reasons:
            explanation_parts.append("\n决策原因：")
            for i, reason in enumerate(reasons, 1):
                explanation_parts.append(f"{i}. {reason.description} (置信度: {reason.confidence:.2f})")
        
        # 预期影响
        if impacts:
            explanation_parts.append("\n预期影响：")
            for i, impact in enumerate(impacts, 1):
                explanation_parts.append(f"{i}. {impact.description} (程度: {impact.magnitude:.2f})")
        
        # 详细程度调整
        if explanation_type == "simple":
            explanation_parts = explanation_parts[:4]  # 只保留概述
        elif explanation_type == "comprehensive":
            # 添加更多细节
            explanation_parts.append("\n详细分析：")
            explanation_parts.append("本决策基于多维度分析，考虑了内外部环境因素，平衡了短期和长期利益。")
        
        return "\n".join(explanation_parts)
    
    def _build_reasoning_chain(self, reasons: List[DecisionReason], impacts: List[DecisionImpact]) -> List[str]:
        """构建推理链"""
        chain = []
        
        # 添加原因链
        chain.append("决策推理链：")
        chain.append("1. 问题识别 → 2. 原因分析 → 3. 方案评估 → 4. 影响预测 → 5. 决策制定")
        
        # 具体原因
        if reasons:
            chain.append("\n关键原因：")
            for reason in reasons:
                chain.append(f"- {reason.category}: {reason.description}")
        
        # 影响预测
        if impacts:
            chain.append("\n影响预测：")
            for impact in impacts:
                chain.append(f"- {impact.category}: {impact.description}")
        
        return chain
    
    def _identify_key_factors(self, reasons: List[DecisionReason], impacts: List[DecisionImpact]) -> List[str]:
        """识别关键因素"""
        factors = []
        
        # 基于权重和置信度识别关键原因
        key_reasons = [r for r in reasons if r.weight > 0.7 and r.confidence > 0.6]
        factors.extend([f"关键原因-{r.category}" for r in key_reasons])
        
        # 基于影响程度识别关键影响
        key_impacts = [i for i in impacts if abs(i.magnitude) > 0.5]
        factors.extend([f"关键影响-{i.category}" for i in key_impacts])
        
        return factors[:10]  # 限制数量
    
    def _analyze_alternatives(self, context: DecisionContext, reasons: List[DecisionReason]) -> Dict[str, str]:
        """分析替代方案"""
        alternatives = {}
        
        # 基于决策类型生成替代方案
        if context.decision_type == DecisionType.STRATEGIC:
            alternatives["激进方案"] = "采取更大胆的行动，追求更高收益但伴随更高风险"
            alternatives["保守方案"] = "采取稳健的行动，确保安全但可能错失机会"
            alternatives["渐进方案"] = "采取渐进式的改进措施，平衡风险和收益"
        
        elif context.decision_type == DecisionType.OPERATIONAL:
            alternatives["自动化方案"] = "通过技术手段提高效率，减少人工干预"
            alternatives["人工方案"] = "保持人工操作，确保灵活性和适应性"
            alternatives["混合方案"] = "结合自动化和人工操作的优势"
        
        else:
            alternatives["标准方案"] = "采用标准的处理流程和方法"
            alternatives["定制方案"] = "根据具体情况定制专门的解决方案"
        
        return alternatives
    
    def _calculate_explanation_confidence(self, reasons: List[DecisionReason], impacts: List[DecisionImpact]) -> float:
        """计算解释置信度"""
        if not reasons and not impacts:
            return 0.0
        
        reason_confidence = np.mean([r.confidence for r in reasons]) if reasons else 0.0
        impact_confidence = np.mean([abs(i.magnitude) for i in impacts]) if impacts else 0.0
        
        return (reason_confidence + impact_confidence) / 2.0
    
    def _calculate_explanation_clarity(self, explanation_text: str) -> float:
        """计算解释清晰度"""
        # 基于文本长度和结构评估清晰度
        sentences = explanation_text.split('。')
        avg_sentence_length = np.mean([len(s.strip()) for s in sentences if s.strip()])
        
        # 理想句子长度在20-50字符之间
        if 20 <= avg_sentence_length <= 50:
            clarity = 1.0
        elif avg_sentence_length < 20:
            clarity = 0.7
        else:
            clarity = 0.8
        
        return clarity
    
    def _generate_reason_based_recommendations(self, reasons: List[DecisionReason]) -> List[DecisionRecommendation]:
        """基于原因生成建议"""
        recommendations = []
        
        for reason in reasons:
            if reason.category == "数据驱动":
                rec = DecisionRecommendation(
                    recommendation_id=str(uuid.uuid4()),
                    decision_id="",
                    title="建立数据监控体系",
                    description=f"针对{reason.description}，建议建立持续的数据监控和分析体系",
                    action_items=["设置关键指标", "建立监控仪表板", "定期数据分析"],
                    priority=DecisionPriority.HIGH
                )
                recommendations.append(rec)
            
            elif reason.category == "风险":
                rec = DecisionRecommendation(
                    recommendation_id=str(uuid.uuid4()),
                    decision_id="",
                    title="风险控制措施",
                    description=f"针对{reason.description}，建议制定风险控制措施",
                    action_items=["风险识别", "风险评估", "风险缓解策略"],
                    priority=DecisionPriority.CRITICAL
                )
                recommendations.append(rec)
        
        return recommendations
    
    def _generate_impact_based_recommendations(self, impacts: List[DecisionImpact]) -> List[DecisionRecommendation]:
        """基于影响生成建议"""
        recommendations = []
        
        for impact in impacts:
            if impact.magnitude > 0.5:  # 正面影响
                rec = DecisionRecommendation(
                    recommendation_id=str(uuid.uuid4()),
                    decision_id="",
                    title=f"放大{impact.category}影响",
                    description=f"为了最大化{impact.description}的正面影响，建议采取以下措施",
                    action_items=["制定行动计划", "分配资源", "跟踪进展"],
                    priority=DecisionPriority.HIGH
                )
                recommendations.append(rec)
            
            elif impact.magnitude < -0.5:  # 负面影响
                rec = DecisionRecommendation(
                    recommendation_id=str(uuid.uuid4()),
                    decision_id="",
                    title=f"缓解{impact.category}风险",
                    description=f"为了减少{impact.description}的负面影响，建议采取缓解措施",
                    action_items=["风险评估", "预防措施", "应急预案"],
                    priority=DecisionPriority.CRITICAL
                )
                recommendations.append(rec)
        
        return recommendations
    
    def _generate_pattern_based_recommendations(self, context: DecisionContext, 
                                              reasons: List[DecisionReason], 
                                              impacts: List[DecisionImpact]) -> List[DecisionRecommendation]:
        """基于历史模式生成建议"""
        recommendations = []
        
        # 基于成功模式生成建议
        context_key = f"{context.decision_type.value}_{context.priority.value}"
        if context_key in self.success_patterns:
            for pattern in self.success_patterns[context_key]:
                rec = DecisionRecommendation(
                    recommendation_id=str(uuid.uuid4()),
                    decision_id="",
                    title="基于历史成功经验",
                    description=f"根据历史成功模式：{pattern}",
                    action_items=["应用最佳实践", "借鉴成功经验", "持续改进"],
                    priority=DecisionPriority.MEDIUM
                )
                recommendations.append(rec)
        
        return recommendations
    
    def _deduplicate_recommendations(self, recommendations: List[DecisionRecommendation]) -> List[DecisionRecommendation]:
        """去重建议"""
        seen_titles = set()
        unique_recs = []
        
        for rec in recommendations:
            if rec.title not in seen_titles:
                seen_titles.add(rec.title)
                unique_recs.append(rec)
        
        return unique_recs
    
    def _sort_recommendations_by_priority(self, recommendations: List[DecisionRecommendation]) -> List[DecisionRecommendation]:
        """按优先级排序建议"""
        priority_order = {
            DecisionPriority.CRITICAL: 0,
            DecisionPriority.HIGH: 1,
            DecisionPriority.MEDIUM: 2,
            DecisionPriority.LOW: 3,
            DecisionPriority.BACKGROUND: 4
        }
        
        return sorted(recommendations, key=lambda x: priority_order.get(x.priority, 5))
    
    def _learn_from_feedback(self, feedback: DecisionFeedback):
        """从反馈中学习"""
        try:
            decision_data = self.active_decisions.get(feedback.decision_id)
            if not decision_data:
                return
            
            context = decision_data['context']
            context_key = f"{context.decision_type.value}_{context.priority.value}"
            
            if feedback.rating >= 4.0:  # 正面反馈
                self.success_patterns[context_key].append(f"成功模式_{feedback.feedback_id[:8]}")
            elif feedback.rating <= 2.0:  # 负面反馈
                self.failure_patterns[context_key].append(f"失败模式_{feedback.feedback_id[:8]}")
            
            self.logger.debug(f"学习反馈完成: {context_key}")
            
        except Exception as e:
            self.logger.warning(f"学习反馈失败: {e}")
    
    def _generate_executive_summary(self, context: DecisionContext, reasons: List[DecisionReason], 
                                  impacts: List[DecisionImpact]) -> str:
        """生成执行摘要"""
        summary_parts = []
        
        summary_parts.append(f"决策：{context.title}")
        summary_parts.append(f"类型：{context.decision_type.value} | 优先级：{context.priority.value}")
        
        if reasons:
            top_reasons = sorted(reasons, key=lambda x: x.weight * x.confidence, reverse=True)[:3]
            summary_parts.append(f"关键原因：{', '.join([r.description for r in top_reasons])}")
        
        if impacts:
            significant_impacts = [i for i in impacts if abs(i.magnitude) > 0.3]
            if significant_impacts:
                summary_parts.append(f"主要影响：{', '.join([i.description for i in significant_impacts[:3]])}")
        
        return " | ".join(summary_parts)
    
    def _perform_comprehensive_analysis(self, decision_id: str) -> Dict[str, Any]:
        """执行综合分析"""
        decision_data = self.active_decisions[decision_id]
        
        analysis = {
            "decision_complexity": self._assess_decision_complexity(decision_data),
            "reason_strength": self._assess_reason_strength(decision_data['reasons']),
            "impact_scope": self._assess_impact_scope(decision_data['impacts']),
            "stakeholder_alignment": self._assess_stakeholder_alignment(decision_data['context']),
            "risk_level": self._assess_risk_level(decision_data),
            "success_probability": self._estimate_success_probability(decision_data)
        }
        
        return analysis
    
    def _comprehensive_impact_assessment(self, impacts: List[DecisionImpact]) -> Dict[str, Any]:
        """综合影响评估"""
        if not impacts:
            return {"overall_impact": 0.0, "impact_categories": {}}
        
        # 按类别分组影响
        categories = {}
        for impact in impacts:
            if impact.category not in categories:
                categories[impact.category] = []
            categories[impact.category].append(impact)
        
        # 计算各类别影响
        category_scores = {}
        for category, category_impacts in categories.items():
            weighted_score = sum(i.magnitude * i.probability for i in category_impacts)
            category_scores[category] = weighted_score
        
        # 计算总体影响
        overall_impact = sum(category_scores.values()) / len(category_scores)
        
        return {
            "overall_impact": overall_impact,
            "impact_categories": category_scores,
            "positive_impacts": [i for i in impacts if i.magnitude > 0],
            "negative_impacts": [i for i in impacts if i.magnitude < 0],
            "high_probability_impacts": [i for i in impacts if i.probability > 0.7]
        }
    
    def _extract_learning_insights(self, decision_id: str) -> List[str]:
        """提取学习洞察"""
        insights = []
        
        decision_data = self.active_decisions[decision_id]
        context = decision_data['context']
        
        # 基于决策类型的洞察
        if context.decision_type == DecisionType.STRATEGIC:
            insights.append("战略决策需要更全面的风险评估和长期影响分析")
        
        elif context.decision_type == DecisionType.EMERGENCY:
            insights.append("紧急决策需要在速度和准确性之间找到平衡")
        
        # 基于优先级的洞察
        if context.priority == DecisionPriority.CRITICAL:
            insights.append("关键决策需要更严格的验证和多重确认机制")
        
        # 基于历史模式的洞察
        if self.enable_learning:
            context_key = f"{context.decision_type.value}_{context.priority.value}"
            if context_key in self.success_patterns:
                insights.append("该决策类型具有成功的历史模式可供参考")
            if context_key in self.failure_patterns:
                insights.append("该决策类型存在需要避免的失败模式")
        
        return insights
    
    def _generate_decision_charts(self, decision_id: str, report_type: str) -> List[str]:
        """生成决策图表"""
        charts = []
        
        try:
            decision_data = self.active_decisions[decision_id]
            
            # 生成原因权重图
            if decision_data['reasons']:
                chart_path = self._create_reason_weight_chart(decision_data['reasons'], decision_id)
                charts.append(chart_path)
            
            # 生成影响评估图
            if decision_data['impacts']:
                chart_path = self._create_impact_assessment_chart(decision_data['impacts'], decision_id)
                charts.append(chart_path)
            
            # 生成决策轨迹图
            if decision_id in self.decision_traces:
                chart_path = self._create_decision_trace_chart(decision_id)
                charts.append(chart_path)
            
        except Exception as e:
            self.logger.warning(f"生成图表失败: {e}")
        
        return charts
    
    def _create_reason_weight_chart(self, reasons: List[DecisionReason], decision_id: str) -> str:
        """创建原因权重图"""
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            reason_labels = [f"{r.category}\n{r.description[:30]}..." for r in reasons]
            weights = [r.weight * r.confidence for r in reasons]
            
            bars = ax.barh(reason_labels, weights)
            ax.set_xlabel('权重 × 置信度')
            ax.set_title(f'决策原因分析 - {decision_id[:8]}')
            
            # 设置颜色
            for i, bar in enumerate(bars):
                bar.set_color(plt.cm.viridis(weights[i]))
            
            plt.tight_layout()
            
            chart_path = f"/tmp/decision_{decision_id[:8]}_reasons.png"
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
            
        except Exception as e:
            self.logger.warning(f"创建原因权重图失败: {e}")
            return ""
    
    def _create_impact_assessment_chart(self, impacts: List[DecisionImpact], decision_id: str) -> str:
        """创建影响评估图"""
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            impact_labels = [f"{i.category}\n{i.description[:30]}..." for i in impacts]
            magnitudes = [i.magnitude for i in impacts]
            probabilities = [i.probability for i in impacts]
            
            # 创建散点图
            scatter = ax.scatter(probabilities, magnitudes, 
                               s=[abs(m) * 200 for m in magnitudes],
                               c=magnitudes, cmap='RdYlBu_r', alpha=0.7)
            
            ax.set_xlabel('发生概率')
            ax.set_ylabel('影响程度')
            ax.set_title(f'决策影响评估 - {decision_id[:8]}')
            ax.grid(True, alpha=0.3)
            
            # 添加颜色条
            plt.colorbar(scatter, ax=ax, label='影响程度')
            
            plt.tight_layout()
            
            chart_path = f"/tmp/decision_{decision_id[:8]}_impacts.png"
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
            
        except Exception as e:
            self.logger.warning(f"创建影响评估图失败: {e}")
            return ""
    
    def _create_decision_trace_chart(self, decision_id: str) -> str:
        """创建决策轨迹图"""
        try:
            trace = self.decision_traces[decision_id]
            
            # 创建网络图
            G = nx.DiGraph()
            
            # 添加节点
            for node in trace.nodes:
                G.add_node(node.node_id, **node.metadata)
            
            # 添加边（基于时间顺序）
            for i in range(len(trace.nodes) - 1):
                G.add_edge(trace.nodes[i].node_id, trace.nodes[i + 1].node_id)
            
            # 绘制网络图
            plt.figure(figsize=(12, 8))
            pos = nx.spring_layout(G, k=1, iterations=50)
            
            # 绘制节点
            node_colors = []
            for node in trace.nodes:
                color_map = {
                    'context': 'lightblue',
                    'reason': 'lightgreen',
                    'impact': 'lightcoral',
                    'explanation': 'lightyellow',
                    'recommendation': 'lightpink',
                    'feedback': 'lightgray'
                }
                node_colors.append(color_map.get(node.node_type, 'white'))
            
            nx.draw_networkx_nodes(G, pos, node_color=node_colors, 
                                 node_size=1000, alpha=0.8)
            nx.draw_networkx_edges(G, pos, edge_color='gray', 
                                 arrows=True, arrowsize=20, alpha=0.6)
            
            # 添加标签
            labels = {node.node_id: node.node_type for node in trace.nodes}
            nx.draw_networkx_labels(G, pos, labels, font_size=8)
            
            plt.title(f'决策轨迹图 - {decision_id[:8]}')
            plt.axis('off')
            plt.tight_layout()
            
            chart_path = f"/tmp/decision_{decision_id[:8]}_trace.png"
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
            
        except Exception as e:
            self.logger.warning(f"创建决策轨迹图失败: {e}")
            return ""
    
    def _assess_context_completeness(self, context: DecisionContext) -> float:
        """评估上下文完整性"""
        score = 0.0
        
        if context.description:
            score += 0.3
        if context.stakeholders:
            score += 0.2
        if context.constraints:
            score += 0.2
        if context.objectives:
            score += 0.3
        
        return score
    
    def _assess_reason_clarity(self, reasons: List[DecisionReason]) -> float:
        """评估原因清晰度"""
        if not reasons:
            return 0.0
        
        clarity_scores = []
        for reason in reasons:
            score = 0.0
            if reason.description:
                score += 0.4
            if reason.evidence:
                score += 0.3
            if reason.confidence > 0:
                score += 0.3
            clarity_scores.append(score)
        
        return np.mean(clarity_scores)
    
    def _assess_impact_coverage(self, impacts: List[DecisionImpact]) -> float:
        """评估影响覆盖度"""
        if not impacts:
            return 0.0
        
        coverage_score = 0.0
        
        # 检查不同类型的影响
        categories = set(i.category for i in impacts)
        expected_categories = {'经济', '社会', '技术', '环境', '风险'}
        
        coverage_score = len(categories & expected_categories) / len(expected_categories)
        
        return coverage_score
    
    def _assess_explanation_quality(self, explanations: List[DecisionExplanation]) -> float:
        """评估解释质量"""
        if not explanations:
            return 0.0
        
        quality_scores = []
        for explanation in explanations:
            score = 0.0
            score += explanation.confidence_score * 0.4
            score += explanation.clarity_score * 0.4
            score += len(explanation.reasoning_chain) / 10.0 * 0.2  # 推理链长度
            quality_scores.append(min(1.0, score))
        
        return np.mean(quality_scores)
    
    def _assess_decision_complexity(self, decision_data: Dict[str, Any]) -> float:
        """评估决策复杂性"""
        context = decision_data['context']
        reasons = decision_data['reasons']
        impacts = decision_data['impacts']
        
        complexity = 0.0
        
        # 基于决策类型
        type_complexity = {
            DecisionType.STRATEGIC: 1.0,
            DecisionType.COMPLEX: 0.9,
            DecisionType.TACTICAL: 0.7,
            DecisionType.OPERATIONAL: 0.5,
            DecisionType.EMERGENCY: 0.3,
            DecisionType.ROUTINE: 0.2,
            DecisionType.SIMPLE: 0.1
        }
        complexity += type_complexity.get(context.decision_type, 0.5)
        
        # 基于原因数量
        complexity += min(0.3, len(reasons) * 0.05)
        
        # 基于影响数量
        complexity += min(0.2, len(impacts) * 0.03)
        
        return min(1.0, complexity)
    
    def _assess_reason_strength(self, reasons: List[DecisionReason]) -> float:
        """评估原因强度"""
        if not reasons:
            return 0.0
        
        strengths = [r.weight * r.confidence for r in reasons]
        return np.mean(strengths)
    
    def _assess_impact_scope(self, impacts: List[DecisionImpact]) -> float:
        """评估影响范围"""
        if not impacts:
            return 0.0
        
        # 基于影响实体数量
        all_entities = set()
        for impact in impacts:
            all_entities.update(impact.affected_entities)
        
        scope_score = min(1.0, len(all_entities) / 10.0)
        
        return scope_score
    
    def _assess_stakeholder_alignment(self, context: DecisionContext) -> float:
        """评估利益相关者一致性"""
        if not context.stakeholders:
            return 0.5  # 中性分数
        
        # 假设利益相关者数量越多，一致性越难达成
        alignment_score = 1.0 - (len(context.stakeholders) - 1) * 0.1
        return max(0.0, min(1.0, alignment_score))
    
    def _assess_risk_level(self, decision_data: Dict[str, Any]) -> float:
        """评估风险等级"""
        impacts = decision_data['impacts']
        
        if not impacts:
            return 0.5
        
        # 计算负面影响的总体风险
        negative_impacts = [i for i in impacts if i.magnitude < 0]
        risk_score = sum(abs(i.magnitude) * i.probability for i in negative_impacts)
        
        return min(1.0, risk_score)
    
    def _estimate_success_probability(self, decision_data: Dict[str, Any]) -> float:
        """估算成功概率"""
        context = decision_data['context']
        reasons = decision_data['reasons']
        impacts = decision_data['impacts']
        
        # 基于多个因素估算成功概率
        probability = 0.5  # 基础概率
        
        # 原因质量影响
        if reasons:
            reason_quality = np.mean([r.confidence for r in reasons])
            probability += reason_quality * 0.2
        
        # 正面影响比例
        if impacts:
            positive_ratio = len([i for i in impacts if i.magnitude > 0]) / len(impacts)
            probability += positive_ratio * 0.2
        
        # 决策类型影响
        type_probability = {
            DecisionType.ROUTINE: 0.8,
            DecisionType.SIMPLE: 0.7,
            DecisionType.OPERATIONAL: 0.6,
            DecisionType.TACTICAL: 0.5,
            DecisionType.STRATEGIC: 0.4,
            DecisionType.COMPLEX: 0.3,
            DecisionType.EMERGENCY: 0.4
        }
        probability += type_probability.get(context.decision_type, 0.5) - 0.5
        
        return max(0.0, min(1.0, probability))
    
    def _export_to_csv(self, data: Dict[str, Any], filepath: str):
        """导出为CSV格式"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # 写入决策信息
            writer.writerow(['决策信息'])
            writer.writerow(['决策ID', data['decision_id']])
            writer.writerow(['决策标题', data['context']['title']])
            writer.writerow(['决策类型', data['context']['decision_type']])
            writer.writerow([])
            
            # 写入原因
            writer.writerow(['决策原因'])
            writer.writerow(['类别', '描述', '权重', '置信度'])
            for reason in data['reasons']:
                writer.writerow([reason['category'], reason['description'], 
                               reason['weight'], reason['confidence']])
            writer.writerow([])
            
            # 写入影响
            writer.writerow(['决策影响'])
            writer.writerow(['类别', '描述', '程度', '概率'])
            for impact in data['impacts']:
                writer.writerow([impact['category'], impact['description'],
                               impact['magnitude'], impact['probability']])
    
    def _export_to_excel(self, data: Dict[str, Any], filepath: str):
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            
            # 决策信息工作表
            ws1 = wb.active
            ws1.title = "决策信息"
            ws1['A1'] = "决策ID"
            ws1['B1'] = data['decision_id']
            ws1['A2'] = "决策标题"
            ws1['B2'] = data['context']['title']
            
            # 原因工作表
            ws2 = wb.create_sheet("决策原因")
            headers = ['类别', '描述', '权重', '置信度']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, reason in enumerate(data['reasons'], 2):
                ws2.cell(row=row, column=1, value=reason['category'])
                ws2.cell(row=row, column=2, value=reason['description'])
                ws2.cell(row=row, column=3, value=reason['weight'])
                ws2.cell(row=row, column=4, value=reason['confidence'])
            
            # 影响工作表
            ws3 = wb.create_sheet("决策影响")
            headers = ['类别', '描述', '程度', '概率']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, impact in enumerate(data['impacts'], 2):
                ws3.cell(row=row, column=1, value=impact['category'])
                ws3.cell(row=row, column=2, value=impact['description'])
                ws3.cell(row=row, column=3, value=impact['magnitude'])
                ws3.cell(row=row, column=4, value=impact['probability'])
            
            wb.save(filepath)
            
        except ImportError:
            # 如果没有openpyxl，回退到CSV
            self.logger.warning("openpyxl未安装，回退到CSV格式")
            csv_filepath = filepath.replace('.xlsx', '.csv')
            self._export_to_csv(data, csv_filepath)
    
    def get_decision_statistics(self) -> Dict[str, Any]:
        """获取决策统计信息"""
        with self._lock:
            stats = {
                'total_decisions': len(self.active_decisions),
                'completed_decisions': len([d for d in self.active_decisions.values() 
                                          if d.get('status') == DecisionStatus.COMPLETED]),
                'average_transparency_score': np.mean([
                    self.get_decision_transparency_score(did) 
                    for did in self.active_decisions.keys()
                ]) if self.active_decisions else 0.0,
                'decision_types': {},
                'priority_distribution': {},
                'feedback_count': len(self.feedback_history),
                'learning_patterns_count': len(self.learning_patterns)
            }
            
            # 决策类型分布
            for decision_data in self.active_decisions.values():
                decision_type = decision_data['context'].decision_type.value
                stats['decision_types'][decision_type] = stats['decision_types'].get(decision_type, 0) + 1
            
            # 优先级分布
            for decision_data in self.active_decisions.values():
                priority = decision_data['context'].priority.value
                stats['priority_distribution'][priority] = stats['priority_distribution'].get(priority, 0) + 1
            
            return stats
    
    def cleanup_old_data(self, days: int = 30):
        """清理旧数据"""
        cutoff_date = datetime.now() - timedelta(days=days)
        
        with self._lock:
            # 清理过期的决策
            expired_decisions = []
            for decision_id, decision_data in self.active_decisions.items():
                if decision_data['created_at'] < cutoff_date:
                    expired_decisions.append(decision_id)
            
            for decision_id in expired_decisions:
                del self.active_decisions[decision_id]
                if decision_id in self.decision_traces:
                    del self.decision_traces[decision_id]
            
            # 清理过期的反馈
            self.feedback_history = [
                f for f in self.feedback_history 
                if f.timestamp > cutoff_date
            ]
            
            self.logger.info(f"清理完成，删除{len(expired_decisions)}个过期决策")


# 示例使用函数
def demo_decision_explainer():
    """决策解释器演示"""
    print("=== G8决策解释器演示 ===")
    
    # 创建决策解释器
    explainer = DecisionExplainer(
        max_history_size=1000,
        enable_learning=True,
        explanation_depth="detailed"
    )
    
    # 1. 创建决策上下文
    decision_id = explainer.create_decision_context(
        decision_type=DecisionType.STRATEGIC,
        priority=DecisionPriority.HIGH,
        title="实施新的客户管理系统",
        description="为了提高客户服务质量和运营效率，决定实施新的CRM系统",
        context_data={
            "budget": 500000,
            "timeline": "6个月",
            "department": "客户服务部"
        },
        stakeholders=["客服经理", "IT总监", "财务总监", "CEO"],
        constraints=["预算限制", "时间限制", "技术兼容性"],
        objectives=["提高客户满意度", "降低运营成本", "提升工作效率"]
    )
    
    print(f"创建决策: {decision_id}")
    
    # 2. 分析决策原因
    reasons = [
        "现有系统老化，维护成本高昂",
        "客户满意度调查显示系统使用困难",
        "竞争对手已采用更先进的CRM系统",
        "新系统能够集成多个销售渠道"
    ]
    
    evidence = {
        "reason_0": ["系统使用年限5年", "年维护费用20万", "故障频率增加"],
        "reason_1": ["满意度评分3.2/5", "客户投诉增加15%"],
        "reason_2": ["市场份额下降2%", "客户流失率上升"],
        "reason_3": ["可整合微信、APP、网站等渠道"]
    }
    
    analyzed_reasons = explainer.analyze_decision_reasons(decision_id, reasons, evidence)
    print(f"分析原因: {len(analyzed_reasons)}个")
    
    # 3. 评估决策影响
    impacts = [
        {
            "category": "经济",
            "description": "初期投资50万，但预期3年内节省运营成本30万/年",
            "timeframe": "3年",
            "affected_entities": ["IT部门", "客服部门", "财务部门"],
            "measurement": "ROI"
        },
        {
            "category": "社会",
            "description": "客户满意度预期提升20%",
            "timeframe": "6个月",
            "affected_entities": ["客户", "客服团队"],
            "measurement": "满意度评分"
        },
        {
            "category": "技术",
            "description": "系统集成复杂度增加，需要额外培训",
            "timeframe": "1年",
            "affected_entities": ["IT团队", "客服人员"],
            "measurement": "培训时长"
        },
        {
            "category": "风险",
            "description": "系统迁移期间可能出现业务中断",
            "timeframe": "2个月",
            "affected_entities": ["所有客户", "客服业务"],
            "measurement": "业务中断时间"
        }
    ]
    
    assessed_impacts = explainer.assess_decision_impacts(decision_id, impacts)
    print(f"评估影响: {len(assessed_impacts)}个")
    
    # 4. 生成决策解释
    explanation = explainer.generate_decision_explanation(decision_id, "comprehensive")
    print(f"生成解释: 置信度{explanation.confidence_score:.2f}")
    
    # 5. 生成决策建议
    recommendations = explainer.generate_recommendations(decision_id)
    print(f"生成建议: {len(recommendations)}条")
    
    # 6. 计算透明度分数
    transparency_score = explainer.get_decision_transparency_score(decision_id)
    print(f"透明度分数: {transparency_score:.2f}")
    
    # 7. 生成决策报告
    report = explainer.generate_decision_report(decision_id)
    print(f"生成报告: {report.report_id}")
    
    # 8. 收集反馈
    feedback_id = explainer.collect_feedback(
        decision_id=decision_id,
        feedback_type="user",
        rating=4.5,
        comments="决策过程清晰，建议实用性强",
        outcome_data={"implementation_success": True, "roi_achieved": True}
    )
    print(f"收集反馈: {feedback_id}")
    
    # 9. 导出数据
    export_file = explainer.export_decision_data(decision_id, format="json")
    print(f"导出数据: {export_file}")
    
    # 10. 获取统计信息
    stats = explainer.get_decision_statistics()
    print(f"统计信息: {stats}")
    
    print("\n=== 演示完成 ===")
    return explainer, decision_id


if __name__ == "__main__":
    # 运行演示
    explainer, decision_id = demo_decision_explainer()
    
    print(f"\n决策解释器已创建，共处理{len(explainer.active_decisions)}个决策")
    print(f"透明度分数: {explainer.get_decision_transparency_score(decision_id):.2f}")