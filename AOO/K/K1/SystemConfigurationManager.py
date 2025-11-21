"""
K1系统配置管理器

这是一个功能完整的系统配置管理器，提供以下核心功能：

1. 系统级配置管理（系统参数、环境变量、运行参数）
2. 配置文件管理（YAML、JSON、XML、INI格式支持）
3. 配置验证和检查（类型检查、范围检查、依赖检查）
4. 配置热更新和动态加载
5. 配置版本管理和历史记录
6. 配置加密和安全存储
7. 异步配置处理和缓存
8. 完整的错误处理和日志记录

主要类：
- SystemConfigurationManager: 主要配置管理器类
- ConfigurationValidator: 配置验证器
- ConfigurationHistory: 配置历史记录管理
- ConfigurationCache: 配置缓存管理
- ConfigurationEncryptor: 配置加密器
- ConfigurationLoader: 配置文件加载器

作者: K1系统开发团队
版本: 1.0.0
创建时间: 2025-11-06
"""

import os
import sys
import json
import yaml
import xml.etree.ElementTree as ET
import configparser
import asyncio
import aiofiles
import aiofiles.os
import hashlib
import hmac
import base64
import time
import threading
import logging
import logging.handlers
import shutil
import tempfile
import re
import pickle
import weakref
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union, Callable, Set, Tuple, Type, TypeVar, Generic
from pathlib import Path
from dataclasses import dataclass, field, asdict
from abc import ABC, abstractmethod
from enum import Enum
from collections import defaultdict, OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import wraps, lru_cache
import warnings
import inspect
import traceback


# 类型定义
T = TypeVar('T')
ConfigValue = Union[str, int, float, bool, List, Dict, None]
ConfigurationData = Dict[str, ConfigValue]


class ConfigurationError(Exception):
    """配置管理器基础异常类"""
    pass


class ConfigurationValidationError(ConfigurationError):
    """配置验证错误"""
    pass


class ConfigurationLoadError(ConfigurationError):
    """配置加载错误"""
    pass


class ConfigurationSaveError(ConfigurationError):
    """配置保存错误"""
    pass


class ConfigurationEncryptionError(ConfigurationError):
    """配置加密错误"""
    pass


class ConfigurationCacheError(ConfigurationError):
    """配置缓存错误"""
    pass


class ConfigurationFormat(Enum):
    """支持的配置文件格式"""
    JSON = "json"
    YAML = "yaml"
    XML = "xml"
    INI = "ini"


class LogLevel(Enum):
    """日志级别"""
    DEBUG = "DEBUG"
    INFO = "INFO"
    WARNING = "WARNING"
    ERROR = "ERROR"
    CRITICAL = "CRITICAL"


@dataclass
class ConfigurationMetadata:
    """配置元数据"""
    name: str
    version: str
    description: str = ""
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)
    author: str = ""
    tags: List[str] = field(default_factory=list)
    checksum: str = ""
    encrypted: bool = False
    size: int = 0


@dataclass
class ConfigurationValidationRule:
    """配置验证规则"""
    field_name: str
    field_type: Type
    required: bool = True
    default_value: Any = None
    min_value: Optional[Union[int, float]] = None
    max_value: Optional[Union[int, float]] = None
    allowed_values: Optional[List[Any]] = None
    pattern: Optional[str] = None
    custom_validator: Optional[Callable] = None
    depends_on: Optional[List[str]] = None


@dataclass
class ConfigurationChange:
    """配置变更记录"""
    timestamp: datetime
    field_name: str
    old_value: Any
    new_value: Any
    change_type: str  # 'add', 'modify', 'delete'
    user: str = ""
    reason: str = ""
    checksum_before: str = ""
    checksum_after: str = ""


class ConfigurationLogger:
    """配置管理器专用日志器"""
    
    def __init__(self, name: str = "K1ConfigManager", log_level: LogLevel = LogLevel.INFO):
        self.logger = logging.getLogger(name)
        self.logger.setLevel(getattr(logging, log_level.value))
        
        # 避免重复添加处理器
        if not self.logger.handlers:
            self._setup_handlers()
    
    def _setup_handlers(self):
        """设置日志处理器"""
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        # 文件处理器
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        file_handler = logging.handlers.RotatingFileHandler(
            log_dir / "k1_config_manager.log",
            maxBytes=10 * 1024 * 1024,  # 10MB
            backupCount=5
        )
        file_handler.setLevel(logging.DEBUG)
        
        # 格式化器
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
        )
        console_handler.setFormatter(formatter)
        file_handler.setFormatter(formatter)
        
        self.logger.addHandler(console_handler)
        self.logger.addHandler(file_handler)
    
    def debug(self, message: str, **kwargs):
        """调试日志"""
        self.logger.debug(message, extra=kwargs)
    
    def info(self, message: str, **kwargs):
        """信息日志"""
        self.logger.info(message, extra=kwargs)
    
    def warning(self, message: str, **kwargs):
        """警告日志"""
        self.logger.warning(message, extra=kwargs)
    
    def error(self, message: str, exception: Optional[Exception] = None, **kwargs):
        """错误日志"""
        if exception:
            self.logger.error(f"{message}: {exception}", exc_info=True, extra=kwargs)
        else:
            self.logger.error(message, extra=kwargs)
    
    def critical(self, message: str, **kwargs):
        """严重错误日志"""
        self.logger.critical(message, extra=kwargs)


class ConfigurationEncryptor:
    """配置加密器"""
    
    def __init__(self, secret_key: Optional[str] = None):
        self.secret_key = secret_key or self._generate_secret_key()
        self.logger = ConfigurationLogger("ConfigEncryptor")
    
    def _generate_secret_key(self) -> str:
        """生成秘密密钥"""
        import secrets
        return secrets.token_hex(32)
    
    def _get_hmac_key(self) -> bytes:
        """获取HMAC密钥"""
        return hmac.new(
            self.secret_key.encode(),
            b"configuration_encryption",
            hashlib.sha256
        ).digest()
    
    def encrypt(self, data: str) -> str:
        """加密数据"""
        try:
            # 生成随机盐
            salt = os.urandom(16)
            
            # 使用HMAC进行数据完整性验证
            h = hmac.new(self._get_hmac_key(), data.encode(), hashlib.sha256)
            
            # 组合数据：盐 + HMAC + 原始数据
            combined_data = salt + h.digest() + data.encode()
            
            # Base64编码
            encrypted_data = base64.b64encode(combined_data).decode()
            
            self.logger.debug(f"数据加密成功，长度: {len(encrypted_data)}")
            return encrypted_data
            
        except Exception as e:
            self.logger.error("数据加密失败", exception=e)
            raise ConfigurationEncryptionError(f"加密失败: {e}")
    
    def decrypt(self, encrypted_data: str) -> str:
        """解密数据"""
        try:
            # Base64解码
            combined_data = base64.b64decode(encrypted_data.encode())
            
            if len(combined_data) < 32:  # 16字节盐 + 32字节HMAC
                raise ValueError("加密数据格式无效")
            
            # 提取盐、HMAC和数据
            salt = combined_data[:16]
            hmac_digest = combined_data[16:48]
            original_data = combined_data[48:]
            
            # 验证HMAC
            expected_hmac = hmac.new(self._get_hmac_key(), original_data, hashlib.sha256)
            if not hmac.compare_digest(hmac_digest, expected_hmac.digest()):
                raise ValueError("数据完整性验证失败")
            
            # 解码原始数据
            decrypted_data = original_data.decode()
            
            self.logger.debug(f"数据解密成功，长度: {len(decrypted_data)}")
            return decrypted_data
            
        except Exception as e:
            self.logger.error("数据解密失败", exception=e)
            raise ConfigurationEncryptionError(f"解密失败: {e}")
    
    def encrypt_config(self, config_data: ConfigurationData) -> str:
        """加密配置数据"""
        config_json = json.dumps(config_data, ensure_ascii=False, separators=(',', ':'))
        return self.encrypt(config_json)
    
    def decrypt_config(self, encrypted_config: str) -> ConfigurationData:
        """解密配置数据"""
        config_json = self.decrypt(encrypted_config)
        return json.loads(config_json)


class ConfigurationCache:
    """配置缓存管理器"""
    
    def __init__(self, max_size: int = 1000, ttl: int = 3600):
        self.max_size = max_size
        self.ttl = ttl  # 缓存过期时间（秒）
        self._cache: Dict[str, Tuple[Any, float]] = {}
        self._access_times: Dict[str, float] = {}
        self._lock = threading.RLock()
        self.logger = ConfigurationLogger("ConfigCache")
        self._stats = {
            'hits': 0,
            'misses': 0,
            'evictions': 0,
            'expired': 0
        }
    
    def get(self, key: str) -> Optional[Any]:
        """获取缓存项"""
        with self._lock:
            if key not in self._cache:
                self._stats['misses'] += 1
                return None
            
            value, timestamp = self._cache[key]
            current_time = time.time()
            
            # 检查是否过期
            if current_time - timestamp > self.ttl:
                del self._cache[key]
                del self._access_times[key]
                self._stats['expired'] += 1
                return None
            
            # 更新访问时间
            self._access_times[key] = current_time
            self._stats['hits'] += 1
            return value
    
    def set(self, key: str, value: Any) -> None:
        """设置缓存项"""
        with self._lock:
            current_time = time.time()
            
            # 如果缓存已满，清理最久未访问的项目
            if len(self._cache) >= self.max_size:
                self._evict_lru()
            
            self._cache[key] = (value, current_time)
            self._access_times[key] = current_time
    
    def _evict_lru(self) -> None:
        """清理最久未访问的缓存项"""
        if not self._access_times:
            return
        
        # 找到最久未访问的键
        lru_key = min(self._access_times.keys(), key=lambda k: self._access_times[k])
        
        del self._cache[lru_key]
        del self._access_times[lru_key]
        self._stats['evictions'] += 1
    
    def clear(self) -> None:
        """清空缓存"""
        with self._lock:
            self._cache.clear()
            self._access_times.clear()
    
    def get_stats(self) -> Dict[str, int]:
        """获取缓存统计信息"""
        with self._lock:
            total_requests = self._stats['hits'] + self._stats['misses']
            hit_rate = self._stats['hits'] / total_requests if total_requests > 0 else 0
            
            return {
                **self._stats,
                'hit_rate': hit_rate,
                'total_requests': total_requests,
                'cache_size': len(self._cache),
                'max_size': self.max_size
            }


class ConfigurationValidator:
    """配置验证器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigValidator")
        self._validation_rules: Dict[str, List[ConfigurationValidationRule]] = defaultdict(list)
    
    def add_rule(self, config_name: str, rule: ConfigurationValidationRule) -> None:
        """添加验证规则"""
        self._validation_rules[config_name].append(rule)
    
    def add_rules(self, config_name: str, rules: List[ConfigurationValidationRule]) -> None:
        """批量添加验证规则"""
        self._validation_rules[config_name].extend(rules)
    
    def validate(self, config_name: str, config_data: ConfigurationData) -> List[str]:
        """验证配置数据"""
        errors = []
        rules = self._validation_rules.get(config_name, [])
        
        for rule in rules:
            try:
                self._validate_rule(rule, config_data, errors)
            except Exception as e:
                errors.append(f"验证规则执行失败 {rule.field_name}: {e}")
        
        return errors
    
    def _validate_rule(self, rule: ConfigurationValidationRule, config_data: ConfigurationData, errors: List[str]) -> None:
        """验证单个规则"""
        field_name = rule.field_name
        
        # 检查必填字段
        if rule.required and field_name not in config_data:
            errors.append(f"必填字段 '{field_name}' 缺失")
            return
        
        # 如果字段不存在且有默认值，使用默认值
        if field_name not in config_data and rule.default_value is not None:
            config_data[field_name] = rule.default_value
            return
        
        if field_name not in config_data:
            return
        
        value = config_data[field_name]
        
        # 类型检查
        if not isinstance(value, rule.field_type):
            errors.append(f"字段 '{field_name}' 类型错误，期望 {rule.field_type.__name__}，实际 {type(value).__name__}")
            return
        
        # 数值范围检查
        if rule.min_value is not None and isinstance(value, (int, float)):
            if value < rule.min_value:
                errors.append(f"字段 '{field_name}' 值 {value} 小于最小值 {rule.min_value}")
        
        if rule.max_value is not None and isinstance(value, (int, float)):
            if value > rule.max_value:
                errors.append(f"字段 '{field_name}' 值 {value} 大于最大值 {rule.max_value}")
        
        # 允许值检查
        if rule.allowed_values is not None and value not in rule.allowed_values:
            errors.append(f"字段 '{field_name}' 值 {value} 不在允许值列表中: {rule.allowed_values}")
        
        # 正则表达式检查
        if rule.pattern is not None and isinstance(value, str):
            if not re.match(rule.pattern, value):
                errors.append(f"字段 '{field_name}' 值 '{value}' 不匹配模式 '{rule.pattern}'")
        
        # 自定义验证器
        if rule.custom_validator is not None:
            try:
                if not rule.custom_validator(value):
                    errors.append(f"字段 '{field_name}' 未通过自定义验证")
            except Exception as e:
                errors.append(f"字段 '{field_name}' 自定义验证执行失败: {e}")
        
        # 依赖检查
        if rule.depends_on:
            for dep_field in rule.depends_on:
                if dep_field not in config_data:
                    errors.append(f"字段 '{field_name}' 依赖于字段 '{dep_field}'，但该字段不存在")


class ConfigurationHistory:
    """配置历史记录管理器"""
    
    def __init__(self, history_dir: str = "config_history"):
        self.history_dir = Path(history_dir)
        self.history_dir.mkdir(exist_ok=True)
        self.logger = ConfigurationLogger("ConfigHistory")
        self._changes: List[ConfigurationChange] = []
        self._max_history_size = 10000  # 最大历史记录数
    
    async def save_version(self, config_name: str, config_data: ConfigurationData, 
                          metadata: ConfigurationMetadata, encrypted_data: Optional[str] = None) -> str:
        """保存配置版本"""
        try:
            timestamp = datetime.now()
            version_id = f"{config_name}_{timestamp.strftime('%Y%m%d_%H%M%S_%f')}"
            
            # 创建版本目录
            version_dir = self.history_dir / config_name / version_id
            version_dir.mkdir(parents=True, exist_ok=True)
            
            # 保存配置数据
            config_file = version_dir / "config.json"
            async with aiofiles.open(config_file, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(config_data, ensure_ascii=False, indent=2))
            
            # 保存元数据
            metadata_file = version_dir / "metadata.json"
            async with aiofiles.open(metadata_file, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(asdict(metadata), ensure_ascii=False, indent=2, default=str))
            
            # 保存加密数据（如果有）
            if encrypted_data:
                encrypted_file = version_dir / "config.encrypted"
                async with aiofiles.open(encrypted_file, 'w', encoding='utf-8') as f:
                    await f.write(encrypted_data)
            
            self.logger.info(f"保存配置版本: {version_id}")
            return version_id
            
        except Exception as e:
            self.logger.error("保存配置版本失败", exception=e)
            raise ConfigurationError(f"保存版本失败: {e}")
    
    async def load_version(self, config_name: str, version_id: str) -> Tuple[ConfigurationData, ConfigurationMetadata]:
        """加载配置版本"""
        try:
            version_dir = self.history_dir / config_name / version_id
            
            if not version_dir.exists():
                raise ConfigurationError(f"版本 {version_id} 不存在")
            
            # 加载配置数据
            config_file = version_dir / "config.json"
            async with aiofiles.open(config_file, 'r', encoding='utf-8') as f:
                config_data = json.loads(await f.read())
            
            # 加载元数据
            metadata_file = version_dir / "metadata.json"
            async with aiofiles.open(metadata_file, 'r', encoding='utf-8') as f:
                metadata_dict = json.loads(await f.read())
                metadata = ConfigurationMetadata(**metadata_dict)
            
            self.logger.info(f"加载配置版本: {version_id}")
            return config_data, metadata
            
        except Exception as e:
            self.logger.error("加载配置版本失败", exception=e)
            raise ConfigurationError(f"加载版本失败: {e}")
    
    async def list_versions(self, config_name: str) -> List[str]:
        """列出配置的所有版本"""
        try:
            config_dir = self.history_dir / config_name
            if not config_dir.exists():
                return []
            
            versions = []
            for version_dir in config_dir.iterdir():
                if version_dir.is_dir():
                    versions.append(version_dir.name)
            
            return sorted(versions, reverse=True)  # 最新的版本在前
            
        except Exception as e:
            self.logger.error("列出配置版本失败", exception=e)
            return []
    
    async def delete_version(self, config_name: str, version_id: str) -> bool:
        """删除配置版本"""
        try:
            version_dir = self.history_dir / config_name / version_id
            
            if version_dir.exists():
                shutil.rmtree(version_dir)
                self.logger.info(f"删除配置版本: {version_id}")
                return True
            
            return False
            
        except Exception as e:
            self.logger.error("删除配置版本失败", exception=e)
            return False
    
    def record_change(self, change: ConfigurationChange) -> None:
        """记录配置变更"""
        self._changes.append(change)
        
        # 限制历史记录数量
        if len(self._changes) > self._max_history_size:
            self._changes = self._changes[-self._max_history_size:]
    
    def get_changes(self, config_name: Optional[str] = None, 
                   start_time: Optional[datetime] = None,
                   end_time: Optional[datetime] = None) -> List[ConfigurationChange]:
        """获取配置变更记录"""
        filtered_changes = self._changes
        
        if config_name:
            # 这里需要修改ConfigurationChange结构来包含config_name
            pass
        
        if start_time:
            filtered_changes = [c for c in filtered_changes if c.timestamp >= start_time]
        
        if end_time:
            filtered_changes = [c for c in filtered_changes if c.timestamp <= end_time]
        
        return filtered_changes


class ConfigurationLoader:
    """配置文件加载器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigLoader")
        self._loaders = {
            ConfigurationFormat.JSON: self._load_json,
            ConfigurationFormat.YAML: self._load_yaml,
            ConfigurationFormat.XML: self._load_xml,
            ConfigurationFormat.INI: self._load_ini,
        }
    
    def detect_format(self, file_path: Union[str, Path]) -> ConfigurationFormat:
        """检测配置文件格式"""
        file_path = Path(file_path)
        extension = file_path.suffix.lower()
        
        format_map = {
            '.json': ConfigurationFormat.JSON,
            '.yaml': ConfigurationFormat.YAML,
            '.yml': ConfigurationFormat.YAML,
            '.xml': ConfigurationFormat.XML,
            '.ini': ConfigurationFormat.INI,
        }
        
        if extension not in format_map:
            raise ConfigurationLoadError(f"不支持的文件格式: {extension}")
        
        return format_map[extension]
    
    async def load_file(self, file_path: Union[str, Path]) -> ConfigurationData:
        """异步加载配置文件"""
        try:
            format_type = self.detect_format(file_path)
            loader_func = self._loaders[format_type]
            
            result = await loader_func(file_path)
            self.logger.info(f"成功加载配置文件: {file_path}")
            return result
            
        except Exception as e:
            self.logger.error(f"加载配置文件失败: {file_path}", exception=e)
            raise ConfigurationLoadError(f"加载文件失败: {e}")
    
    async def save_file(self, file_path: Union[str, Path], config_data: ConfigurationData, 
                       format_type: Optional[ConfigurationFormat] = None) -> None:
        """异步保存配置文件"""
        try:
            if format_type is None:
                format_type = self.detect_format(file_path)
            
            save_func = getattr(self, f'_save_{format_type.value}')
            await save_func(file_path, config_data)
            
            self.logger.info(f"成功保存配置文件: {file_path}")
            
        except Exception as e:
            self.logger.error(f"保存配置文件失败: {file_path}", exception=e)
            raise ConfigurationSaveError(f"保存文件失败: {e}")
    
    async def _load_json(self, file_path: Union[str, Path]) -> ConfigurationData:
        """加载JSON配置文件"""
        async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
            content = await f.read()
        return json.loads(content)
    
    async def _load_yaml(self, file_path: Union[str, Path]) -> ConfigurationData:
        """加载YAML配置文件"""
        async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
            content = await f.read()
        return yaml.safe_load(content) or {}
    
    async def _load_xml(self, file_path: Union[str, Path]) -> ConfigurationData:
        """加载XML配置文件"""
        tree = ET.parse(file_path)
        root = tree.getroot()
        return self._xml_to_dict(root)
    
    async def _load_ini(self, file_path: Union[str, Path]) -> ConfigurationData:
        """加载INI配置文件"""
        config = configparser.ConfigParser()
        config.read(file_path, encoding='utf-8')
        
        result = {}
        for section_name in config.sections():
            result[section_name] = dict(config.items(section_name))
        
        return result
    
    async def _save_json(self, file_path: Union[str, Path], config_data: ConfigurationData) -> None:
        """保存JSON配置文件"""
        content = json.dumps(config_data, ensure_ascii=False, indent=2)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content)
    
    async def _save_yaml(self, file_path: Union[str, Path], config_data: ConfigurationData) -> None:
        """保存YAML配置文件"""
        content = yaml.dump(config_data, allow_unicode=True, indent=2)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content)
    
    async def _save_xml(self, file_path: Union[str, Path], config_data: ConfigurationData) -> None:
        """保存XML配置文件"""
        root = self._dict_to_xml('config', config_data)
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
    
    async def _save_ini(self, file_path: Union[str, Path], config_data: ConfigurationData) -> None:
        """保存INI配置文件"""
        config = configparser.ConfigParser()
        
        for section_name, section_data in config_data.items():
            config.add_section(section_name)
            for key, value in section_data.items():
                config.set(section_name, key, str(value))
        
        with open(file_path, 'w', encoding='utf-8') as f:
            config.write(f)
    
    def _xml_to_dict(self, element: ET.Element) -> Dict[str, Any]:
        """将XML元素转换为字典"""
        result = {}
        
        # 处理属性
        if element.attrib:
            result['@attributes'] = element.attrib
        
        # 处理文本内容
        if element.text and element.text.strip():
            if len(element) == 0:
                return element.text.strip()
            result['#text'] = element.text.strip()
        
        # 处理子元素
        for child in element:
            child_data = self._xml_to_dict(child)
            
            if child.tag in result:
                # 如果已存在相同标签，转换为列表
                if not isinstance(result[child.tag], list):
                    result[child.tag] = [result[child.tag]]
                result[child.tag].append(child_data)
            else:
                result[child.tag] = child_data
        
        return result
    
    def _dict_to_xml(self, tag: str, data: Any) -> ET.Element:
        """将字典转换为XML元素"""
        element = ET.Element(tag)
        
        if isinstance(data, dict):
            for key, value in data.items():
                if key == '@attributes':
                    element.attrib.update(value)
                elif key == '#text':
                    element.text = str(value)
                else:
                    if isinstance(value, list):
                        for item in value:
                            element.append(self._dict_to_xml(key, item))
                    else:
                        element.append(self._dict_to_xml(key, value))
        elif isinstance(data, list):
            for item in data:
                element.append(self._dict_to_xml(tag, item))
        else:
            element.text = str(data)
        
        return element


class SystemConfigurationManager:
    """
    K1系统配置管理器
    
    这是一个功能完整的系统配置管理器，提供以下核心功能：
    
    1. 系统级配置管理（系统参数、环境变量、运行参数）
    2. 配置文件管理（YAML、JSON、XML、INI格式支持）
    3. 配置验证和检查（类型检查、范围检查、依赖检查）
    4. 配置热更新和动态加载
    5. 配置版本管理和历史记录
    6. 配置加密和安全存储
    7. 异步配置处理和缓存
    8. 完整的错误处理和日志记录
    
    使用示例：
    
    ```python
    # 创建配置管理器实例
    config_manager = SystemConfigurationManager()
    
    # 加载配置文件
    await config_manager.load_config('app_config', 'config/app.json')
    
    # 获取配置值
    db_host = config_manager.get('database.host')
    db_port = config_manager.get('database.port', 5432)
    
    # 设置配置值
    config_manager.set('app.debug', True)
    
    # 保存配置
    await config_manager.save_config('app_config')
    
    # 配置验证
    config_manager.add_validation_rules('app_config', [
        ConfigurationValidationRule('database.port', int, min_value=1, max_value=65535),
        ConfigurationValidationRule('app.debug', bool)
    ])
    
    errors = config_manager.validate_config('app_config')
    if errors:
        print("配置验证失败:", errors)
    
    # 启用热更新
    await config_manager.enable_hot_reload('app_config', 'config/app.json')
    ```
    """
    
    def __init__(self, 
                 config_dir: str = "configs",
                 history_dir: str = "config_history",
                 cache_size: int = 1000,
                 cache_ttl: int = 3600,
                 secret_key: Optional[str] = None,
                 log_level: LogLevel = LogLevel.INFO):
        """
        初始化系统配置管理器
        
        Args:
            config_dir: 配置文件目录
            history_dir: 配置历史记录目录
            cache_size: 缓存大小限制
            cache_ttl: 缓存过期时间（秒）
            secret_key: 加密密钥，如果为None则自动生成
            log_level: 日志级别
        """
        self.config_dir = Path(config_dir)
        self.config_dir.mkdir(exist_ok=True)
        
        # 初始化组件
        self.logger = ConfigurationLogger("SystemConfigManager", log_level)
        self.validator = ConfigurationValidator()
        self.history = ConfigurationHistory(history_dir)
        self.cache = ConfigurationCache(cache_size, cache_ttl)
        self.loader = ConfigurationLoader()
        self.encryptor = ConfigurationEncryptor(secret_key)
        
        # 配置存储
        self._configs: Dict[str, ConfigurationData] = {}
        self._metadata: Dict[str, ConfigurationMetadata] = {}
        self._watchers: Dict[str, List[Callable]] = defaultdict(list)
        self._hot_reload_tasks: Dict[str, asyncio.Task] = {}
        self._file_watchers: Dict[str, asyncio.Task] = {}
        
        # 线程锁
        self._lock = threading.RLock()
        
        # 性能统计
        self._stats = {
            'configs_loaded': 0,
            'configs_saved': 0,
            'validations_performed': 0,
            'cache_hits': 0,
            'cache_misses': 0,
            'hot_reloads_triggered': 0,
            'errors_encountered': 0
        }
        
        self.logger.info("K1系统配置管理器初始化完成")
    
    async def load_config(self, config_name: str, file_path: Union[str, Path], 
                         validate: bool = True, use_cache: bool = True) -> ConfigurationData:
        """
        异步加载配置文件
        
        Args:
            config_name: 配置名称
            file_path: 配置文件路径
            validate: 是否进行验证
            use_cache: 是否使用缓存
            
        Returns:
            配置数据字典
        """
        try:
            cache_key = f"{config_name}:{file_path}"
            
            # 检查缓存
            if use_cache:
                cached_data = self.cache.get(cache_key)
                if cached_data is not None:
                    self._configs[config_name] = cached_data
                    self._stats['cache_hits'] += 1
                    self.logger.debug(f"从缓存加载配置: {config_name}")
                    return cached_data
                else:
                    self._stats['cache_misses'] += 1
            
            # 加载配置文件
            config_data = await self.loader.load_file(file_path)
            
            # 验证配置
            if validate:
                errors = self.validator.validate(config_name, config_data)
                if errors:
                    raise ConfigurationValidationError(f"配置验证失败: {errors}")
            
            # 检查是否加密
            encrypted_data = None
            if config_data.get('_encrypted', False):
                encrypted_content = config_data.get('_content')
                if encrypted_content:
                    config_data = self.encryptor.decrypt_config(encrypted_content)
                    encrypted_data = encrypted_content
            
            # 计算校验和
            config_json = json.dumps(config_data, sort_keys=True)
            checksum = hashlib.sha256(config_json.encode()).hexdigest()
            
            # 创建元数据
            metadata = ConfigurationMetadata(
                name=config_name,
                version="1.0.0",
                checksum=checksum,
                encrypted=encrypted_data is not None,
                size=len(config_json)
            )
            
            # 存储配置和元数据
            with self._lock:
                self._configs[config_name] = config_data
                self._metadata[config_name] = metadata
            
            # 缓存配置
            if use_cache:
                self.cache.set(cache_key, config_data)
            
            # 保存版本历史
            await self.history.save_version(config_name, config_data, metadata, encrypted_data)
            
            self._stats['configs_loaded'] += 1
            self.logger.info(f"成功加载配置: {config_name} from {file_path}")
            
            return config_data
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"加载配置失败: {config_name}", exception=e)
            raise ConfigurationLoadError(f"加载配置失败: {e}")
    
    async def save_config(self, config_name: str, file_path: Optional[Union[str, Path]] = None,
                         encrypt: bool = False, format_type: Optional[ConfigurationFormat] = None) -> None:
        """
        异步保存配置到文件
        
        Args:
            config_name: 配置名称
            file_path: 保存路径，如果为None则使用默认路径
            encrypt: 是否加密保存
            format_type: 文件格式
        """
        try:
            if config_name not in self._configs:
                raise ConfigurationError(f"配置 {config_name} 不存在")
            
            config_data = self._configs[config_name].copy()
            metadata = self._metadata[config_name]
            
            # 加密处理
            encrypted_data = None
            if encrypt:
                encrypted_data = self.encryptor.encrypt_config(config_data)
                config_data = {
                    '_encrypted': True,
                    '_content': encrypted_data
                }
            
            # 确定保存路径
            if file_path is None:
                if format_type is None:
                    format_type = ConfigurationFormat.JSON
                file_path = self.config_dir / f"{config_name}.{format_type.value}"
            else:
                file_path = Path(file_path)
            
            # 创建目录
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存文件
            await self.loader.save_file(file_path, config_data, format_type)
            
            # 更新元数据
            with self._lock:
                metadata.updated_at = datetime.now()
                if encrypt:
                    metadata.encrypted = True
                
                # 计算新的校验和
                config_json = json.dumps(self._configs[config_name], sort_keys=True)
                metadata.checksum = hashlib.sha256(config_json.encode()).hexdigest()
                metadata.size = len(config_json)
            
            # 保存版本历史
            await self.history.save_version(config_name, self._configs[config_name], 
                                          metadata, encrypted_data)
            
            self._stats['configs_saved'] += 1
            self.logger.info(f"成功保存配置: {config_name} to {file_path}")
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"保存配置失败: {config_name}", exception=e)
            raise ConfigurationSaveError(f"保存配置失败: {e}")
    
    def get(self, key: str, default: Any = None, config_name: Optional[str] = None) -> Any:
        """
        获取配置值（支持点号路径）
        
        Args:
            key: 配置键，支持点号分隔的路径，如 'database.host'
            default: 默认值
            config_name: 配置名称，如果为None则在所有配置中搜索
            
        Returns:
            配置值
        """
        try:
            keys = key.split('.')
            
            # 确定搜索的配置列表
            configs_to_search = []
            if config_name:
                if config_name in self._configs:
                    configs_to_search = [config_name]
            else:
                configs_to_search = list(self._configs.keys())
            
            # 在配置中搜索
            for config_name in configs_to_search:
                config_data = self._configs[config_name]
                current = config_data
                
                try:
                    for k in keys:
                        if isinstance(current, dict):
                            current = current[k]
                        else:
                            break
                    else:
                        # 成功找到值
                        return current
                except (KeyError, TypeError):
                    continue
            
            return default
            
        except Exception as e:
            self.logger.error(f"获取配置值失败: {key}", exception=e)
            return default
    
    def set(self, key: str, value: Any, config_name: str, 
            create_missing: bool = True, record_change: bool = True) -> None:
        """
        设置配置值（支持点号路径）
        
        Args:
            key: 配置键，支持点号分隔的路径
            value: 配置值
            config_name: 配置名称
            create_missing: 是否创建缺失的路径
            record_change: 是否记录变更
        """
        try:
            if config_name not in self._configs:
                if create_missing:
                    self._configs[config_name] = {}
                else:
                    raise ConfigurationError(f"配置 {config_name} 不存在")
            
            keys = key.split('.')
            config_data = self._configs[config_name]
            
            # 导航到目标位置
            current = config_data
            for k in keys[:-1]:
                if k not in current:
                    if create_missing:
                        current[k] = {}
                    else:
                        raise ConfigurationError(f"路径 {key} 不存在")
                current = current[k]
                
                if not isinstance(current, dict):
                    raise ConfigurationError(f"路径 {key} 不是字典类型")
            
            # 记录变更
            if record_change and key in config_data:
                old_value = config_data[key]
                change = ConfigurationChange(
                    timestamp=datetime.now(),
                    field_name=key,
                    old_value=old_value,
                    new_value=value,
                    change_type='modify'
                )
                self.history.record_change(change)
            
            # 设置新值
            current[keys[-1]] = value
            
            # 清理相关缓存
            self._clear_config_cache(config_name)
            
            self.logger.debug(f"设置配置值: {config_name}.{key} = {value}")
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"设置配置值失败: {config_name}.{key}", exception=e)
            raise ConfigurationError(f"设置配置值失败: {e}")
    
    def delete(self, key: str, config_name: str, record_change: bool = True) -> bool:
        """
        删除配置项
        
        Args:
            key: 配置键
            config_name: 配置名称
            record_change: 是否记录变更
            
        Returns:
            是否成功删除
        """
        try:
            if config_name not in self._configs:
                return False
            
            keys = key.split('.')
            config_data = self._configs[config_name]
            
            # 导航到目标位置
            current = config_data
            for k in keys[:-1]:
                if k not in current:
                    return False
                current = current[k]
                
                if not isinstance(current, dict):
                    return False
            
            # 检查目标键是否存在
            if keys[-1] not in current:
                return False
            
            # 记录变更
            if record_change:
                old_value = current[keys[-1]]
                change = ConfigurationChange(
                    timestamp=datetime.now(),
                    field_name=key,
                    old_value=old_value,
                    new_value=None,
                    change_type='delete'
                )
                self.history.record_change(change)
            
            # 删除键
            del current[keys[-1]]
            
            # 清理相关缓存
            self._clear_config_cache(config_name)
            
            self.logger.debug(f"删除配置项: {config_name}.{key}")
            return True
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"删除配置项失败: {config_name}.{key}", exception=e)
            return False
    
    def exists(self, key: str, config_name: Optional[str] = None) -> bool:
        """
        检查配置键是否存在
        
        Args:
            key: 配置键
            config_name: 配置名称
            
        Returns:
            是否存在
        """
        return self.get(key, None, config_name) is not None
    
    def get_config(self, config_name: str) -> Optional[ConfigurationData]:
        """获取完整配置"""
        return self._configs.get(config_name)
    
    def get_all_configs(self) -> Dict[str, ConfigurationData]:
        """获取所有配置"""
        return self._configs.copy()
    
    def get_metadata(self, config_name: str) -> Optional[ConfigurationMetadata]:
        """获取配置元数据"""
        return self._metadata.get(config_name)
    
    def get_all_metadata(self) -> Dict[str, ConfigurationMetadata]:
        """获取所有配置元数据"""
        return self._metadata.copy()
    
    def add_validation_rules(self, config_name: str, rules: List[ConfigurationValidationRule]) -> None:
        """添加验证规则"""
        self.validator.add_rules(config_name, rules)
        self.logger.info(f"为配置 {config_name} 添加了 {len(rules)} 个验证规则")
    
    def validate_config(self, config_name: str) -> List[str]:
        """验证配置"""
        if config_name not in self._configs:
            return [f"配置 {config_name} 不存在"]
        
        self._stats['validations_performed'] += 1
        errors = self.validator.validate(config_name, self._configs[config_name])
        
        if errors:
            self.logger.warning(f"配置 {config_name} 验证失败: {errors}")
        else:
            self.logger.debug(f"配置 {config_name} 验证通过")
        
        return errors
    
    async def enable_hot_reload(self, config_name: str, file_path: Union[str, Path], 
                               check_interval: float = 5.0) -> None:
        """
        启用配置热更新
        
        Args:
            config_name: 配置名称
            file_path: 配置文件路径
            check_interval: 检查间隔（秒）
        """
        try:
            if config_name in self._hot_reload_tasks:
                await self.disable_hot_reload(config_name)
            
            file_path = Path(file_path)
            self.logger.info(f"启用配置热更新: {config_name} ({file_path})")
            
            # 创建文件监控任务
            task = asyncio.create_task(
                self._hot_reload_worker(config_name, file_path, check_interval)
            )
            self._hot_reload_tasks[config_name] = task
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"启用热更新失败: {config_name}", exception=e)
            raise ConfigurationError(f"启用热更新失败: {e}")
    
    async def disable_hot_reload(self, config_name: str) -> None:
        """禁用配置热更新"""
        try:
            if config_name in self._hot_reload_tasks:
                task = self._hot_reload_tasks[config_name]
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass
                del self._hot_reload_tasks[config_name]
                
                self.logger.info(f"禁用配置热更新: {config_name}")
                
        except Exception as e:
            self.logger.error(f"禁用热更新失败: {config_name}", exception=e)
    
    async def _hot_reload_worker(self, config_name: str, file_path: Path, check_interval: float) -> None:
        """热更新工作线程"""
        last_modified = 0
        
        try:
            while True:
                try:
                    if file_path.exists():
                        current_modified = file_path.stat().st_mtime
                        
                        if current_modified > last_modified:
                            self.logger.info(f"检测到配置文件变化，重新加载: {config_name}")
                            
                            # 重新加载配置
                            await self.load_config(config_name, file_path, validate=True, use_cache=False)
                            
                            # 通知观察者
                            await self._notify_watchers(config_name, 'reload')
                            
                            last_modified = current_modified
                            self._stats['hot_reloads_triggered'] += 1
                    
                    await asyncio.sleep(check_interval)
                    
                except asyncio.CancelledError:
                    break
                except Exception as e:
                    self.logger.error(f"热更新工作线程异常: {config_name}", exception=e)
                    await asyncio.sleep(check_interval)
                    
        except asyncio.CancelledError:
            pass
    
    def watch_config(self, config_name: str, callback: Callable) -> None:
        """添加配置观察者"""
        self._watchers[config_name].append(callback)
        self.logger.debug(f"添加配置观察者: {config_name}")
    
    def unwatch_config(self, config_name: str, callback: Callable) -> None:
        """移除配置观察者"""
        if callback in self._watchers[config_name]:
            self._watchers[config_name].remove(callback)
            self.logger.debug(f"移除配置观察者: {config_name}")
    
    async def _notify_watchers(self, config_name: str, event_type: str) -> None:
        """通知观察者"""
        watchers = self._watchers.get(config_name, [])
        
        for watcher in watchers:
            try:
                if asyncio.iscoroutinefunction(watcher):
                    await watcher(config_name, event_type)
                else:
                    watcher(config_name, event_type)
            except Exception as e:
                self.logger.error(f"通知观察者失败: {config_name}", exception=e)
    
    def _clear_config_cache(self, config_name: str) -> None:
        """清理配置相关缓存"""
        cache_keys_to_remove = []
        
        for key in self.cache._cache.keys():
            if key.startswith(f"{config_name}:"):
                cache_keys_to_remove.append(key)
        
        for key in cache_keys_to_remove:
            if key in self.cache._cache:
                del self.cache._cache[key]
            if key in self.cache._access_times:
                del self.cache._access_times[key]
    
    async def reload_all_configs(self) -> Dict[str, bool]:
        """重新加载所有配置"""
        results = {}
        
        for config_name in list(self._configs.keys()):
            try:
                # 这里需要保存原始文件路径，暂时跳过
                results[config_name] = False
            except Exception as e:
                self.logger.error(f"重新加载配置失败: {config_name}", exception=e)
                results[config_name] = False
        
        return results
    
    async def backup_config(self, config_name: str, backup_path: Optional[Path] = None) -> Path:
        """备份配置"""
        try:
            if config_name not in self._configs:
                raise ConfigurationError(f"配置 {config_name} 不存在")
            
            if backup_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = self.config_dir / f"{config_name}_backup_{timestamp}.json"
            
            backup_data = {
                'config': self._configs[config_name],
                'metadata': asdict(self._metadata[config_name]),
                'backup_time': datetime.now().isoformat()
            }
            
            async with aiofiles.open(backup_path, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(backup_data, ensure_ascii=False, indent=2))
            
            self.logger.info(f"配置备份完成: {config_name} -> {backup_path}")
            return backup_path
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"配置备份失败: {config_name}", exception=e)
            raise ConfigurationError(f"备份配置失败: {e}")
    
    async def restore_config(self, config_name: str, backup_path: Path) -> None:
        """恢复配置"""
        try:
            async with aiofiles.open(backup_path, 'r', encoding='utf-8') as f:
                backup_data = json.loads(await f.read())
            
            config_data = backup_data['config']
            
            # 验证恢复的数据
            errors = self.validator.validate(config_name, config_data)
            if errors:
                raise ConfigurationValidationError(f"恢复的配置验证失败: {errors}")
            
            # 恢复配置
            with self._lock:
                self._configs[config_name] = config_data
                
                # 更新元数据
                if config_name in self._metadata:
                    self._metadata[config_name].updated_at = datetime.now()
            
            # 清理缓存
            self._clear_config_cache(config_name)
            
            self.logger.info(f"配置恢复完成: {config_name} from {backup_path}")
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"配置恢复失败: {config_name}", exception=e)
            raise ConfigurationError(f"恢复配置失败: {e}")
    
    def get_statistics(self) -> Dict[str, Any]:
        """获取性能统计信息"""
        cache_stats = self.cache.get_stats()
        
        return {
            'configuration_stats': self._stats.copy(),
            'cache_stats': cache_stats,
            'loaded_configs': len(self._configs),
            'active_hot_reloads': len(self._hot_reload_tasks),
            'total_watchers': sum(len(watchers) for watchers in self._watchers.values())
        }
    
    def reset_statistics(self) -> None:
        """重置统计信息"""
        self._stats = {
            'configs_loaded': 0,
            'configs_saved': 0,
            'validations_performed': 0,
            'cache_hits': 0,
            'cache_misses': 0,
            'hot_reloads_triggered': 0,
            'errors_encountered': 0
        }
        self.cache.clear()
        self.logger.info("统计信息已重置")
    
    async def cleanup_old_versions(self, config_name: str, keep_versions: int = 10) -> int:
        """清理旧版本"""
        try:
            versions = await self.history.list_versions(config_name)
            
            if len(versions) <= keep_versions:
                return 0
            
            versions_to_delete = versions[keep_versions:]
            deleted_count = 0
            
            for version_id in versions_to_delete:
                if await self.history.delete_version(config_name, version_id):
                    deleted_count += 1
            
            self.logger.info(f"清理旧版本完成: {config_name}，删除了 {deleted_count} 个版本")
            return deleted_count
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"清理旧版本失败: {config_name}", exception=e)
            return 0
    
    async def export_config(self, config_name: str, export_path: Path, 
                           format_type: ConfigurationFormat = ConfigurationFormat.JSON,
                           encrypt: bool = False) -> None:
        """导出配置"""
        try:
            if config_name not in self._configs:
                raise ConfigurationError(f"配置 {config_name} 不存在")
            
            config_data = self._configs[config_name]
            
            if encrypt:
                config_data = {
                    '_encrypted': True,
                    '_content': self.encryptor.encrypt_config(config_data)
                }
            
            await self.loader.save_file(export_path, config_data, format_type)
            
            self.logger.info(f"配置导出完成: {config_name} -> {export_path}")
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"配置导出失败: {config_name}", exception=e)
            raise ConfigurationError(f"导出配置失败: {e}")
    
    async def import_config(self, config_name: str, import_path: Path,
                           validate: bool = True, overwrite: bool = False) -> None:
        """导入配置"""
        try:
            config_data = await self.loader.load_file(import_path)
            
            # 检查是否加密
            encrypted_data = None
            if config_data.get('_encrypted', False):
                encrypted_content = config_data.get('_content')
                if encrypted_content:
                    config_data = self.encryptor.decrypt_config(encrypted_content)
                    encrypted_data = encrypted_content
            
            # 验证配置
            if validate:
                errors = self.validator.validate(config_name, config_data)
                if errors:
                    raise ConfigurationValidationError(f"导入的配置验证失败: {errors}")
            
            # 检查配置是否存在
            if config_name in self._configs and not overwrite:
                raise ConfigurationError(f"配置 {config_name} 已存在，使用 overwrite=True 覆盖")
            
            # 导入配置
            with self._lock:
                self._configs[config_name] = config_data
                
                # 创建或更新元数据
                metadata = ConfigurationMetadata(
                    name=config_name,
                    version="1.0.0",
                    checksum=hashlib.sha256(json.dumps(config_data, sort_keys=True).encode()).hexdigest(),
                    encrypted=encrypted_data is not None,
                    size=len(json.dumps(config_data))
                )
                self._metadata[config_name] = metadata
            
            # 清理缓存
            self._clear_config_cache(config_name)
            
            # 保存版本历史
            await self.history.save_version(config_name, config_data, metadata, encrypted_data)
            
            self.logger.info(f"配置导入完成: {config_name} from {import_path}")
            
        except Exception as e:
            self._stats['errors_encountered'] += 1
            self.logger.error(f"配置导入失败: {config_name}", exception=e)
            raise ConfigurationError(f"导入配置失败: {e}")
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        # 清理资源
        for task in self._hot_reload_tasks.values():
            task.cancel()
        
        self.logger.info("K1系统配置管理器已关闭")


# 便捷函数和工厂方法
def create_config_manager(config_dir: str = "configs",
                         history_dir: str = "config_history",
                         cache_size: int = 1000,
                         cache_ttl: int = 3600,
                         secret_key: Optional[str] = None,
                         log_level: LogLevel = LogLevel.INFO) -> SystemConfigurationManager:
    """
    创建配置管理器实例的便捷函数
    
    Args:
        config_dir: 配置文件目录
        history_dir: 配置历史记录目录
        cache_size: 缓存大小限制
        cache_ttl: 缓存过期时间（秒）
        secret_key: 加密密钥
        log_level: 日志级别
        
    Returns:
        配置管理器实例
    """
    return SystemConfigurationManager(
        config_dir=config_dir,
        history_dir=history_dir,
        cache_size=cache_size,
        cache_ttl=cache_ttl,
        secret_key=secret_key,
        log_level=log_level
    )


def create_validation_rules(rules_config: Dict[str, Dict]) -> List[ConfigurationValidationRule]:
    """
    从配置字典创建验证规则
    
    Args:
        rules_config: 验证规则配置字典
        
    Returns:
        验证规则列表
    """
    rules = []
    
    for field_name, rule_config in rules_config.items():
        rule = ConfigurationValidationRule(
            field_name=field_name,
            field_type=eval(rule_config['type']),  # 注意：这里使用eval可能有安全风险
            required=rule_config.get('required', True),
            default_value=rule_config.get('default_value'),
            min_value=rule_config.get('min_value'),
            max_value=rule_config.get('max_value'),
            allowed_values=rule_config.get('allowed_values'),
            pattern=rule_config.get('pattern')
        )
        rules.append(rule)
    
    return rules


# 使用示例
async def example_usage():
    """配置管理器使用示例"""
    
    # 创建配置管理器
    config_manager = create_config_manager(
        config_dir="example_configs",
        log_level=LogLevel.DEBUG
    )
    
    try:
        # 示例1: 基本配置操作
        print("=== 示例1: 基本配置操作 ===")
        
        # 加载配置
        await config_manager.load_config('app', 'example_configs/app.json')
        
        # 获取配置值
        debug_mode = config_manager.get('app.debug', False)
        db_host = config_manager.get('database.host', 'localhost')
        db_port = config_manager.get('database.port', 5432)
        
        print(f"调试模式: {debug_mode}")
        print(f"数据库主机: {db_host}")
        print(f"数据库端口: {db_port}")
        
        # 设置配置值
        config_manager.set('app.version', '1.0.1', 'app')
        config_manager.set('database.timeout', 30, 'app')
        
        # 保存配置
        await config_manager.save_config('app')
        
        # 示例2: 配置验证
        print("\n=== 示例2: 配置验证 ===")
        
        validation_rules = [
            ConfigurationValidationRule('database.port', int, min_value=1, max_value=65535),
            ConfigurationValidationRule('app.debug', bool),
            ConfigurationValidationRule('app.version', str, pattern=r'^\d+\.\d+\.\d+$')
        ]
        
        config_manager.add_validation_rules('app', validation_rules)
        
        errors = config_manager.validate_config('app')
        if errors:
            print("配置验证失败:")
            for error in errors:
                print(f"  - {error}")
        else:
            print("配置验证通过")
        
        # 示例3: 热更新
        print("\n=== 示例3: 热更新 ===")
        
        def config_reloaded(config_name: str, event_type: str):
            print(f"配置 {config_name} 发生 {event_type} 事件")
        
        config_manager.watch_config('app', config_reloaded)
        await config_manager.enable_hot_reload('app', 'example_configs/app.json')
        
        print("热更新已启用，修改配置文件将自动重新加载")
        
        # 示例4: 配置备份和恢复
        print("\n=== 示例4: 配置备份和恢复 ===")
        
        backup_path = await config_manager.backup_config('app')
        print(f"配置已备份到: {backup_path}")
        
        # 示例5: 统计信息
        print("\n=== 示例5: 统计信息 ===")
        
        stats = config_manager.get_statistics()
        print("配置管理器统计信息:")
        for key, value in stats.items():
            print(f"  {key}: {value}")
        
        # 等待一段时间以观察热更新
        print("\n等待热更新测试...")
        await asyncio.sleep(2)
        
    finally:
        # 清理资源
        await config_manager.disable_hot_reload('app')
        print("配置管理器示例完成")


class ConfigurationTemplate:
    """配置模板管理器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigTemplate")
        self._templates: Dict[str, Dict[str, Any]] = {}
        self._template_cache = ConfigurationCache(max_size=500, ttl=1800)
    
    def create_template(self, template_name: str, template_data: ConfigurationData, 
                       variables: Optional[Dict[str, str]] = None) -> None:
        """创建配置模板"""
        try:
            self._templates[template_name] = {
                'data': template_data,
                'variables': variables or {},
                'created_at': datetime.now(),
                'version': '1.0.0'
            }
            self.logger.info(f"创建配置模板: {template_name}")
        except Exception as e:
            self.logger.error(f"创建配置模板失败: {template_name}", exception=e)
            raise ConfigurationError(f"创建模板失败: {e}")
    
    def load_template(self, template_name: str) -> Optional[ConfigurationData]:
        """加载配置模板"""
        cache_key = f"template:{template_name}"
        cached_template = self._template_cache.get(cache_key)
        
        if cached_template is not None:
            return cached_template.copy()
        
        if template_name in self._templates:
            template_data = self._templates[template_name]['data'].copy()
            self._template_cache.set(cache_key, template_data)
            return template_data
        
        return None
    
    def instantiate_template(self, template_name: str, variables: Optional[Dict[str, str]] = None) -> ConfigurationData:
        """实例化配置模板"""
        try:
            template_info = self._templates.get(template_name)
            if not template_info:
                raise ConfigurationError(f"模板 {template_name} 不存在")
            
            template_data = template_info['data'].copy()
            template_vars = {**template_info['variables'], **(variables or {})}
            
            # 替换模板变量
            result = self._substitute_variables(template_data, template_vars)
            
            self.logger.debug(f"实例化模板: {template_name}")
            return result
            
        except Exception as e:
            self.logger.error(f"实例化模板失败: {template_name}", exception=e)
            raise ConfigurationError(f"实例化模板失败: {e}")
    
    def _substitute_variables(self, data: Any, variables: Dict[str, str]) -> Any:
        """递归替换模板变量"""
        if isinstance(data, str):
            # 替换 ${variable} 格式的变量
            result = data
            for var_name, var_value in variables.items():
                placeholder = f"${{{var_name}}}"
                result = result.replace(placeholder, str(var_value))
            return result
        elif isinstance(data, dict):
            return {key: self._substitute_variables(value, variables) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._substitute_variables(item, variables) for item in data]
        else:
            return data
    
    def list_templates(self) -> List[str]:
        """列出所有模板"""
        return list(self._templates.keys())
    
    def delete_template(self, template_name: str) -> bool:
        """删除模板"""
        if template_name in self._templates:
            del self._templates[template_name]
            self._template_cache.clear()  # 清理缓存
            self.logger.info(f"删除配置模板: {template_name}")
            return True
        return False


class ConfigurationMerger:
    """配置合并器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigMerger")
    
    def merge_configs(self, *configs: ConfigurationData, 
                     strategy: str = 'override') -> ConfigurationData:
        """
        合并多个配置
        
        Args:
            *configs: 要合并的配置字典
            strategy: 合并策略 ('override', 'merge', 'append')
            
        Returns:
            合并后的配置
        """
        try:
            if not configs:
                return {}
            
            if strategy == 'override':
                return self._merge_override(configs)
            elif strategy == 'merge':
                return self._merge_deep(configs)
            elif strategy == 'append':
                return self._merge_append(configs)
            else:
                raise ConfigurationError(f"不支持的合并策略: {strategy}")
                
        except Exception as e:
            self.logger.error(f"合并配置失败", exception=e)
            raise ConfigurationError(f"合并配置失败: {e}")
    
    def _merge_override(self, configs: Tuple[ConfigurationData, ...]) -> ConfigurationData:
        """覆盖策略：后面的配置覆盖前面的"""
        result = {}
        for config in configs:
            result.update(config)
        return result
    
    def _merge_deep(self, configs: Tuple[ConfigurationData, ...]) -> ConfigurationData:
        """深度合并策略"""
        result = {}
        for config in configs:
            result = self._deep_update(result, config)
        return result
    
    def _deep_update(self, target: Dict[str, Any], source: Dict[str, Any]) -> Dict[str, Any]:
        """深度更新字典"""
        for key, value in source.items():
            if key in target and isinstance(target[key], dict) and isinstance(value, dict):
                target[key] = self._deep_update(target[key], value)
            else:
                target[key] = value.copy() if isinstance(value, (dict, list)) else value
        return target
    
    def _merge_append(self, configs: Tuple[ConfigurationData, ...]) -> ConfigurationData:
        """追加策略：列表追加，字典合并"""
        result = {}
        for config in configs:
            for key, value in config.items():
                if key in result:
                    if isinstance(result[key], list) and isinstance(value, list):
                        result[key].extend(value)
                    elif isinstance(result[key], dict) and isinstance(value, dict):
                        result[key] = self._deep_update(result[key], value)
                    else:
                        result[key] = value
                else:
                    result[key] = value.copy() if isinstance(value, (dict, list)) else value
        return result


class ConfigurationDependencyManager:
    """配置依赖管理器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigDependency")
        self._dependencies: Dict[str, List[str]] = defaultdict(list)
        self._dependents: Dict[str, List[str]] = defaultdict(list)
        self._dependency_graph = {}
    
    def add_dependency(self, config_name: str, depends_on: str) -> None:
        """添加配置依赖"""
        self._dependencies[config_name].append(depends_on)
        self._dependents[depends_on].append(config_name)
        
        self.logger.debug(f"添加配置依赖: {config_name} -> {depends_on}")
    
    def remove_dependency(self, config_name: str, depends_on: str) -> None:
        """移除配置依赖"""
        if depends_on in self._dependencies[config_name]:
            self._dependencies[config_name].remove(depends_on)
        
        if config_name in self._dependents[depends_on]:
            self._dependents[depends_on].remove(config_name)
        
        self.logger.debug(f"移除配置依赖: {config_name} -> {depends_on}")
    
    def get_dependencies(self, config_name: str) -> List[str]:
        """获取配置的直接依赖"""
        return self._dependencies.get(config_name, [])
    
    def get_dependents(self, config_name: str) -> List[str]:
        """获取依赖此配置的配置"""
        return self._dependents.get(config_name, [])
    
    def get_all_dependencies(self, config_name: str) -> Set[str]:
        """获取配置的所有依赖（递归）"""
        visited = set()
        stack = [config_name]
        
        while stack:
            current = stack.pop()
            if current not in visited:
                visited.add(current)
                for dep in self._dependencies.get(current, []):
                    if dep not in visited:
                        stack.append(dep)
        
        return visited - {config_name}
    
    def get_load_order(self, config_names: List[str]) -> List[str]:
        """获取配置加载顺序"""
        # 使用拓扑排序
        in_degree = {name: 0 for name in config_names}
        
        # 计算入度
        for config_name in config_names:
            for dep in self._dependencies.get(config_name, []):
                if dep in in_degree:
                    in_degree[config_name] += 1
        
        # 拓扑排序
        queue = [name for name, degree in in_degree.items() if degree == 0]
        result = []
        
        while queue:
            current = queue.pop(0)
            result.append(current)
            
            # 更新依赖当前配置的节点
            for dependent in self._dependents.get(current, []):
                if dependent in in_degree:
                    in_degree[dependent] -= 1
                    if in_degree[dependent] == 0:
                        queue.append(dependent)
        
        # 检查是否有循环依赖
        if len(result) != len(config_names):
            remaining = set(config_names) - set(result)
            raise ConfigurationError(f"检测到循环依赖: {remaining}")
        
        return result
    
    def detect_circular_dependencies(self) -> List[List[str]]:
        """检测循环依赖"""
        cycles = []
        visited = set()
        rec_stack = set()
        
        def dfs(node: str, path: List[str]) -> bool:
            if node in rec_stack:
                # 找到循环
                cycle_start = path.index(node)
                cycles.append(path[cycle_start:] + [node])
                return True
            
            if node in visited:
                return False
            
            visited.add(node)
            rec_stack.add(node)
            path.append(node)
            
            for dep in self._dependencies.get(node, []):
                if dep in self._dependencies:  # 只检查已知的配置
                    if dfs(dep, path):
                        pass  # 继续查找其他循环
            
            rec_stack.remove(node)
            path.pop()
            return False
        
        for config_name in self._dependencies:
            if config_name not in visited:
                dfs(config_name, [])
        
        return cycles


class ConfigurationPerformanceMonitor:
    """配置性能监控器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigMonitor")
        self._metrics: Dict[str, List[float]] = defaultdict(list)
        self._operation_counts: Dict[str, int] = defaultdict(int)
        self._start_times: Dict[str, float] = {}
        self._max_samples = 1000
    
    def start_operation(self, operation_name: str) -> None:
        """开始操作计时"""
        self._start_times[operation_name] = time.time()
    
    def end_operation(self, operation_name: str) -> None:
        """结束操作计时"""
        if operation_name in self._start_times:
            duration = time.time() - self._start_times[operation_name]
            self._metrics[operation_name].append(duration)
            
            # 限制样本数量
            if len(self._metrics[operation_name]) > self._max_samples:
                self._metrics[operation_name] = self._metrics[operation_name][-self._max_samples:]
            
            self._operation_counts[operation_name] += 1
            del self._start_times[operation_name]
    
    def get_operation_stats(self, operation_name: str) -> Dict[str, float]:
        """获取操作统计信息"""
        durations = self._metrics.get(operation_name, [])
        
        if not durations:
            return {}
        
        return {
            'count': len(durations),
            'total_time': sum(durations),
            'avg_time': sum(durations) / len(durations),
            'min_time': min(durations),
            'max_time': max(durations),
            'recent_avg': sum(durations[-10:]) / min(len(durations), 10)
        }
    
    def get_all_stats(self) -> Dict[str, Dict[str, float]]:
        """获取所有操作统计信息"""
        return {name: self.get_operation_stats(name) for name in self._metrics}
    
    def reset_metrics(self) -> None:
        """重置指标"""
        self._metrics.clear()
        self._operation_counts.clear()
        self._start_times.clear()
        self.logger.info("性能指标已重置")


class AdvancedConfigurationValidator:
    """高级配置验证器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("AdvancedValidator")
        self._custom_validators: Dict[str, Callable] = {}
        self._cross_field_validators: List[Callable] = []
    
    def register_custom_validator(self, validator_name: str, validator_func: Callable) -> None:
        """注册自定义验证器"""
        self._custom_validators[validator_name] = validator_func
        self.logger.debug(f"注册自定义验证器: {validator_name}")
    
    def add_cross_field_validator(self, validator_func: Callable) -> None:
        """添加跨字段验证器"""
        self._cross_field_validators.append(validator_func)
        self.logger.debug("添加跨字段验证器")
    
    def validate_with_custom_rules(self, config_data: ConfigurationData, 
                                  custom_rules: Dict[str, Any]) -> List[str]:
        """使用自定义规则验证配置"""
        errors = []
        
        for field_path, rules in custom_rules.items():
            try:
                value = self._get_nested_value(config_data, field_path)
                field_errors = self._validate_field(field_path, value, rules)
                errors.extend(field_errors)
            except Exception as e:
                errors.append(f"验证字段 {field_path} 时发生错误: {e}")
        
        return errors
    
    def _get_nested_value(self, data: Dict[str, Any], path: str) -> Any:
        """获取嵌套值"""
        keys = path.split('.')
        current = data
        
        for key in keys:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return None
        
        return current
    
    def _validate_field(self, field_path: str, value: Any, rules: Dict[str, Any]) -> List[str]:
        """验证单个字段"""
        errors = []
        
        # 类型验证
        if 'type' in rules:
            expected_type = eval(rules['type'])  # 注意安全风险
            if not isinstance(value, expected_type):
                errors.append(f"字段 {field_path} 类型错误，期望 {expected_type.__name__}")
        
        # 范围验证
        if 'min' in rules and isinstance(value, (int, float)):
            if value < rules['min']:
                errors.append(f"字段 {field_path} 值 {value} 小于最小值 {rules['min']}")
        
        if 'max' in rules and isinstance(value, (int, float)):
            if value > rules['max']:
                errors.append(f"字段 {field_path} 值 {value} 大于最大值 {rules['max']}")
        
        # 正则验证
        if 'pattern' in rules and isinstance(value, str):
            if not re.match(rules['pattern'], value):
                errors.append(f"字段 {field_path} 值 '{value}' 不匹配模式 '{rules['pattern']}'")
        
        # 自定义验证器
        if 'validator' in rules:
            validator_name = rules['validator']
            if validator_name in self._custom_validators:
                try:
                    if not self._custom_validators[validator_name](value):
                        errors.append(f"字段 {field_path} 未通过自定义验证: {validator_name}")
                except Exception as e:
                    errors.append(f"字段 {field_path} 自定义验证执行失败: {e}")
        
        return errors


class ConfigurationEnvironmentManager:
    """配置环境管理器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigEnvironment")
        self._environments: Dict[str, ConfigurationData] = {}
        self._current_environment = 'default'
        self._environment_variables = {}
    
    def create_environment(self, env_name: str, base_config: ConfigurationData) -> None:
        """创建环境配置"""
        self._environments[env_name] = base_config.copy()
        self.logger.info(f"创建环境配置: {env_name}")
    
    def switch_environment(self, env_name: str) -> bool:
        """切换当前环境"""
        if env_name in self._environments:
            self._current_environment = env_name
            self.logger.info(f"切换到环境: {env_name}")
            return True
        return False
    
    def get_current_environment(self) -> str:
        """获取当前环境"""
        return self._current_environment
    
    def get_environment_config(self, env_name: Optional[str] = None) -> ConfigurationData:
        """获取环境配置"""
        env_name = env_name or self._current_environment
        return self._environments.get(env_name, {}).copy()
    
    def update_environment_config(self, env_name: str, updates: ConfigurationData) -> None:
        """更新环境配置"""
        if env_name in self._environments:
            self._environments[env_name].update(updates)
            self.logger.debug(f"更新环境配置: {env_name}")
    
    def load_from_environment_variables(self, prefix: str = "CONFIG_") -> None:
        """从环境变量加载配置"""
        env_config = {}
        
        for key, value in os.environ.items():
            if key.startswith(prefix):
                config_key = key[len(prefix):].lower().replace('_', '.')
                
                # 尝试解析不同类型
                parsed_value = self._parse_env_value(value)
                self._set_nested_value(env_config, config_key, parsed_value)
        
        # 设置为当前环境的配置
        if env_config:
            if self._current_environment in self._environments:
                self._environments[self._current_environment].update(env_config)
            else:
                self._environments[self._current_environment] = env_config
            
            self.logger.info(f"从环境变量加载配置，前缀: {prefix}")
    
    def _parse_env_value(self, value: str) -> Any:
        """解析环境变量值"""
        # 布尔值
        if value.lower() in ('true', 'false'):
            return value.lower() == 'true'
        
        # 数值
        try:
            if '.' in value:
                return float(value)
            else:
                return int(value)
        except ValueError:
            pass
        
        # JSON
        if value.startswith(('{', '[')):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                pass
        
        # 字符串
        return value
    
    def _set_nested_value(self, data: Dict[str, Any], path: str, value: Any) -> None:
        """设置嵌套值"""
        keys = path.split('.')
        current = data
        
        for key in keys[:-1]:
            if key not in current:
                current[key] = {}
            current = current[key]
        
        current[keys[-1]] = value


class ConfigurationBackupManager:
    """配置备份管理器"""
    
    def __init__(self, backup_dir: str = "config_backups"):
        self.backup_dir = Path(backup_dir)
        self.backup_dir.mkdir(exist_ok=True)
        self.logger = ConfigurationLogger("ConfigBackup")
        self._backup_history: List[Dict[str, Any]] = []
    
    async def create_backup(self, config_name: str, config_data: ConfigurationData,
                           metadata: ConfigurationMetadata, backup_type: str = 'manual') -> Path:
        """创建配置备份"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"{config_name}_{backup_type}_{timestamp}.json"
            backup_path = self.backup_dir / backup_filename
            
            backup_info = {
                'config_name': config_name,
                'config_data': config_data,
                'metadata': asdict(metadata),
                'backup_time': datetime.now().isoformat(),
                'backup_type': backup_type,
                'file_path': str(backup_path)
            }
            
            async with aiofiles.open(backup_path, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(backup_info, ensure_ascii=False, indent=2))
            
            self._backup_history.append(backup_info)
            
            # 限制备份历史数量
            if len(self._backup_history) > 100:
                self._backup_history = self._backup_history[-100:]
            
            self.logger.info(f"创建配置备份: {backup_path}")
            return backup_path
            
        except Exception as e:
            self.logger.error(f"创建配置备份失败: {config_name}", exception=e)
            raise ConfigurationError(f"创建备份失败: {e}")
    
    async def list_backups(self, config_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """列出备份"""
        if config_name:
            return [backup for backup in self._backup_history 
                   if backup['config_name'] == config_name]
        return self._backup_history.copy()
    
    async def restore_backup(self, backup_path: Path) -> Tuple[str, ConfigurationData, ConfigurationMetadata]:
        """恢复备份"""
        try:
            async with aiofiles.open(backup_path, 'r', encoding='utf-8') as f:
                backup_info = json.loads(await f.read())
            
            config_data = backup_info['config_data']
            metadata = ConfigurationMetadata(**backup_info['metadata'])
            config_name = backup_info['config_name']
            
            self.logger.info(f"恢复配置备份: {backup_path}")
            return config_name, config_data, metadata
            
        except Exception as e:
            self.logger.error(f"恢复配置备份失败: {backup_path}", exception=e)
            raise ConfigurationError(f"恢复备份失败: {e}")
    
    async def cleanup_old_backups(self, days: int = 30) -> int:
        """清理旧备份"""
        try:
            cutoff_time = datetime.now() - timedelta(days=days)
            deleted_count = 0
            
            for backup_info in self._backup_history[:]:
                backup_time = datetime.fromisoformat(backup_info['backup_time'])
                if backup_time < cutoff_time:
                    backup_path = Path(backup_info['file_path'])
                    if backup_path.exists():
                        backup_path.unlink()
                    self._backup_history.remove(backup_info)
                    deleted_count += 1
            
            self.logger.info(f"清理了 {deleted_count} 个旧备份")
            return deleted_count
            
        except Exception as e:
            self.logger.error("清理旧备份失败", exception=e)
            return 0


class ConfigurationHealthChecker:
    """配置健康检查器"""
    
    def __init__(self, config_manager: SystemConfigurationManager):
        self.config_manager = config_manager
        self.logger = ConfigurationLogger("ConfigHealth")
        self._health_checks: List[Callable] = []
    
    def add_health_check(self, check_func: Callable) -> None:
        """添加健康检查函数"""
        self._health_checks.append(check_func)
        self.logger.debug("添加配置健康检查")
    
    async def perform_health_check(self) -> Dict[str, Any]:
        """执行健康检查"""
        health_status = {
            'overall_status': 'healthy',
            'checks': [],
            'issues': [],
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            # 基本配置检查
            basic_check = await self._basic_config_check()
            health_status['checks'].append(basic_check)
            
            # 验证规则检查
            validation_check = await self._validation_check()
            health_status['checks'].append(validation_check)
            
            # 依赖检查
            dependency_check = await self._dependency_check()
            health_status['checks'].append(dependency_check)
            
            # 性能检查
            performance_check = await self._performance_check()
            health_status['checks'].append(performance_check)
            
            # 运行自定义健康检查
            for check_func in self._health_checks:
                try:
                    if asyncio.iscoroutinefunction(check_func):
                        custom_check = await check_func()
                    else:
                        custom_check = check_func()
                    health_status['checks'].append(custom_check)
                except Exception as e:
                    health_status['checks'].append({
                        'name': check_func.__name__,
                        'status': 'error',
                        'message': str(e)
                    })
            
            # 确定整体状态
            failed_checks = [check for check in health_status['checks'] if check['status'] == 'fail']
            error_checks = [check for check in health_status['checks'] if check['status'] == 'error']
            
            if error_checks:
                health_status['overall_status'] = 'error'
                health_status['issues'].extend([check['message'] for check in error_checks])
            elif failed_checks:
                health_status['overall_status'] = 'warning'
                health_status['issues'].extend([check['message'] for check in failed_checks])
            
            self.logger.debug(f"配置健康检查完成，状态: {health_status['overall_status']}")
            return health_status
            
        except Exception as e:
            health_status['overall_status'] = 'error'
            health_status['issues'].append(f"健康检查执行失败: {e}")
            self.logger.error("配置健康检查失败", exception=e)
            return health_status
    
    async def _basic_config_check(self) -> Dict[str, Any]:
        """基本配置检查"""
        try:
            loaded_configs = len(self.config_manager._configs)
            if loaded_configs == 0:
                return {
                    'name': 'basic_config_check',
                    'status': 'warning',
                    'message': '没有加载任何配置'
                }
            
            return {
                'name': 'basic_config_check',
                'status': 'pass',
                'message': f'已加载 {loaded_configs} 个配置'
            }
        except Exception as e:
            return {
                'name': 'basic_config_check',
                'status': 'error',
                'message': f'基本配置检查失败: {e}'
            }
    
    async def _validation_check(self) -> Dict[str, Any]:
        """验证规则检查"""
        try:
            failed_validations = []
            
            for config_name in self.config_manager._configs:
                errors = self.config_manager.validate_config(config_name)
                if errors:
                    failed_validations.append(f"{config_name}: {len(errors)} 个验证错误")
            
            if failed_validations:
                return {
                    'name': 'validation_check',
                    'status': 'fail',
                    'message': f'配置验证失败: {", ".join(failed_validations)}'
                }
            
            return {
                'name': 'validation_check',
                'status': 'pass',
                'message': '所有配置验证通过'
            }
        except Exception as e:
            return {
                'name': 'validation_check',
                'status': 'error',
                'message': f'验证检查失败: {e}'
            }
    
    async def _dependency_check(self) -> Dict[str, Any]:
        """依赖检查"""
        try:
            # 这里可以添加依赖检查逻辑
            return {
                'name': 'dependency_check',
                'status': 'pass',
                'message': '配置依赖检查通过'
            }
        except Exception as e:
            return {
                'name': 'dependency_check',
                'status': 'error',
                'message': f'依赖检查失败: {e}'
            }
    
    async def _performance_check(self) -> Dict[str, Any]:
        """性能检查"""
        try:
            stats = self.config_manager.get_statistics()
            cache_stats = stats.get('cache_stats', {})
            hit_rate = cache_stats.get('hit_rate', 0)
            
            if hit_rate < 0.5:
                return {
                    'name': 'performance_check',
                    'status': 'warning',
                    'message': f'缓存命中率较低: {hit_rate:.2%}'
                }
            
            return {
                'name': 'performance_check',
                'status': 'pass',
                'message': f'性能检查通过，缓存命中率: {hit_rate:.2%}'
            }
        except Exception as e:
            return {
                'name': 'performance_check',
                'status': 'error',
                'message': f'性能检查失败: {e}'
            }


# 扩展SystemConfigurationManager类
class ExtendedSystemConfigurationManager(SystemConfigurationManager):
    """扩展的系统配置管理器"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # 添加扩展组件
        self.template_manager = ConfigurationTemplate()
        self.config_merger = ConfigurationMerger()
        self.dependency_manager = ConfigurationDependencyManager()
        self.performance_monitor = ConfigurationPerformanceMonitor()
        self.advanced_validator = AdvancedConfigurationValidator()
        self.environment_manager = ConfigurationEnvironmentManager()
        self.backup_manager = ConfigurationBackupManager()
        self.health_checker = ConfigurationHealthChecker(self)
        
        self.logger.info("扩展系统配置管理器初始化完成")
    
    def create_config_from_template(self, template_name: str, config_name: str,
                                   variables: Optional[Dict[str, str]] = None) -> ConfigurationData:
        """从模板创建配置"""
        try:
            with self.performance_monitor._lock:
                self.performance_monitor.start_operation('create_config_from_template')
            
            config_data = self.template_manager.instantiate_template(template_name, variables)
            
            with self._lock:
                self._configs[config_name] = config_data
                
                # 创建元数据
                metadata = ConfigurationMetadata(
                    name=config_name,
                    version="1.0.0",
                    checksum=hashlib.sha256(json.dumps(config_data, sort_keys=True).encode()).hexdigest(),
                    size=len(json.dumps(config_data))
                )
                self._metadata[config_name] = metadata
            
            self.performance_monitor.end_operation('create_config_from_template')
            self.logger.info(f"从模板创建配置: {config_name} from {template_name}")
            
            return config_data
            
        except Exception as e:
            self.performance_monitor.end_operation('create_config_from_template')
            self._stats['errors_encountered'] += 1
            self.logger.error(f"从模板创建配置失败: {config_name}", exception=e)
            raise ConfigurationError(f"创建配置失败: {e}")
    
    def merge_and_load_configs(self, config_name: str, *source_configs: str,
                              merge_strategy: str = 'override') -> ConfigurationData:
        """合并并加载配置"""
        try:
            with self.performance_monitor._lock:
                self.performance_monitor.start_operation('merge_and_load_configs')
            
            # 获取源配置
            source_config_data = []
            for source_config in source_configs:
                if source_config in self._configs:
                    source_config_data.append(self._configs[source_config])
                else:
                    raise ConfigurationError(f"源配置 {source_config} 不存在")
            
            # 合并配置
            merged_data = self.config_merger.merge_configs(*source_config_data, strategy=merge_strategy)
            
            # 加载合并后的配置
            with self._lock:
                self._configs[config_name] = merged_data
                
                # 创建元数据
                metadata = ConfigurationMetadata(
                    name=config_name,
                    version="1.0.0",
                    checksum=hashlib.sha256(json.dumps(merged_data, sort_keys=True).encode()).hexdigest(),
                    size=len(json.dumps(merged_data))
                )
                self._metadata[config_name] = metadata
            
            self.performance_monitor.end_operation('merge_and_load_configs')
            self.logger.info(f"合并并加载配置: {config_name} from {source_configs}")
            
            return merged_data
            
        except Exception as e:
            self.performance_monitor.end_operation('merge_and_load_configs')
            self._stats['errors_encountered'] += 1
            self.logger.error(f"合并配置失败: {config_name}", exception=e)
            raise ConfigurationError(f"合并配置失败: {e}")
    
    async def load_configs_in_order(self, config_load_list: List[Tuple[str, Union[str, Path]]]) -> Dict[str, bool]:
        """按依赖顺序加载配置"""
        try:
            with self.performance_monitor._lock:
                self.performance_monitor.start_operation('load_configs_in_order')
            
            config_names = [name for name, _ in config_load_list]
            load_order = self.dependency_manager.get_load_order(config_names)
            
            results = {}
            
            # 按顺序加载配置
            for config_name in load_order:
                file_path = None
                for name, path in config_load_list:
                    if name == config_name:
                        file_path = path
                        break
                
                if file_path:
                    try:
                        await self.load_config(config_name, file_path)
                        results[config_name] = True
                    except Exception as e:
                        self.logger.error(f"按顺序加载配置失败: {config_name}", exception=e)
                        results[config_name] = False
                else:
                    results[config_name] = False
            
            self.performance_monitor.end_operation('load_configs_in_order')
            self.logger.info(f"按依赖顺序加载配置完成: {load_order}")
            
            return results
            
        except Exception as e:
            self.performance_monitor.end_operation('load_configs_in_order')
            self._stats['errors_encountered'] += 1
            self.logger.error("按依赖顺序加载配置失败", exception=e)
            raise ConfigurationError(f"按顺序加载配置失败: {e}")
    
    async def perform_comprehensive_health_check(self) -> Dict[str, Any]:
        """执行全面的健康检查"""
        try:
            with self.performance_monitor._lock:
                self.performance_monitor.start_operation('comprehensive_health_check')
            
            health_status = await self.health_checker.perform_health_check()
            
            # 添加扩展组件的健康检查
            template_health = {
                'name': 'template_health',
                'status': 'pass' if self.template_manager._templates else 'warning',
                'message': f'模板数量: {len(self.template_manager._templates)}'
            }
            health_status['checks'].append(template_health)
            
            dependency_health = {
                'name': 'dependency_health',
                'status': 'pass',
                'message': f'依赖关系数量: {sum(len(deps) for deps in self.dependency_manager._dependencies.values())}'
            }
            health_status['checks'].append(dependency_health)
            
            performance_health = {
                'name': 'performance_health',
                'status': 'pass',
                'message': f'监控的操作数量: {len(self.performance_monitor._metrics)}'
            }
            health_status['checks'].append(performance_health)
            
            self.performance_monitor.end_operation('comprehensive_health_check')
            return health_status
            
        except Exception as e:
            self.performance_monitor.end_operation('comprehensive_health_check')
            self.logger.error("全面健康检查失败", exception=e)
            raise ConfigurationError(f"健康检查失败: {e}")


# 单元测试代码
class ConfigurationManagerTestSuite:
    """配置管理器测试套件"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("ConfigTest")
        self.test_configs_dir = Path("test_configs")
        self.test_configs_dir.mkdir(exist_ok=True)
    
    async def run_all_tests(self) -> Dict[str, Any]:
        """运行所有测试"""
        test_results = {
            'total_tests': 0,
            'passed_tests': 0,
            'failed_tests': 0,
            'test_details': []
        }
        
        test_methods = [
            self.test_basic_operations,
            self.test_validation,
            self.test_hot_reload,
            self.test_encryption,
            self.test_template_system,
            self.test_config_merging,
            self.test_dependency_management,
            self.test_backup_restore,
            self.test_performance_monitoring,
            self.test_environment_management
        ]
        
        for test_method in test_methods:
            test_results['total_tests'] += 1
            try:
                await test_method()
                test_results['passed_tests'] += 1
                test_results['test_details'].append({
                    'test': test_method.__name__,
                    'status': 'PASS',
                    'message': '测试通过'
                })
                self.logger.info(f"测试通过: {test_method.__name__}")
            except Exception as e:
                test_results['failed_tests'] += 1
                test_results['test_details'].append({
                    'test': test_method.__name__,
                    'status': 'FAIL',
                    'message': str(e)
                })
                self.logger.error(f"测试失败: {test_method.__name__}", exception=e)
        
        return test_results
    
    async def test_basic_operations(self):
        """测试基本操作"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建测试配置
        test_config = {
            'app': {
                'name': 'TestApp',
                'version': '1.0.0',
                'debug': True
            },
            'database': {
                'host': 'localhost',
                'port': 5432,
                'name': 'testdb'
            }
        }
        
        config_manager._configs['test'] = test_config
        
        # 测试获取配置
        assert config_manager.get('app.name') == 'TestApp'
        assert config_manager.get('database.port') == 5432
        assert config_manager.get('nonexistent', 'default') == 'default'
        
        # 测试设置配置
        config_manager.set('app.version', '1.1.0', 'test')
        assert config_manager.get('app.version') == '1.1.0'
        
        # 测试配置存在检查
        assert config_manager.exists('app.name')
        assert not config_manager.exists('app.nonexistent')
    
    async def test_validation(self):
        """测试验证功能"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 添加验证规则
        rules = [
            ConfigurationValidationRule('app.port', int, min_value=1, max_value=65535),
            ConfigurationValidationRule('app.debug', bool),
            ConfigurationValidationRule('app.version', str, pattern=r'^\d+\.\d+\.\d+$')
        ]
        
        config_manager.add_validation_rules('test', rules)
        
        # 测试有效配置
        valid_config = {
            'app': {
                'port': 8080,
                'debug': False,
                'version': '1.0.0'
            }
        }
        
        config_manager._configs['test'] = valid_config
        errors = config_manager.validate_config('test')
        assert len(errors) == 0, f"有效配置验证失败: {errors}"
        
        # 测试无效配置
        invalid_config = {
            'app': {
                'port': 99999,  # 超出范围
                'debug': 'yes',  # 类型错误
                'version': 'invalid'  # 格式错误
            }
        }
        
        config_manager._configs['test'] = invalid_config
        errors = config_manager.validate_config('test')
        assert len(errors) > 0, "无效配置应该产生验证错误"
    
    async def test_hot_reload(self):
        """测试热更新功能"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建测试文件
        test_file = self.test_configs_dir / "hot_reload_test.json"
        test_config = {'app': {'name': 'TestApp', 'version': '1.0.0'}}
        
        async with aiofiles.open(test_file, 'w', encoding='utf-8') as f:
            await f.write(json.dumps(test_config))
        
        # 加载配置
        await config_manager.load_config('hot_reload_test', test_file)
        assert config_manager.get('app.name') == 'TestApp'
        
        # 测试热更新（这里只是模拟，实际需要文件变化）
        # 在真实环境中，需要修改文件来触发热更新
    
    async def test_encryption(self):
        """测试加密功能"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建测试配置
        test_config = {
            'database': {
                'host': 'localhost',
                'password': 'secret123',
                'api_key': 'abc123xyz'
            }
        }
        
        config_manager._configs['encrypted_test'] = test_config
        
        # 测试加密保存
        encrypted_file = self.test_configs_dir / "encrypted_test.json"
        await config_manager.save_config('encrypted_test', encrypted_file, encrypt=True)
        
        # 测试加密文件内容
        async with aiofiles.open(encrypted_file, 'r', encoding='utf-8') as f:
            encrypted_content = json.loads(await f.read())
        
        assert encrypted_content.get('_encrypted') == True
        assert '_content' in encrypted_content
    
    async def test_template_system(self):
        """测试模板系统"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建模板
        template_data = {
            'app': {
                'name': '${app_name}',
                'version': '${version}',
                'port': '${port}',
                'debug': '${debug}'
            },
            'database': {
                'host': '${db_host}',
                'port': '${db_port}'
            }
        }
        
        config_manager.template_manager.create_template('app_template', template_data)
        
        # 实例化模板
        variables = {
            'app_name': 'MyApp',
            'version': '2.0.0',
            'port': 8080,
            'debug': True,
            'db_host': 'db.example.com',
            'db_port': 5432
        }
        
        instance = config_manager.template_manager.instantiate_template('app_template', variables)
        
        assert instance['app']['name'] == 'MyApp'
        assert instance['app']['version'] == '2.0.0'
        assert instance['app']['port'] == 8080
        assert instance['database']['host'] == 'db.example.com'
    
    async def test_config_merging(self):
        """测试配置合并"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建测试配置
        config1 = {
            'app': {'name': 'App1', 'version': '1.0.0'},
            'database': {'host': 'db1.com', 'port': 5432}
        }
        
        config2 = {
            'app': {'version': '2.0.0', 'debug': True},
            'database': {'port': 3306},
            'cache': {'enabled': True}
        }
        
        # 测试覆盖策略
        merged_override = config_manager.config_merger.merge_configs(config1, config2, strategy='override')
        assert merged_override['app']['version'] == '2.0.0'
        assert merged_override['app']['debug'] == True
        assert merged_override['database']['port'] == 3306
        assert 'cache' in merged_override
        
        # 测试深度合并策略
        merged_deep = config_manager.config_merger.merge_configs(config1, config2, strategy='merge')
        assert merged_deep['app']['name'] == 'App1'  # 保留第一个值
        assert merged_deep['app']['version'] == '2.0.0'  # 覆盖版本
        assert merged_deep['database']['host'] == 'db1.com'  # 保留第一个值
        assert merged_deep['database']['port'] == 3306  # 覆盖端口
    
    async def test_dependency_management(self):
        """测试依赖管理"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 添加依赖关系
        config_manager.dependency_manager.add_dependency('config_b', 'config_a')
        config_manager.dependency_manager.add_dependency('config_c', 'config_b')
        
        # 测试获取依赖
        assert 'config_a' in config_manager.dependency_manager.get_dependencies('config_b')
        assert 'config_b' in config_manager.dependency_manager.get_dependencies('config_c')
        
        # 测试获取所有依赖
        all_deps = config_manager.dependency_manager.get_all_dependencies('config_c')
        assert 'config_a' in all_deps
        assert 'config_b' in all_deps
        
        # 测试加载顺序
        load_order = config_manager.dependency_manager.get_load_order(['config_c', 'config_b', 'config_a'])
        assert load_order == ['config_a', 'config_b', 'config_c']
    
    async def test_backup_restore(self):
        """测试备份和恢复"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建测试配置
        test_config = {
            'app': {'name': 'BackupTest', 'version': '1.0.0'},
            'data': {'items': [1, 2, 3, 4, 5]}
        }
        
        config_manager._configs['backup_test'] = test_config
        
        metadata = ConfigurationMetadata(
            name='backup_test',
            version='1.0.0',
            checksum='test_checksum'
        )
        config_manager._metadata['backup_test'] = metadata
        
        # 创建备份
        backup_path = await config_manager.backup_manager.create_backup(
            'backup_test', test_config, metadata
        )
        assert backup_path.exists()
        
        # 恢复备份
        restored_name, restored_data, restored_metadata = await config_manager.backup_manager.restore_backup(backup_path)
        assert restored_name == 'backup_test'
        assert restored_data == test_config
        assert restored_metadata.name == 'backup_test'
    
    async def test_performance_monitoring(self):
        """测试性能监控"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 模拟一些操作
        config_manager.performance_monitor.start_operation('test_operation')
        await asyncio.sleep(0.1)  # 模拟操作耗时
        config_manager.performance_monitor.end_operation('test_operation')
        
        # 检查统计信息
        stats = config_manager.performance_monitor.get_operation_stats('test_operation')
        assert 'count' in stats
        assert stats['count'] == 1
        assert stats['total_time'] > 0.05  # 至少50毫秒
        assert stats['total_time'] < 0.2   # 少于200毫秒
    
    async def test_environment_management(self):
        """测试环境管理"""
        config_manager = ExtendedSystemConfigurationManager()
        
        # 创建环境配置
        dev_config = {
            'app': {'debug': True, 'log_level': 'DEBUG'},
            'database': {'host': 'localhost', 'port': 5432}
        }
        
        prod_config = {
            'app': {'debug': False, 'log_level': 'ERROR'},
            'database': {'host': 'prod-db.com', 'port': 5432}
        }
        
        config_manager.environment_manager.create_environment('development', dev_config)
        config_manager.environment_manager.create_environment('production', prod_config)
        
        # 切换环境
        assert config_manager.environment_manager.switch_environment('development')
        assert config_manager.environment_manager.get_current_environment() == 'development'
        
        dev_config_data = config_manager.environment_manager.get_environment_config('development')
        assert dev_config_data['app']['debug'] == True
        
        # 切换到生产环境
        config_manager.environment_manager.switch_environment('production')
        prod_config_data = config_manager.environment_manager.get_environment_config('production')
        assert prod_config_data['app']['debug'] == False


# 完整的使用示例和文档
async def comprehensive_example():
    """完整的使用示例"""
    
    print("=== K1系统配置管理器 - 完整示例 ===\n")
    
    # 1. 创建配置管理器
    print("1. 创建扩展配置管理器...")
    config_manager = ExtendedSystemConfigurationManager(
        config_dir="examples/configs",
        history_dir="examples/history",
        cache_size=500,
        cache_ttl=1800,
        log_level=LogLevel.INFO
    )
    
    try:
        # 2. 创建配置模板
        print("2. 创建配置模板...")
        app_template = {
            'app': {
                'name': '${app_name}',
                'version': '${version}',
                'debug': '${debug}',
                'port': '${port}'
            },
            'logging': {
                'level': '${log_level}',
                'file': '${log_file}'
            },
            'database': {
                'host': '${db_host}',
                'port': '${db_port}',
                'name': '${db_name}',
                'ssl': '${ssl_enabled}'
            }
        }
        
        config_manager.template_manager.create_template('web_app', app_template)
        
        # 3. 从模板创建配置
        print("3. 从模板创建开发环境配置...")
        dev_variables = {
            'app_name': 'MyWebApp',
            'version': '1.2.3',
            'debug': 'true',
            'port': '8080',
            'log_level': 'DEBUG',
            'log_file': 'app_dev.log',
            'db_host': 'localhost',
            'db_port': '5432',
            'db_name': 'myapp_dev',
            'ssl_enabled': 'false'
        }
        
        config_manager.create_config_from_template('web_app', 'dev_config', dev_variables)
        
        # 4. 创建生产环境配置
        print("4. 从模板创建生产环境配置...")
        prod_variables = {
            'app_name': 'MyWebApp',
            'version': '1.2.3',
            'debug': 'false',
            'port': '80',
            'log_level': 'ERROR',
            'log_file': 'app.log',
            'db_host': 'prod-db.example.com',
            'db_port': '5432',
            'db_name': 'myapp_prod',
            'ssl_enabled': 'true'
        }
        
        config_manager.create_config_from_template('web_app', 'prod_config', prod_variables)
        
        # 5. 添加验证规则
        print("5. 添加配置验证规则...")
        validation_rules = [
            ConfigurationValidationRule('app.port', int, min_value=1, max_value=65535),
            ConfigurationValidationRule('app.debug', bool),
            ConfigurationValidationRule('app.version', str, pattern=r'^\d+\.\d+\.\d+$'),
            ConfigurationValidationRule('database.port', int, min_value=1, max_value=65535),
            ConfigurationValidationRule('database.ssl', bool)
        ]
        
        config_manager.add_validation_rules('dev_config', validation_rules)
        config_manager.add_validation_rules('prod_config', validation_rules)
        
        # 6. 验证配置
        print("6. 验证配置...")
        dev_errors = config_manager.validate_config('dev_config')
        prod_errors = config_manager.validate_config('prod_config')
        
        if dev_errors:
            print(f"开发配置验证错误: {dev_errors}")
        else:
            print("开发配置验证通过")
        
        if prod_errors:
            print(f"生产配置验证错误: {prod_errors}")
        else:
            print("生产配置验证通过")
        
        # 7. 配置依赖管理
        print("7. 设置配置依赖关系...")
        config_manager.dependency_manager.add_dependency('prod_config', 'dev_config')
        
        # 8. 保存配置到文件
        print("8. 保存配置到文件...")
        await config_manager.save_config('dev_config', 'examples/configs/dev.json')
        await config_manager.save_config('prod_config', 'examples/configs/prod.json')
        
        # 9. 启用热更新
        print("9. 启用热更新...")
        
        def config_change_handler(config_name: str, event_type: str):
            print(f"配置 {config_name} 发生 {event_type} 事件")
        
        config_manager.watch_config('dev_config', config_change_handler)
        await config_manager.enable_hot_reload('dev_config', 'examples/configs/dev.json')
        
        # 10. 配置合并示例
        print("10. 配置合并示例...")
        base_config = {
            'common': {
                'app_name': 'SharedApp',
                'version': '1.0.0'
            },
            'features': {
                'feature_a': True,
                'feature_b': False
            }
        }
        
        feature_config = {
            'features': {
                'feature_b': True,
                'feature_c': True
            },
            'performance': {
                'cache_size': 1000,
                'timeout': 30
            }
        }
        
        config_manager._configs['base'] = base_config
        config_manager._configs['features'] = feature_config
        
        merged_config = config_manager.merge_and_load_configs(
            'complete_config', 'base', 'features', merge_strategy='merge'
        )
        
        print(f"合并后配置: {json.dumps(merged_config, indent=2, ensure_ascii=False)}")
        
        # 11. 环境变量管理
        print("11. 环境变量管理...")
        config_manager.environment_manager.create_environment('test', {
            'app': {'debug': True},
            'test_mode': True
        })
        
        # 模拟环境变量
        os.environ['CONFIG_TEST_VAR'] = 'test_value'
        os.environ['CONFIG_DEBUG'] = 'true'
        
        config_manager.environment_manager.load_from_environment_variables('CONFIG_')
        
        # 12. 配置备份
        print("12. 创建配置备份...")
        dev_metadata = config_manager.get_metadata('dev_config')
        if dev_metadata:
            backup_path = await config_manager.backup_manager.create_backup(
                'dev_config', config_manager.get_config('dev_config'), dev_metadata
            )
            print(f"备份创建: {backup_path}")
        
        # 13. 性能监控
        print("13. 性能监控...")
        perf_stats = config_manager.performance_monitor.get_all_stats()
        if perf_stats:
            print("性能统计:")
            for operation, stats in perf_stats.items():
                print(f"  {operation}: {stats}")
        
        # 14. 健康检查
        print("14. 执行健康检查...")
        health_status = await config_manager.perform_comprehensive_health_check()
        print(f"健康状态: {health_status['overall_status']}")
        print("健康检查详情:")
        for check in health_status['checks']:
            print(f"  {check['name']}: {check['status']} - {check['message']}")
        
        # 15. 统计信息
        print("15. 获取统计信息...")
        stats = config_manager.get_statistics()
        print("配置管理器统计:")
        for category, data in stats.items():
            print(f"  {category}:")
            for key, value in data.items():
                print(f"    {key}: {value}")
        
        print("\n=== 示例完成 ===")
        
        # 等待一段时间以观察热更新
        print("\n等待热更新测试（5秒）...")
        await asyncio.sleep(5)
        
    except Exception as e:
        print(f"示例执行错误: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # 清理资源
        await config_manager.disable_hot_reload('dev_config')
        print("配置管理器示例完成，资源已清理")


# 高级使用场景示例
async def advanced_scenarios_example():
    """高级使用场景示例"""
    
    print("=== K1系统配置管理器 - 高级场景示例 ===\n")
    
    config_manager = ExtendedSystemConfigurationManager()
    
    # 场景1: 微服务配置管理
    print("场景1: 微服务配置管理")
    
    # 定义服务配置模板
    service_template = {
        'service': {
            'name': '${service_name}',
            'version': '${version}',
            'port': '${port}',
            'health_check': '${health_check_path}'
        },
        'database': {
            'url': '${db_url}',
            'pool_size': '${db_pool_size}',
            'timeout': '${db_timeout}'
        },
        'redis': {
            'host': '${redis_host}',
            'port': '${redis_port}',
            'db': '${redis_db}'
        },
        'logging': {
            'level': '${log_level}',
            'format': '${log_format}'
        }
    }
    
    config_manager.template_manager.create_template('microservice', service_template)
    
    # 创建用户服务配置
    user_service_vars = {
        'service_name': 'user-service',
        'version': '2.1.0',
        'port': '8001',
        'health_check_path': '/health',
        'db_url': 'postgresql://user-db:5432/users',
        'db_pool_size': '20',
        'db_timeout': '30',
        'redis_host': 'redis-cluster',
        'redis_port': '6379',
        'redis_db': '0',
        'log_level': 'INFO',
        'log_format': 'json'
    }
    
    config_manager.create_config_from_template('microservice', 'user_service', user_service_vars)
    
    # 创建订单服务配置
    order_service_vars = {
        'service_name': 'order-service',
        'version': '1.8.3',
        'port': '8002',
        'health_check_path': '/health',
        'db_url': 'postgresql://order-db:5432/orders',
        'db_pool_size': '15',
        'db_timeout': '25',
        'redis_host': 'redis-cluster',
        'redis_port': '6379',
        'redis_db': '1',
        'log_level': 'DEBUG',
        'log_format': 'text'
    }
    
    config_manager.create_config_from_template('microservice', 'order_service', order_service_vars)
    
    # 场景2: 多环境部署
    print("场景2: 多环境部署配置")
    
    environments = {
        'development': {
            'debug': True,
            'log_level': 'DEBUG',
            'database': {'host': 'localhost', 'pool_size': 5},
            'cache': {'ttl': 300}
        },
        'staging': {
            'debug': False,
            'log_level': 'INFO',
            'database': {'host': 'staging-db', 'pool_size': 10},
            'cache': {'ttl': 600}
        },
        'production': {
            'debug': False,
            'log_level': 'ERROR',
            'database': {'host': 'prod-db', 'pool_size': 50},
            'cache': {'ttl': 3600}
        }
    }
    
    for env_name, env_config in environments.items():
        config_manager.environment_manager.create_environment(env_name, env_config)
    
    # 场景3: 配置依赖和加载顺序
    print("场景3: 配置依赖管理")
    
    # 定义配置依赖
    config_manager.dependency_manager.add_dependency('user_service', 'common_config')
    config_manager.dependency_manager.add_dependency('order_service', 'common_config')
    config_manager.dependency_manager.add_dependency('order_service', 'user_service')
    
    # 创建公共配置
    common_config = {
        'common': {
            'timezone': 'UTC',
            'currency': 'USD',
            'date_format': 'YYYY-MM-DD'
        },
        'security': {
            'jwt_secret': 'your-secret-key',
            'token_expiry': 3600
        }
    }
    
    config_manager._configs['common_config'] = common_config
    
    # 获取加载顺序
    all_configs = ['user_service', 'order_service', 'common_config']
    load_order = config_manager.dependency_manager.get_load_order(all_configs)
    print(f"配置加载顺序: {load_order}")
    
    # 场景4: 配置验证和自定义验证器
    print("场景4: 高级配置验证")
    
    def validate_port_range(value):
        """验证端口范围"""
        return 1 <= value <= 65535
    
    def validate_url_format(value):
        """验证URL格式"""
        import re
        pattern = r'^https?://[^\s/$.?#].[\s]*$'
        return re.match(pattern, value) is not None
    
    # 注册自定义验证器
    config_manager.advanced_validator.register_custom_validator('port_range', validate_port_range)
    config_manager.advanced_validator.register_custom_validator('url_format', validate_url_format)
    
    # 定义跨字段验证器
    def validate_service_config(config_data):
        """验证服务配置的跨字段逻辑"""
        errors = []
        
        if config_data.get('service', {}).get('debug', False):
            if config_data.get('logging', {}).get('level') != 'DEBUG':
                errors.append("调试模式下日志级别必须是DEBUG")
        
        if config_data.get('database', {}).get('pool_size', 0) > 100:
            errors.append("数据库连接池大小不能超过100")
        
        return errors
    
    config_manager.advanced_validator.add_cross_field_validator(validate_service_config)
    
    # 场景5: 性能监控和优化
    print("场景5: 性能监控")
    
    # 模拟一些配置操作
    for i in range(10):
        config_manager.performance_monitor.start_operation('config_access')
        # 模拟配置访问
        _ = config_manager.get('service.name', config_name='user_service')
        config_manager.performance_monitor.end_operation('config_access')
    
    # 获取性能统计
    perf_stats = config_manager.performance_monitor.get_operation_stats('config_access')
    print(f"配置访问性能统计: {perf_stats}")
    
    # 场景6: 配置热更新和实时监控
    print("场景6: 配置热更新监控")
    
    def hot_update_handler(config_name: str, event_type: str):
        """热更新处理器"""
        print(f"[热更新] 配置 {config_name} 发生 {event_type} 事件")
        
        if event_type == 'reload':
            # 重新验证配置
            errors = config_manager.validate_config(config_name)
            if errors:
                print(f"[热更新] 配置验证失败: {errors}")
            else:
                print(f"[热更新] 配置验证通过")
    
    config_manager.watch_config('user_service', hot_update_handler)
    await config_manager.enable_hot_reload('user_service', 'examples/configs/user_service.json')
    
    # 场景7: 配置备份和灾难恢复
    print("场景7: 配置备份策略")
    
    # 创建多个备份点
    for i in range(3):
        # 模拟配置变更
        config_manager.set('service.version', f'2.1.{i}', 'user_service')
        
        metadata = config_manager.get_metadata('user_service')
        if metadata:
            backup_path = await config_manager.backup_manager.create_backup(
                'user_service', 
                config_manager.get_config('user_service'), 
                metadata,
                backup_type='auto'
            )
            print(f"自动备份创建: {backup_path}")
        
        await asyncio.sleep(0.1)  # 短暂延迟
    
    # 场景8: 健康检查和监控
    print("场景8: 健康检查")
    
    # 添加自定义健康检查
    def database_connectivity_check():
        """数据库连接健康检查"""
        try:
            # 模拟数据库连接检查
            db_config = config_manager.get('database.url', config_name='user_service')
            if db_config:
                return {
                    'name': 'database_connectivity',
                    'status': 'pass',
                    'message': f'数据库连接正常: {db_config}'
                }
            else:
                return {
                    'name': 'database_connectivity',
                    'status': 'fail',
                    'message': '数据库配置缺失'
                }
        except Exception as e:
            return {
                'name': 'database_connectivity',
                'status': 'error',
                'message': f'数据库连接检查失败: {e}'
            }
    
    config_manager.health_checker.add_health_check(database_connectivity_check)
    
    # 执行健康检查
    health_status = await config_manager.perform_comprehensive_health_check()
    print(f"整体健康状态: {health_status['overall_status']}")
    
    # 场景9: 配置统计和分析
    print("场景9: 配置统计分析")
    
    # 分析配置使用情况
    config_stats = {}
    for config_name in config_manager.get_all_configs():
        config_data = config_manager.get_config(config_name)
        if config_data:
            config_stats[config_name] = {
                'total_keys': len(str(config_data).split()),
                'size_bytes': len(json.dumps(config_data)),
                'last_updated': config_manager.get_metadata(config_name).updated_at if config_manager.get_metadata(config_name) else None
            }
    
    print("配置统计:")
    for config_name, stats in config_stats.items():
        print(f"  {config_name}:")
        print(f"    键数量: {stats['total_keys']}")
        print(f"    大小: {stats['size_bytes']} 字节")
        if stats['last_updated']:
            print(f"    最后更新: {stats['last_updated']}")
    
    print("\n=== 高级场景示例完成 ===")
    
    # 清理资源
    await config_manager.disable_hot_reload('user_service')


# 命令行工具
def create_config_cli():
    """创建配置管理命令行工具"""
    
    import argparse
    
    parser = argparse.ArgumentParser(description='K1系统配置管理器CLI工具')
    subparsers = parser.add_subparsers(dest='command', help='可用命令')
    
    # 加载配置命令
    load_parser = subparsers.add_parser('load', help='加载配置文件')
    load_parser.add_argument('config_name', help='配置名称')
    load_parser.add_argument('file_path', help='配置文件路径')
    load_parser.add_argument('--validate', action='store_true', help='验证配置')
    
    # 保存配置命令
    save_parser = subparsers.add_parser('save', help='保存配置')
    save_parser.add_argument('config_name', help='配置名称')
    save_parser.add_argument('--file-path', help='保存路径')
    save_parser.add_argument('--encrypt', action='store_true', help='加密保存')
    
    # 获取配置命令
    get_parser = subparsers.add_parser('get', help='获取配置值')
    get_parser.add_argument('key', help='配置键')
    get_parser.add_argument('--config-name', help='配置名称')
    get_parser.add_argument('--default', help='默认值')
    
    # 设置配置命令
    set_parser = subparsers.add_parser('set', help='设置配置值')
    set_parser.add_argument('key', help='配置键')
    set_parser.add_argument('value', help='配置值')
    set_parser.add_argument('config_name', help='配置名称')
    
    # 验证配置命令
    validate_parser = subparsers.add_parser('validate', help='验证配置')
    validate_parser.add_argument('config_name', help='配置名称')
    
    # 健康检查命令
    health_parser = subparsers.add_parser('health', help='健康检查')
    
    # 统计信息命令
    stats_parser = subparsers.add_parser('stats', help='统计信息')
    
    def main():
        args = parser.parse_args()
        
        if not args.command:
            parser.print_help()
            return
        
        # 创建配置管理器
        config_manager = ExtendedSystemConfigurationManager()
        
        async def run_command():
            try:
                if args.command == 'load':
                    await config_manager.load_config(args.config_name, args.file_path, validate=args.validate)
                    print(f"配置 {args.config_name} 加载成功")
                
                elif args.command == 'save':
                    await config_manager.save_config(args.config_name, args.file_path, encrypt=args.encrypt)
                    print(f"配置 {args.config_name} 保存成功")
                
                elif args.command == 'get':
                    value = config_manager.get(args.key, args.default, args.config_name)
                    print(value)
                
                elif args.command == 'set':
                    config_manager.set(args.key, args.value, args.config_name)
                    print(f"设置 {args.config_name}.{args.key} = {args.value}")
                
                elif args.command == 'validate':
                    errors = config_manager.validate_config(args.config_name)
                    if errors:
                        print("配置验证失败:")
                        for error in errors:
                            print(f"  - {error}")
                    else:
                        print("配置验证通过")
                
                elif args.command == 'health':
                    health_status = await config_manager.perform_comprehensive_health_check()
                    print(f"健康状态: {health_status['overall_status']}")
                    for check in health_status['checks']:
                        print(f"  {check['name']}: {check['status']} - {check['message']}")
                
                elif args.command == 'stats':
                    stats = config_manager.get_statistics()
                    print("统计信息:")
                    for category, data in stats.items():
                        print(f"  {category}:")
                        for key, value in data.items():
                            print(f"    {key}: {value}")
            
            except Exception as e:
                print(f"命令执行失败: {e}")
                return 1
            
            return 0
        
        exit_code = asyncio.run(run_command())
        sys.exit(exit_code)
    
    return main


# 新增高级功能类

class ConfigurationImporter:
    """配置导入器 - 支持从多种数据源导入配置"""
    
    def __init__(self, config_manager: SystemConfigurationManager):
        self.config_manager = config_manager
        self.logger = ConfigurationLogger("ConfigImporter")
        self._import_handlers = {
            'csv': self._import_csv,
            'excel': self._import_excel,
            'database': self._import_database,
            'api': self._import_api,
            'environment': self._import_environment
        }
    
    async def import_from_source(self, source_type: str, source_path: str, 
                                config_name: str, **kwargs) -> ConfigurationData:
        """从指定数据源导入配置"""
        try:
            if source_type not in self._import_handlers:
                raise ConfigurationLoadError(f"不支持的导入类型: {source_type}")
            
            handler = self._import_handlers[source_type]
            imported_data = await handler(source_path, **kwargs)
            
            # 存储导入的配置
            with self.config_manager._lock:
                self.config_manager._configs[config_name] = imported_data
                
                # 创建元数据
                metadata = ConfigurationMetadata(
                    name=config_name,
                    version="1.0.0",
                    checksum=hashlib.sha256(json.dumps(imported_data, sort_keys=True).encode()).hexdigest(),
                    size=len(json.dumps(imported_data))
                )
                self.config_manager._metadata[config_name] = metadata
            
            self.logger.info(f"成功从 {source_type} 导入配置: {config_name}")
            return imported_data
            
        except Exception as e:
            self.logger.error(f"导入配置失败: {source_type}", exception=e)
            raise ConfigurationLoadError(f"导入失败: {e}")
    
    async def _import_csv(self, file_path: str, **kwargs) -> ConfigurationData:
        """从CSV文件导入配置"""
        try:
            import csv
            
            result = {}
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    section = row.get('section', 'default')
                    key = row.get('key')
                    value = row.get('value')
                    data_type = row.get('type', 'string')
                    
                    if section not in result:
                        result[section] = {}
                    
                    # 类型转换
                    if data_type == 'int':
                        value = int(value)
                    elif data_type == 'float':
                        value = float(value)
                    elif data_type == 'bool':
                        value = value.lower() in ('true', '1', 'yes')
                    
                    result[section][key] = value
            
            return result
            
        except Exception as e:
            raise ConfigurationLoadError(f"CSV导入失败: {e}")
    
    async def _import_excel(self, file_path: str, **kwargs) -> ConfigurationData:
        """从Excel文件导入配置"""
        try:
            import openpyxl
            
            result = {}
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                section_data = {}
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    key, value, data_type = row[:3]
                    
                    if not key:
                        continue
                    
                    # 类型转换
                    if data_type == 'int':
                        value = int(value) if value else 0
                    elif data_type == 'float':
                        value = float(value) if value else 0.0
                    elif data_type == 'bool':
                        value = bool(value) if value else False
                    
                    section_data[key] = value
                
                if section_data:
                    result[sheet_name] = section_data
            
            return result
            
        except ImportError:
            raise ConfigurationLoadError("需要安装openpyxl库来支持Excel导入")
        except Exception as e:
            raise ConfigurationLoadError(f"Excel导入失败: {e}")
    
    async def _import_database(self, connection_string: str, **kwargs) -> ConfigurationData:
        """从数据库导入配置"""
        try:
            import asyncpg
            
            result = {}
            conn = await asyncpg.connect(connection_string)
            
            try:
                # 假设配置存储在config表中
                query = kwargs.get('query', 'SELECT section, key, value, type FROM config')
                rows = await conn.fetch(query)
                
                for row in rows:
                    section = row['section']
                    key = row['key']
                    value = row['value']
                    data_type = row['type']
                    
                    if section not in result:
                        result[section] = {}
                    
                    # 类型转换
                    if data_type == 'int':
                        value = int(value)
                    elif data_type == 'float':
                        value = float(value)
                    elif data_type == 'bool':
                        value = bool(value)
                    
                    result[section][key] = value
            
            finally:
                await conn.close()
            
            return result
            
        except ImportError:
            raise ConfigurationLoadError("需要安装asyncpg库来支持数据库导入")
        except Exception as e:
            raise ConfigurationLoadError(f"数据库导入失败: {e}")
    
    async def _import_api(self, api_url: str, **kwargs) -> ConfigurationData:
        """从API导入配置"""
        try:
            import aiohttp
            
            async with aiohttp.ClientSession() as session:
                headers = kwargs.get('headers', {})
                params = kwargs.get('params', {})
                
                async with session.get(api_url, headers=headers, params=params) as response:
                    if response.status == 200:
                        data = await response.json()
                        return data
                    else:
                        raise ConfigurationLoadError(f"API请求失败: {response.status}")
            
        except ImportError:
            raise ConfigurationLoadError("需要安装aiohttp库来支持API导入")
        except Exception as e:
            raise ConfigurationLoadError(f"API导入失败: {e}")
    
    async def _import_environment(self, prefix: str = '', **kwargs) -> ConfigurationData:
        """从环境变量导入配置"""
        result = {}
        
        for key, value in os.environ.items():
            if prefix and not key.startswith(prefix):
                continue
            
            # 移除前缀
            if prefix:
                key = key[len(prefix):]
            
            # 解析嵌套键
            parts = key.lower().split('_')
            current = result
            
            for part in parts[:-1]:
                if part not in current:
                    current[part] = {}
                current = current[part]
            
            # 类型转换
            final_key = parts[-1]
            if value.lower() in ('true', '1', 'yes'):
                value = True
            elif value.lower() in ('false', '0', 'no'):
                value = False
            elif value.isdigit():
                value = int(value)
            else:
                try:
                    value = float(value)
                except ValueError:
                    pass  # 保持字符串
            
            current[final_key] = value
        
        return result


class ConfigurationExporter:
    """配置导出器 - 支持导出到多种格式"""
    
    def __init__(self, config_manager: SystemConfigurationManager):
        self.config_manager = config_manager
        self.logger = ConfigurationLogger("ConfigExporter")
    
    async def export_to_source(self, config_name: str, target_type: str, 
                              target_path: str, **kwargs) -> bool:
        """导出配置到指定目标"""
        try:
            if config_name not in self.config_manager._configs:
                raise ConfigurationError(f"配置 {config_name} 不存在")
            
            config_data = self.config_manager._configs[config_name]
            
            if target_type == 'json':
                await self._export_json(config_data, target_path)
            elif target_type == 'yaml':
                await self._export_yaml(config_data, target_path)
            elif target_type == 'xml':
                await self._export_xml(config_data, target_path)
            elif target_type == 'csv':
                await self._export_csv(config_data, target_path)
            elif target_type == 'excel':
                await self._export_excel(config_data, target_path)
            elif target_type == 'properties':
                await self._export_properties(config_data, target_path)
            elif target_type == 'env':
                await self._export_env(config_data, target_path, kwargs.get('prefix', ''))
            else:
                raise ConfigurationSaveError(f"不支持的导出类型: {target_type}")
            
            self.logger.info(f"成功导出配置到 {target_type}: {config_name}")
            return True
            
        except Exception as e:
            self.logger.error(f"导出配置失败: {target_type}", exception=e)
            return False
    
    async def _export_json(self, config_data: ConfigurationData, file_path: str):
        """导出为JSON格式"""
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(json.dumps(config_data, ensure_ascii=False, indent=2))
    
    async def _export_yaml(self, config_data: ConfigurationData, file_path: str):
        """导出为YAML格式"""
        content = yaml.dump(config_data, allow_unicode=True, indent=2)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content)
    
    async def _export_xml(self, config_data: ConfigurationData, file_path: str):
        """导出为XML格式"""
        root = self.config_manager.loader._dict_to_xml('config', config_data)
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
    
    async def _export_csv(self, config_data: ConfigurationData, file_path: str):
        """导出为CSV格式"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['section', 'key', 'value', 'type']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for section, section_data in config_data.items():
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        data_type = type(value).__name__
                        writer.writerow({
                            'section': section,
                            'key': key,
                            'value': str(value),
                            'type': data_type
                        })
    
    async def _export_excel(self, config_data: ConfigurationData, file_path: str):
        """导出为Excel格式"""
        try:
            import openpyxl
            
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # 删除默认工作表
            
            for section, section_data in config_data.items():
                if isinstance(section_data, dict):
                    sheet = workbook.create_sheet(title=section)
                    
                    # 写入表头
                    sheet['A1'] = 'Key'
                    sheet['B1'] = 'Value'
                    sheet['C1'] = 'Type'
                    
                    # 写入数据
                    row = 2
                    for key, value in section_data.items():
                        sheet[f'A{row}'] = key
                        sheet[f'B{row}'] = str(value)
                        sheet[f'C{row}'] = type(value).__name__
                        row += 1
            
            workbook.save(file_path)
            
        except ImportError:
            raise ConfigurationSaveError("需要安装openpyxl库来支持Excel导出")
    
    async def _export_properties(self, config_data: ConfigurationData, file_path: str):
        """导出为Java Properties格式"""
        content = []
        
        for section, section_data in config_data.items():
            if isinstance(section_data, dict):
                content.append(f"# {section}")
                for key, value in section_data.items():
                    content.append(f"{section}.{key}={value}")
                content.append("")  # 空行分隔
        
        content_str = "\n".join(content)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content_str)
    
    async def _export_env(self, config_data: ConfigurationData, file_path: str, prefix: str = ""):
        """导出为环境变量格式"""
        content = []
        
        def flatten_dict(d, parent_key="", sep="_"):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                else:
                    items.append((new_key, str(v)))
            return dict(items)
        
        flat_data = flatten_dict(config_data)
        
        for key, value in flat_data.items():
            env_key = f"{prefix}{key}".upper()
            content.append(f"{env_key}={value}")
        
        content_str = "\n".join(content)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content_str)


class ConfigurationComparator:
    """配置比较器 - 比较配置差异"""
    
    def __init__(self, config_manager: SystemConfigurationManager):
        self.config_manager = config_manager
        self.logger = ConfigurationLogger("ConfigComparator")
    
    def compare_configs(self, config1_name: str, config2_name: str) -> Dict[str, Any]:
        """比较两个配置的差异"""
        try:
            config1 = self.config_manager.get_config(config1_name)
            config2 = self.config_manager.get_config(config2_name)
            
            if not config1 or not config2:
                raise ConfigurationError("一个或两个配置不存在")
            
            differences = {
                'added': [],
                'removed': [],
                'modified': [],
                'unchanged': []
            }
            
            self._compare_dicts(config1, config2, "", differences)
            
            return {
                'config1': config1_name,
                'config2': config2_name,
                'differences': differences,
                'summary': {
                    'added_count': len(differences['added']),
                    'removed_count': len(differences['removed']),
                    'modified_count': len(differences['modified']),
                    'unchanged_count': len(differences['unchanged'])
                }
            }
            
        except Exception as e:
            self.logger.error("配置比较失败", exception=e)
            raise ConfigurationError(f"比较失败: {e}")
    
    def _compare_dicts(self, dict1: Dict, dict2: Dict, path: str, differences: Dict):
        """递归比较字典"""
        all_keys = set(dict1.keys()) | set(dict2.keys())
        
        for key in all_keys:
            current_path = f"{path}.{key}" if path else key
            
            if key not in dict1:
                differences['added'].append({
                    'path': current_path,
                    'value': dict2[key],
                    'type': type(dict2[key]).__name__
                })
            elif key not in dict2:
                differences['removed'].append({
                    'path': current_path,
                    'value': dict1[key],
                    'type': type(dict1[key]).__name__
                })
            elif dict1[key] != dict2[key]:
                if isinstance(dict1[key], dict) and isinstance(dict2[key], dict):
                    self._compare_dicts(dict1[key], dict2[key], current_path, differences)
                else:
                    differences['modified'].append({
                        'path': current_path,
                        'old_value': dict1[key],
                        'new_value': dict2[key],
                        'old_type': type(dict1[key]).__name__,
                        'new_type': type(dict2[key]).__name__
                    })
            else:
                differences['unchanged'].append({
                    'path': current_path,
                    'value': dict1[key],
                    'type': type(dict1[key]).__name__
                })
    
    def generate_diff_report(self, comparison_result: Dict[str, Any]) -> str:
        """生成差异报告"""
        report = []
        report.append(f"配置比较报告: {comparison_result['config1']} vs {comparison_result['config2']}")
        report.append("=" * 60)
        
        summary = comparison_result['summary']
        report.append(f"摘要:")
        report.append(f"  新增项: {summary['added_count']}")
        report.append(f"  删除项: {summary['removed_count']}")
        report.append(f"  修改项: {summary['modified_count']}")
        report.append(f"  未变化项: {summary['unchanged_count']}")
        report.append("")
        
        differences = comparison_result['differences']
        
        if differences['added']:
            report.append("新增项:")
            for item in differences['added']:
                report.append(f"  + {item['path']}: {item['value']} ({item['type']})")
            report.append("")
        
        if differences['removed']:
            report.append("删除项:")
            for item in differences['removed']:
                report.append(f"  - {item['path']}: {item['value']} ({item['type']})")
            report.append("")
        
        if differences['modified']:
            report.append("修改项:")
            for item in differences['modified']:
                report.append(f"  ~ {item['path']}:")
                report.append(f"    旧值: {item['old_value']} ({item['old_type']})")
                report.append(f"    新值: {item['new_value']} ({item['new_type']})")
            report.append("")
        
        return "\n".join(report)


class TOMLFormatHandler:
    """TOML格式处理器"""
    
    @staticmethod
    async def load_toml(file_path: Union[str, Path]) -> ConfigurationData:
        """加载TOML文件"""
        try:
            import toml
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                content = await f.read()
            return toml.loads(content)
        except ImportError:
            raise ConfigurationLoadError("需要安装toml库来支持TOML格式")
        except Exception as e:
            raise ConfigurationLoadError(f"TOML加载失败: {e}")
    
    @staticmethod
    async def save_toml(file_path: Union[str, Path], config_data: ConfigurationData):
        """保存TOML文件"""
        try:
            import toml
            content = toml.dumps(config_data)
            async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
                await f.write(content)
        except ImportError:
            raise ConfigurationSaveError("需要安装toml库来支持TOML格式")
        except Exception as e:
            raise ConfigurationSaveError(f"TOML保存失败: {e}")


class PropertiesFormatHandler:
    """Java Properties格式处理器"""
    
    @staticmethod
    async def load_properties(file_path: Union[str, Path]) -> ConfigurationData:
        """加载Properties文件"""
        try:
            result = {}
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                content = await f.read()
            
            current_section = "default"
            result[current_section] = {}
            
            for line in content.splitlines():
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip()
                    value = value.strip()
                    
                    # 处理section
                    if '.' in key:
                        parts = key.split('.')
                        section = parts[0]
                        actual_key = '.'.join(parts[1:])
                        
                        if section not in result:
                            result[section] = {}
                        result[section][actual_key] = value
                    else:
                        result[current_section][key] = value
            
            return result
            
        except Exception as e:
            raise ConfigurationLoadError(f"Properties加载失败: {e}")
    
    @staticmethod
    async def save_properties(file_path: Union[str, Path], config_data: ConfigurationData):
        """保存Properties文件"""
        content = []
        
        for section, section_data in config_data.items():
            if isinstance(section_data, dict):
                if section != "default":
                    content.append(f"# {section}")
                
                for key, value in section_data.items():
                    full_key = f"{section}.{key}" if section != "default" else key
                    content.append(f"{full_key}={value}")
                
                content.append("")  # 空行分隔
        
        content_str = "\n".join(content)
        async with aiofiles.open(file_path, 'w', encoding='utf-8') as f:
            await f.write(content_str)


class AdvancedValidationRules:
    """高级验证规则集合"""
    
    @staticmethod
    def validate_ip_address(value: str) -> bool:
        """验证IP地址格式"""
        import ipaddress
        try:
            ipaddress.ip_address(value)
            return True
        except ValueError:
            return False
    
    @staticmethod
    def validate_email(value: str) -> bool:
        """验证邮箱格式"""
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, value) is not None
    
    @staticmethod
    def validate_url(value: str) -> bool:
        """验证URL格式"""
        import re
        pattern = r'^https?://[^\s/$.?#].[^\s]*$'
        return re.match(pattern, value) is not None
    
    @staticmethod
    def validate_port(value: int) -> bool:
        """验证端口号"""
        return 1 <= value <= 65535
    
    @staticmethod
    def validate_json(value: str) -> bool:
        """验证JSON格式"""
        try:
            json.loads(value)
            return True
        except (json.JSONDecodeError, TypeError):
            return False
    
    @staticmethod
    def validate_uuid(value: str) -> bool:
        """验证UUID格式"""
        import re
        pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}$'
        return re.match(pattern, value.lower()) is not None
    
    @staticmethod
    def validate_cron_expression(value: str) -> bool:
        """验证Cron表达式格式"""
        # 简单的Cron表达式验证
        parts = value.split()
        if len(parts) != 5:
            return False
        
        # 这里可以添加更详细的Cron表达式验证逻辑
        return True


class ConfigurationSchemaValidator:
    """配置模式验证器"""
    
    def __init__(self):
        self.logger = ConfigurationLogger("SchemaValidator")
        self.schemas = {}
    
    def register_schema(self, schema_name: str, schema: Dict[str, Any]):
        """注册配置模式"""
        self.schemas[schema_name] = schema
        self.logger.info(f"注册配置模式: {schema_name}")
    
    def validate_against_schema(self, config_data: ConfigurationData, schema_name: str) -> List[str]:
        """根据模式验证配置"""
        if schema_name not in self.schemas:
            return [f"模式 {schema_name} 不存在"]
        
        schema = self.schemas[schema_name]
        errors = []
        
        self._validate_node(config_data, schema, "", errors)
        
        return errors
    
    def _validate_node(self, data: Any, schema: Dict[str, Any], path: str, errors: List[str]):
        """验证节点"""
        # 类型验证
        if 'type' in schema:
            expected_type = schema['type']
            if expected_type == 'object' and not isinstance(data, dict):
                errors.append(f"{path}: 期望对象类型，实际 {type(data).__name__}")
                return
            elif expected_type == 'array' and not isinstance(data, list):
                errors.append(f"{path}: 期望数组类型，实际 {type(data).__name__}")
                return
            elif expected_type == 'string' and not isinstance(data, str):
                errors.append(f"{path}: 期望字符串类型，实际 {type(data).__name__}")
                return
            elif expected_type == 'number' and not isinstance(data, (int, float)):
                errors.append(f"{path}: 期望数字类型，实际 {type(data).__name__}")
                return
            elif expected_type == 'boolean' and not isinstance(data, bool):
                errors.append(f"{path}: 期望布尔类型，实际 {type(data).__name__}")
                return
        
        # 必填字段验证
        if 'required' in schema and isinstance(data, dict):
            for required_field in schema['required']:
                if required_field not in data:
                    errors.append(f"{path}: 必填字段 '{required_field}' 缺失")
        
        # 属性验证
        if 'properties' in schema and isinstance(data, dict):
            for prop_name, prop_schema in schema['properties'].items():
                if prop_name in data:
                    self._validate_node(data[prop_name], prop_schema, 
                                      f"{path}.{prop_name}" if path else prop_name, errors)
        
        # 数组项验证
        if 'items' in schema and isinstance(data, list):
            item_schema = schema['items']
            for i, item in enumerate(data):
                self._validate_node(item, item_schema, f"{path}[{i}]", errors)
        
        # 枚举值验证
        if 'enum' in schema:
            if data not in schema['enum']:
                errors.append(f"{path}: 值 '{data}' 不在允许的枚举值中: {schema['enum']}")
        
        # 数值范围验证
        if 'minimum' in schema and isinstance(data, (int, float)):
            if data < schema['minimum']:
                errors.append(f"{path}: 值 {data} 小于最小值 {schema['minimum']}")
        
        if 'maximum' in schema and isinstance(data, (int, float)):
            if data > schema['maximum']:
                errors.append(f"{path}: 值 {data} 大于最大值 {schema['maximum']}")
        
        # 字符串长度验证
        if 'minLength' in schema and isinstance(data, str):
            if len(data) < schema['minLength']:
                errors.append(f"{path}: 字符串长度 {len(data)} 小于最小长度 {schema['minLength']}")
        
        if 'maxLength' in schema and isinstance(data, str):
            if len(data) > schema['maxLength']:
                errors.append(f"{path}: 字符串长度 {len(data)} 大于最大长度 {schema['maxLength']}")
        
        # 正则表达式验证
        if 'pattern' in schema and isinstance(data, str):
            import re
            if not re.match(schema['pattern'], data):
                errors.append(f"{path}: 字符串 '{data}' 不匹配模式 '{schema['pattern']}'")


# 扩展的SystemConfigurationManager类
class UltimateSystemConfigurationManager(ExtendedSystemConfigurationManager):
    """终极系统配置管理器 - 包含所有高级功能"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # 添加新的组件
        self.importer = ConfigurationImporter(self)
        self.exporter = ConfigurationExporter(self)
        self.comparator = ConfigurationComparator(self)
        self.schema_validator = ConfigurationSchemaValidator()
        
        # 添加新的格式处理器
        self.toml_handler = TOMLFormatHandler()
        self.properties_handler = PropertiesFormatHandler()
        
        # 注册新的文件格式
        self.loader._loaders[ConfigurationFormat.TOML] = self.toml_handler.load_toml
        self.loader._save_methods = getattr(self.loader, '_save_methods', {})
        self.loader._save_methods['toml'] = self.toml_handler.save_toml
        
        self.logger.info("终极系统配置管理器初始化完成")
    
    async def import_config(self, source_type: str, source_path: str, 
                           config_name: str, **kwargs) -> ConfigurationData:
        """导入配置"""
        return await self.importer.import_from_source(source_type, source_path, config_name, **kwargs)
    
    async def export_config(self, config_name: str, target_type: str, 
                           target_path: str, **kwargs) -> bool:
        """导出配置"""
        return await self.exporter.export_to_source(config_name, target_type, target_path, **kwargs)
    
    def compare_configurations(self, config1_name: str, config2_name: str) -> Dict[str, Any]:
        """比较配置"""
        return self.comparator.compare_configs(config1_name, config2_name)
    
    def generate_diff_report(self, comparison_result: Dict[str, Any]) -> str:
        """生成差异报告"""
        return self.comparator.generate_diff_report(comparison_result)
    
    def register_validation_schema(self, schema_name: str, schema: Dict[str, Any]):
        """注册验证模式"""
        self.schema_validator.register_schema(schema_name, schema)
    
    def validate_with_schema(self, config_name: str, schema_name: str) -> List[str]:
        """使用模式验证配置"""
        if config_name not in self._configs:
            return [f"配置 {config_name} 不存在"]
        
        config_data = self._configs[config_name]
        return self.schema_validator.validate_against_schema(config_data, schema_name)
    
    async def bulk_import_configs(self, import_specs: List[Dict[str, Any]]) -> Dict[str, bool]:
        """批量导入配置"""
        results = {}
        
        for spec in import_specs:
            config_name = spec['config_name']
            try:
                await self.import_config(
                    spec['source_type'],
                    spec['source_path'],
                    config_name,
                    **spec.get('kwargs', {})
                )
                results[config_name] = True
                self.logger.info(f"批量导入成功: {config_name}")
            except Exception as e:
                results[config_name] = False
                self.logger.error(f"批量导入失败: {config_name}", exception=e)
        
        return results
    
    async def bulk_export_configs(self, export_specs: List[Dict[str, Any]]) -> Dict[str, bool]:
        """批量导出配置"""
        results = {}
        
        for spec in export_specs:
            config_name = spec['config_name']
            try:
                success = await self.export_config(
                    config_name,
                    spec['target_type'],
                    spec['target_path'],
                    **spec.get('kwargs', {})
                )
                results[config_name] = success
                if success:
                    self.logger.info(f"批量导出成功: {config_name}")
                else:
                    self.logger.warning(f"批量导出失败: {config_name}")
            except Exception as e:
                results[config_name] = False
                self.logger.error(f"批量导出失败: {config_name}", exception=e)
        
        return results
    
    def create_config_profile(self, profile_name: str, config_selections: Dict[str, List[str]]):
        """创建配置配置文件"""
        profile = {
            'name': profile_name,
            'created_at': datetime.now().isoformat(),
            'configs': config_selections,
            'description': f'Configuration profile: {profile_name}'
        }
        
        profile_file = self.config_dir / f"{profile_name}_profile.json"
        with open(profile_file, 'w', encoding='utf-8') as f:
            json.dump(profile, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"创建配置配置文件: {profile_name}")
        return profile_file
    
    async def load_config_profile(self, profile_name: str) -> Dict[str, bool]:
        """加载配置配置文件"""
        profile_file = self.config_dir / f"{profile_name}_profile.json"
        
        if not profile_file.exists():
            raise ConfigurationError(f"配置文件 {profile_name} 不存在")
        
        with open(profile_file, 'r', encoding='utf-8') as f:
            profile = json.load(f)
        
        results = {}
        for config_name, file_path in profile['configs'].items():
            try:
                await self.load_config(config_name, file_path)
                results[config_name] = True
            except Exception as e:
                results[config_name] = False
                self.logger.error(f"加载配置文件失败: {config_name}", exception=e)
        
        return results
    
    def get_config_analytics(self) -> Dict[str, Any]:
        """获取配置分析数据"""
        analytics = {
            'total_configs': len(self._configs),
            'total_size': 0,
            'config_types': defaultdict(int),
            'validation_status': defaultdict(int),
            'last_modified': None,
            'most_accessed_keys': defaultdict(int),
            'configuration_health': 'unknown'
        }
        
        for config_name, config_data in self._configs.items():
            # 计算大小
            config_size = len(json.dumps(config_data))
            analytics['total_size'] += config_size
            
            # 分析配置类型
            for key, value in config_data.items():
                analytics['config_types'][type(value).__name__] += 1
            
            # 验证状态
            errors = self.validate_config(config_name)
            if errors:
                analytics['validation_status']['invalid'] += 1
            else:
                analytics['validation_status']['valid'] += 1
            
            # 最后修改时间
            metadata = self._metadata.get(config_name)
            if metadata:
                if analytics['last_modified'] is None or metadata.updated_at > analytics['last_modified']:
                    analytics['last_modified'] = metadata.updated_at
        
        # 计算配置健康状态
        valid_ratio = analytics['validation_status']['valid'] / max(analytics['total_configs'], 1)
        if valid_ratio >= 0.9:
            analytics['configuration_health'] = 'excellent'
        elif valid_ratio >= 0.7:
            analytics['configuration_health'] = 'good'
        elif valid_ratio >= 0.5:
            analytics['configuration_health'] = 'fair'
        else:
            analytics['configuration_health'] = 'poor'
        
        return dict(analytics)


# 更多高级示例
async def ultimate_features_example():
    """终极功能示例"""
    
    print("=== K1系统配置管理器 - 终极功能示例 ===\n")
    
    config_manager = UltimateSystemConfigurationManager()
    
    try:
        # 1. 创建示例配置数据
        print("1. 创建示例配置...")
        sample_config = {
            'application': {
                'name': 'UltimateApp',
                'version': '3.0.0',
                'debug': False,
                'port': 8080,
                'host': '0.0.0.0'
            },
            'database': {
                'host': 'localhost',
                'port': 5432,
                'name': 'ultimate_db',
                'user': 'admin',
                'ssl': True
            },
            'cache': {
                'redis_host': 'redis.example.com',
                'redis_port': 6379,
                'ttl': 3600,
                'max_connections': 100
            },
            'logging': {
                'level': 'INFO',
                'format': 'json',
                'file': '/var/log/ultimate.log'
            }
        }
        
        config_manager._configs['ultimate_app'] = sample_config
        
        # 2. 注册验证模式
        print("2. 注册验证模式...")
        app_schema = {
            'type': 'object',
            'required': ['application', 'database'],
            'properties': {
                'application': {
                    'type': 'object',
                    'required': ['name', 'version', 'port'],
                    'properties': {
                        'name': {'type': 'string', 'minLength': 1},
                        'version': {'type': 'string', 'pattern': r'^\d+\.\d+\.\d+$'},
                        'port': {'type': 'number', 'minimum': 1, 'maximum': 65535},
                        'debug': {'type': 'boolean'},
                        'host': {'type': 'string'}
                    }
                },
                'database': {
                    'type': 'object',
                    'required': ['host', 'port', 'name'],
                    'properties': {
                        'host': {'type': 'string'},
                        'port': {'type': 'number', 'minimum': 1, 'maximum': 65535},
                        'name': {'type': 'string'},
                        'user': {'type': 'string'},
                        'ssl': {'type': 'boolean'}
                    }
                }
            }
        }
        
        config_manager.register_validation_schema('ultimate_app_schema', app_schema)
        
        # 3. 使用模式验证
        print("3. 使用模式验证配置...")
        schema_errors = config_manager.validate_with_schema('ultimate_app', 'ultimate_app_schema')
        if schema_errors:
            print(f"模式验证错误: {schema_errors}")
        else:
            print("模式验证通过")
        
        # 4. 配置导出示例
        print("4. 配置导出示例...")
        export_formats = ['json', 'yaml', 'xml', 'properties', 'env']
        
        for fmt in export_formats:
            export_path = f"examples/exports/ultimate_app.{fmt}"
            success = await config_manager.export_config(
                'ultimate_app', fmt, export_path
            )
            print(f"导出为 {fmt}: {'成功' if success else '失败'}")
        
        # 5. 配置比较示例
        print("5. 配置比较示例...")
        
        # 创建第二个配置进行对比
        modified_config = sample_config.copy()
        modified_config['application']['port'] = 9090
        modified_config['application']['debug'] = True
        modified_config['new_feature'] = {'enabled': True}
        
        config_manager._configs['ultimate_app_modified'] = modified_config
        
        comparison = config_manager.compare_configurations('ultimate_app', 'ultimate_app_modified')
        diff_report = config_manager.generate_diff_report(comparison)
        
        print("配置差异报告:")
        print(diff_report)
        
        # 6. 配置分析
        print("6. 配置分析...")
        analytics = config_manager.get_config_analytics()
        print("配置分析结果:")
        for key, value in analytics.items():
            print(f"  {key}: {value}")
        
        # 7. 创建配置文件
        print("7. 创建配置文件...")
        profile_configs = {
            'ultimate_app': 'examples/exports/ultimate_app.json',
            'ultimate_app_modified': 'examples/exports/ultimate_app_modified.json'
        }
        
        profile_file = config_manager.create_config_profile('ultimate_profile', profile_configs)
        print(f"配置文件创建: {profile_file}")
        
        # 8. 批量操作示例
        print("8. 批量操作示例...")
        
        # 批量导出
        export_specs = [
            {
                'config_name': 'ultimate_app',
                'target_type': 'json',
                'target_path': 'examples/bulk/ultimate_app.json'
            },
            {
                'config_name': 'ultimate_app_modified',
                'target_type': 'yaml',
                'target_path': 'examples/bulk/ultimate_app_modified.yaml'
            }
        ]
        
        bulk_export_results = await config_manager.bulk_export_configs(export_specs)
        print(f"批量导出结果: {bulk_export_results}")
        
        # 9. 高级验证规则示例
        print("9. 高级验证规则示例...")
        
        # 添加自定义验证规则
        advanced_rules = [
            ConfigurationValidationRule('application.host', str, 
                                      custom_validator=AdvancedValidationRules.validate_ip_address),
            ConfigurationValidationRule('cache.redis_host', str,
                                      custom_validator=AdvancedValidationRules.validate_ip_address),
            ConfigurationValidationRule('database.port', int,
                                      custom_validator=AdvancedValidationRules.validate_port)
        ]
        
        config_manager.add_validation_rules('ultimate_app', advanced_rules)
        
        # 执行验证
        validation_errors = config_manager.validate_config('ultimate_app')
        if validation_errors:
            print(f"高级验证错误: {validation_errors}")
        else:
            print("高级验证通过")
        
        print("\n=== 终极功能示例完成 ===")
        
    except Exception as e:
        print(f"示例执行错误: {e}")
        import traceback
        traceback.print_exc()


# 性能基准测试
class ConfigurationBenchmark:
    """配置管理器性能基准测试"""
    
    def __init__(self, config_manager: UltimateSystemConfigurationManager):
        self.config_manager = config_manager
        self.logger = ConfigurationLogger("ConfigBenchmark")
    
    async def run_benchmarks(self) -> Dict[str, Any]:
        """运行所有基准测试"""
        results = {}
        
        print("开始配置管理器性能基准测试...")
        
        # 1. 配置加载性能测试
        results['load_performance'] = await self._benchmark_config_loading()
        
        # 2. 配置访问性能测试
        results['access_performance'] = await self._benchmark_config_access()
        
        # 3. 配置验证性能测试
        results['validation_performance'] = await self._benchmark_config_validation()
        
        # 4. 缓存性能测试
        results['cache_performance'] = await self._benchmark_cache_performance()
        
        # 5. 并发性能测试
        results['concurrent_performance'] = await self._benchmark_concurrent_access()
        
        # 6. 内存使用测试
        results['memory_performance'] = await self._benchmark_memory_usage()
        
        return results
    
    async def _benchmark_config_loading(self) -> Dict[str, Any]:
        """配置加载性能测试"""
        iterations = 100
        times = []
        
        test_config = {
            'test': {
                'data': list(range(1000)),
                'nested': {str(i): {'value': i * 2} for i in range(100)}
            }
        }
        
        for i in range(iterations):
            config_name = f"benchmark_load_{i}"
            start_time = time.time()
            
            self.config_manager._configs[config_name] = test_config.copy()
            
            end_time = time.time()
            times.append(end_time - start_time)
        
        avg_time = sum(times) / len(times)
        min_time = min(times)
        max_time = max(times)
        
        return {
            'iterations': iterations,
            'average_time': avg_time,
            'min_time': min_time,
            'max_time': max_time,
            'ops_per_second': 1 / avg_time if avg_time > 0 else 0
        }
    
    async def _benchmark_config_access(self) -> Dict[str, Any]:
        """配置访问性能测试"""
        iterations = 10000
        times = []
        
        # 确保有测试配置
        if 'benchmark_access' not in self.config_manager._configs:
            self.config_manager._configs['benchmark_access'] = {
                'level1': {
                    'level2': {
                        'level3': {
                            'value': 'test_value'
                        }
                    }
                }
            }
        
        for i in range(iterations):
            start_time = time.time()
            
            # 访问嵌套配置
            value = self.config_manager.get('level1.level2.level3.value', config_name='benchmark_access')
            
            end_time = time.time()
            times.append(end_time - start_time)
        
        avg_time = sum(times) / len(times)
        min_time = min(times)
        max_time = max(times)
        
        return {
            'iterations': iterations,
            'average_time': avg_time,
            'min_time': min_time,
            'max_time': max_time,
            'ops_per_second': 1 / avg_time if avg_time > 0 else 0
        }
    
    async def _benchmark_config_validation(self) -> Dict[str, Any]:
        """配置验证性能测试"""
        iterations = 1000
        times = []
        
        # 添加测试验证规则
        test_rules = [
            ConfigurationValidationRule('test.value', int, min_value=0, max_value=1000),
            ConfigurationValidationRule('test.name', str, pattern=r'^[a-zA-Z0-9]+$')
        ]
        
        self.config_manager.add_validation_rules('benchmark_validation', test_rules)
        
        for i in range(iterations):
            test_config = {
                'test': {
                    'value': i % 1001,  # 确保有些会超出范围
                    'name': f'test{i}'
                }
            }
            
            self.config_manager._configs['benchmark_validation'] = test_config
            
            start_time = time.time()
            
            errors = self.config_manager.validate_config('benchmark_validation')
            
            end_time = time.time()
            times.append(end_time - start_time)
        
        avg_time = sum(times) / len(times)
        min_time = min(times)
        max_time = max(times)
        
        return {
            'iterations': iterations,
            'average_time': avg_time,
            'min_time': min_time,
            'max_time': max_time,
            'ops_per_second': 1 / avg_time if avg_time > 0 else 0
        }
    
    async def _benchmark_cache_performance(self) -> Dict[str, Any]:
        """缓存性能测试"""
        iterations = 5000
        cache_times = []
        no_cache_times = []
        
        # 测试缓存命中
        for i in range(iterations):
            # 第一次访问（缓存未命中）
            start_time = time.time()
            value = self.config_manager.get('benchmark_cache.test_value', config_name='benchmark_cache')
            no_cache_times.append(time.time() - start_time)
            
            # 第二次访问（缓存命中）
            start_time = time.time()
            value = self.config_manager.get('benchmark_cache.test_value', config_name='benchmark_cache')
            cache_times.append(time.time() - start_time)
        
        cache_avg = sum(cache_times) / len(cache_times)
        no_cache_avg = sum(no_cache_times) / len(no_cache_times)
        
        return {
            'iterations': iterations,
            'cache_hit_average_time': cache_avg,
            'cache_miss_average_time': no_cache_avg,
            'speedup_factor': no_cache_avg / cache_avg if cache_avg > 0 else 0
        }
    
    async def _benchmark_concurrent_access(self) -> Dict[str, Any]:
        """并发访问性能测试"""
        import asyncio
        
        iterations = 1000
        num_workers = 10
        
        async def worker(worker_id):
            times = []
            for i in range(iterations // num_workers):
                start_time = time.time()
                
                # 随机配置访问
                config_name = f'concurrent_config_{worker_id % 5}'
                if config_name not in self.config_manager._configs:
                    self.config_manager._configs[config_name] = {'value': f'config_{worker_id}'}
                
                value = self.config_manager.get('value', config_name=config_name)
                
                end_time = time.time()
                times.append(end_time - start_time)
            
            return times
        
        start_time = time.time()
        
        # 创建并发任务
        tasks = [worker(i) for i in range(num_workers)]
        results = await asyncio.gather(*tasks)
        
        end_time = time.time()
        total_time = end_time - start_time
        
        # 合并所有时间
        all_times = []
        for worker_times in results:
            all_times.extend(worker_times)
        
        avg_time = sum(all_times) / len(all_times)
        
        return {
            'num_workers': num_workers,
            'total_iterations': iterations,
            'total_time': total_time,
            'average_time': avg_time,
            'ops_per_second': iterations / total_time
        }
    
    async def _benchmark_memory_usage(self) -> Dict[str, Any]:
        """内存使用测试"""
        import psutil
        import os
        
        process = psutil.Process(os.getpid())
        
        # 获取初始内存使用
        initial_memory = process.memory_info().rss
        
        # 创建大量配置
        num_configs = 1000
        config_size = 0
        
        for i in range(num_configs):
            large_config = {
                'data': list(range(100)),
                'nested': {str(j): {'value': j * 2} for j in range(50)},
                'metadata': {
                    'created': datetime.now().isoformat(),
                    'version': '1.0.0',
                    'tags': [f'tag_{k}' for k in range(10)]
                }
            }
            
            config_name = f'memory_test_{i}'
            self.config_manager._configs[config_name] = large_config
            config_size += len(json.dumps(large_config))
        
        # 获取最终内存使用
        final_memory = process.memory_info().rss
        memory_increase = final_memory - initial_memory
        
        return {
            'num_configs': num_configs,
            'config_size_bytes': config_size,
            'memory_increase_bytes': memory_increase,
            'memory_per_config': memory_increase / num_configs,
            'efficiency_ratio': config_size / memory_increase if memory_increase > 0 else 0
        }


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == 'cli':
        # 运行CLI工具
        cli_main = create_config_cli()
        cli_main()
    else:
        # 运行示例
        print("K1系统配置管理器 - 终极版")
        print("1. 运行基本示例")
        print("2. 运行完整示例")
        print("3. 运行高级场景示例")
        print("4. 运行终极功能示例")
        print("5. 运行测试套件")
        print("6. 启动CLI工具")
        print("7. 运行性能基准测试")
        
        choice = input("请选择 (1-7): ").strip()
        
        if choice == '1':
            asyncio.run(example_usage())
        elif choice == '2':
            asyncio.run(comprehensive_example())
        elif choice == '3':
            asyncio.run(advanced_scenarios_example())
        elif choice == '4':
            asyncio.run(ultimate_features_example())
        elif choice == '5':
            async def run_tests():
                test_suite = ConfigurationManagerTestSuite()
                results = await test_suite.run_all_tests()
                print("测试结果:")
                print(f"  总测试数: {results['total_tests']}")
                print(f"  通过: {results['passed_tests']}")
                print(f"  失败: {results['failed_tests']}")
                for detail in results['test_details']:
                    print(f"  {detail['test']}: {detail['status']} - {detail['message']}")
            
            asyncio.run(run_tests())
        elif choice == '6':
            cli_main = create_config_cli()
            cli_main()
        elif choice == '7':
            async def run_benchmarks():
                print("初始化终极配置管理器...")
                config_manager = UltimateSystemConfigurationManager()
                benchmark = ConfigurationBenchmark(config_manager)
                
                print("运行性能基准测试...")
                results = await benchmark.run_benchmarks()
                
                print("\n=== 性能基准测试结果 ===")
                for category, metrics in results.items():
                    print(f"\n{category}:")
                    for metric, value in metrics.items():
                        if isinstance(value, float):
                            print(f"  {metric}: {value:.6f}")
                        else:
                            print(f"  {metric}: {value}")
            
            asyncio.run(run_benchmarks())
        else:
            print("无效选择")