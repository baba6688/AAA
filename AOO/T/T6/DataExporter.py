#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
T6数据导出器 (Data Exporter)

一个高性能、多格式的数据导出工具，支持多种数据格式导出、自动化调度、
流式导出、压缩加密、性能优化等功能。

功能特性:
- 多格式数据导出 (CSV, Excel, JSON, XML, Parquet, HDF5等)
- 数据导出调度和自动化
- 数据分批导出和流式导出
- 数据导出压缩和加密
- 数据导出性能优化
- 数据导出错误处理
- 数据导出进度监控
- 数据导出安全控制
- 数据导出日志和审计


日期: 2025-11-05
版本: 1.0.0
"""

import os
import io
import csv
import json
import xml
import gzip
import zipfile
import logging
import hashlib
import threading
import schedule
import time
import traceback
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from typing import (
    Any, Dict, List, Optional, Union, Callable, Iterator, 
    Tuple, Type, BinaryIO, TextIO, Generator
)
from dataclasses import dataclass, field
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
import warnings
import tempfile
import shutil
from functools import wraps

# 第三方库导入（可选）
try:
    import pandas as pd
    import numpy as np
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

try:
    import h5py
    HAS_H5PY = True
except ImportError:
    HAS_H5PY = False

try:
    from cryptography.fernet import Fernet
    HAS_CRYPTO = True
except ImportError:
    HAS_CRYPTO = False

try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False


@dataclass
class ExportConfig:
    """数据导出配置类"""
    # 基本配置
    format: str = "csv"  # 导出格式: csv, excel, json, xml, parquet, hdf5
    compression: Optional[str] = None  # 压缩格式: gzip, zip, bz2
    encryption_key: Optional[str] = None  # 加密密钥
    encoding: str = "utf-8"  # 文件编码
    
    # 分批导出配置
    batch_size: int = 10000  # 批次大小
    enable_streaming: bool = False  # 是否启用流式导出
    chunk_size: int = 1000  # 流式导出块大小
    
    # 性能优化配置
    max_workers: int = 4  # 最大工作线程数
    enable_compression: bool = False  # 是否启用压缩
    memory_limit_mb: int = 512  # 内存限制(MB)
    
    # 调度配置
    schedule_type: Optional[str] = None  # 调度类型: daily, weekly, monthly, cron
    schedule_time: Optional[str] = None  # 调度时间
    schedule_cron: Optional[str] = None  # Cron表达式
    
    # 安全配置
    enable_audit: bool = True  # 是否启用审计
    enable_access_control: bool = True  # 是否启用访问控制
    allowed_users: List[str] = field(default_factory=list)  # 允许的用户列表
    
    # 监控配置
    enable_progress_monitor: bool = True  # 是否启用进度监控
    progress_callback: Optional[Callable] = None  # 进度回调函数
    enable_performance_monitor: bool = True  # 是否启用性能监控
    
    # 日志配置
    log_level: str = "INFO"  # 日志级别
    log_file: Optional[str] = None  # 日志文件路径
    enable_audit_log: bool = True  # 是否启用审计日志


@dataclass
class ExportResult:
    """数据导出结果类"""
    success: bool
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    export_time: Optional[float] = None
    records_exported: Optional[int] = None
    error_message: Optional[str] = None
    checksum: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


class ProgressMonitor:
    """进度监控器"""
    
    def __init__(self, total_records: int, callback: Optional[Callable] = None):
        self.total_records = total_records
        self.processed_records = 0
        self.start_time = time.time()
        self.callback = callback
        self._lock = threading.Lock()
    
    def update(self, increment: int = 1):
        """更新进度"""
        with self._lock:
            self.processed_records += increment
            if self.callback:
                progress = self.processed_records / self.total_records if self.total_records > 0 else 1.0
                elapsed_time = time.time() - self.start_time
                estimated_total_time = elapsed_time / progress if progress > 0 else 0
                remaining_time = estimated_total_time - elapsed_time
                
                self.callback(
                    progress=progress,
                    processed=self.processed_records,
                    total=self.total_records,
                    elapsed_time=elapsed_time,
                    remaining_time=remaining_time
                )
    
    def get_progress(self) -> float:
        """获取进度百分比"""
        return self.processed_records / self.total_records if self.total_records > 0 else 1.0


class SecurityManager:
    """安全管理器"""
    
    def __init__(self, config: ExportConfig):
        self.config = config
        self.audit_log = []
        self.access_log = []
    
    def check_access(self, user_id: str) -> bool:
        """检查用户访问权限"""
        if not self.config.enable_access_control:
            return True
        
        if not self.config.allowed_users or user_id in self.config.allowed_users:
            self._log_access(user_id, True)
            return True
        
        self._log_access(user_id, False)
        return False
    
    def _log_access(self, user_id: str, granted: bool):
        """记录访问日志"""
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "user_id": user_id,
            "access_granted": granted,
            "action": "data_export"
        }
        self.access_log.append(log_entry)
        
        if self.config.enable_audit_log:
            logging.info(f"Access log: {log_entry}")
    
    def audit_export(self, operation: str, details: Dict[str, Any]):
        """审计导出操作"""
        if not self.config.enable_audit:
            return
        
        audit_entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": operation,
            "details": details,
            "user_id": details.get("user_id", "unknown")
        }
        self.audit_log.append(audit_entry)
        
        if self.config.enable_audit_log:
            logging.info(f"Audit log: {audit_entry}")


class PerformanceMonitor:
    """性能监控器"""
    
    def __init__(self):
        self.metrics = {
            "export_count": 0,
            "total_export_time": 0.0,
            "total_records_exported": 0,
            "average_export_time": 0.0,
            "memory_usage": [],
            "cpu_usage": []
        }
        self._lock = threading.Lock()
    
    def record_export(self, export_time: float, records_count: int, file_size: int):
        """记录导出性能指标"""
        with self._lock:
            self.metrics["export_count"] += 1
            self.metrics["total_export_time"] += export_time
            self.metrics["total_records_exported"] += records_count
            self.metrics["average_export_time"] = (
                self.metrics["total_export_time"] / self.metrics["export_count"]
            )
            
            if HAS_PSUTIL:
                process = psutil.Process()
                self.metrics["memory_usage"].append(process.memory_info().rss / 1024 / 1024)  # MB
                self.metrics["cpu_usage"].append(process.cpu_percent())
    
    def get_metrics(self) -> Dict[str, Any]:
        """获取性能指标"""
        with self._lock:
            return self.metrics.copy()
    
    def reset_metrics(self):
        """重置性能指标"""
        with self._lock:
            self.metrics = {
                "export_count": 0,
                "total_export_time": 0.0,
                "total_records_exported": 0,
                "average_export_time": 0.0,
                "memory_usage": [],
                "cpu_usage": []
            }


class FormatHandler(ABC):
    """格式处理器抽象基类"""
    
    @abstractmethod
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出数据"""
        pass
    
    @abstractmethod
    def get_supported_formats(self) -> List[str]:
        """获取支持的格式"""
        pass


class CSVHandler(FormatHandler):
    """CSV格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出CSV格式数据"""
        start_time = time.time()
        
        try:
            if HAS_PANDAS and isinstance(data, (pd.DataFrame, pd.Series)):
                # 使用pandas导出
                data.to_csv(file_path, index=False, encoding=config.encoding, 
                           compression=config.compression if config.enable_compression else None)
                records_count = len(data)
            else:
                # 手动导出CSV
                records_count = 0
                with open(file_path, 'w', newline='', encoding=config.encoding) as csvfile:
                    writer = csv.writer(csvfile)
                    
                    if isinstance(data, list):
                        for row in data:
                            writer.writerow(row)
                            records_count += 1
                    elif isinstance(data, dict):
                        writer.writerow(data.keys())
                        writer.writerow(data.values())
                        records_count = 1
                    else:
                        writer.writerow([data])
                        records_count = 1
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["csv", "tsv"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class ExcelHandler(FormatHandler):
    """Excel格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出Excel格式数据"""
        if not HAS_OPENPYXL:
            return ExportResult(
                success=False,
                error_message="openpyxl库未安装，无法导出Excel格式"
            )
        
        start_time = time.time()
        
        try:
            if HAS_PANDAS and isinstance(data, pd.DataFrame):
                # 使用pandas导出Excel
                data.to_excel(file_path, index=False, engine='openpyxl')
                records_count = len(data)
            else:
                # 使用openpyxl手动导出
                wb = Workbook()
                ws = wb.active
                ws.title = "Data Export"
                
                records_count = 0
                if isinstance(data, list):
                    for row_idx, row in enumerate(data, 1):
                        for col_idx, value in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        records_count += 1
                elif isinstance(data, dict):
                    for col_idx, (key, value) in enumerate(data.items(), 1):
                        ws.cell(row=1, column=col_idx, value=key)
                        ws.cell(row=2, column=col_idx, value=value)
                    records_count = 1
                else:
                    ws.cell(row=1, column=1, value=data)
                    records_count = 1
                
                wb.save(file_path)
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["xlsx", "xls"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class JSONHandler(FormatHandler):
    """JSON格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出JSON格式数据"""
        start_time = time.time()
        
        try:
            if HAS_PANDAS and isinstance(data, pd.DataFrame):
                # DataFrame转换为JSON
                json_data = data.to_json(orient='records', force_ascii=False, indent=2)
                records_count = len(data)
            elif HAS_PANDAS and isinstance(data, pd.Series):
                json_data = data.to_json(force_ascii=False, indent=2)
                records_count = len(data)
            else:
                # 直接序列化
                json_data = json.dumps(data, ensure_ascii=False, indent=2, default=str)
                if isinstance(data, list):
                    records_count = len(data)
                elif isinstance(data, dict):
                    records_count = len(data)
                else:
                    records_count = 1
            
            # 写入文件
            with open(file_path, 'w', encoding=config.encoding) as f:
                f.write(json_data)
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["json"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class XMLHandler(FormatHandler):
    """XML格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出XML格式数据"""
        start_time = time.time()
        
        try:
            import xml.etree.ElementTree as ET
            
            # 创建根元素
            root = ET.Element("data")
            
            if HAS_PANDAS and isinstance(data, pd.DataFrame):
                # DataFrame转换为XML
                for _, row in data.iterrows():
                    record = ET.SubElement(root, "record")
                    for col, value in row.items():
                        field = ET.SubElement(record, col)
                        field.text = str(value)
                records_count = len(data)
            elif isinstance(data, list):
                # 列表转换为XML
                for item in data:
                    record = ET.SubElement(root, "item")
                    if isinstance(item, dict):
                        for key, value in item.items():
                            field = ET.SubElement(record, key)
                            field.text = str(value)
                    else:
                        record.text = str(item)
                records_count = len(data)
            elif isinstance(data, dict):
                # 字典转换为XML
                for key, value in data.items():
                    field = ET.SubElement(root, key)
                    field.text = str(value)
                records_count = 1
            else:
                # 单值转换为XML
                root.text = str(data)
                records_count = 1
            
            # 写入文件
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ", level=0)
            tree.write(file_path, encoding=config.encoding, xml_declaration=True)
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["xml"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class ParquetHandler(FormatHandler):
    """Parquet格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出Parquet格式数据"""
        if not HAS_PYARROW:
            return ExportResult(
                success=False,
                error_message="pyarrow库未安装，无法导出Parquet格式"
            )
        
        start_time = time.time()
        
        try:
            if HAS_PANDAS and isinstance(data, pd.DataFrame):
                # 使用pandas导出Parquet
                data.to_parquet(file_path, compression=config.compression if config.enable_compression else None)
                records_count = len(data)
            else:
                # 转换为pandas DataFrame再导出
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data])
                else:
                    df = pd.DataFrame([data])
                
                df.to_parquet(file_path, compression=config.compression if config.enable_compression else None)
                records_count = len(df)
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["parquet"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class HDF5Handler(FormatHandler):
    """HDF5格式处理器"""
    
    def export(self, data: Any, file_path: str, config: ExportConfig) -> ExportResult:
        """导出HDF5格式数据"""
        if not HAS_H5PY:
            return ExportResult(
                success=False,
                error_message="h5py库未安装，无法导出HDF5格式"
            )
        
        start_time = time.time()
        
        try:
            with h5py.File(file_path, 'w') as f:
                if HAS_PANDAS and isinstance(data, pd.DataFrame):
                    # DataFrame存储到HDF5
                    f.create_dataset('data', data=data.values)
                    f.create_dataset('columns', data=[col.encode() for col in data.columns])
                    f.create_dataset('index', data=data.index.values)
                    records_count = len(data)
                elif isinstance(data, list):
                    # 列表存储到HDF5
                    if data and isinstance(data[0], (list, tuple)):
                        # 二维数据
                        f.create_dataset('data', data=np.array(data))
                        records_count = len(data)
                    else:
                        # 一维数据
                        f.create_dataset('data', data=np.array(data))
                        records_count = len(data)
                elif isinstance(data, dict):
                    # 字典存储到HDF5
                    for key, value in data.items():
                        f.create_dataset(f'item_{key}', data=value)
                    records_count = len(data)
                else:
                    # 单值存储到HDF5
                    f.create_dataset('data', data=data)
                    records_count = 1
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=self._calculate_checksum(file_path)
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=time.time() - start_time
            )
    
    def get_supported_formats(self) -> List[str]:
        return ["hdf5", "h5"]
    
    def _calculate_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()


class CompressionManager:
    """压缩管理器"""
    
    @staticmethod
    def compress_file(file_path: str, compression_type: str) -> str:
        """压缩文件"""
        compressed_path = f"{file_path}.{compression_type}"
        
        if compression_type == "gzip":
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        elif compression_type == "bz2":
            import bz2
            with open(file_path, 'rb') as f_in:
                with bz2.open(compressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        elif compression_type == "zip":
            with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED) as f:
                f.write(file_path, os.path.basename(file_path))
        else:
            raise ValueError(f"不支持的压缩格式: {compression_type}")
        
        # 删除原文件
        os.remove(file_path)
        return compressed_path
    
    @staticmethod
    def decompress_file(compressed_path: str) -> str:
        """解压文件"""
        if compressed_path.endswith('.gz'):
            decompressed_path = compressed_path[:-3]
            with gzip.open(compressed_path, 'rb') as f_in:
                with open(decompressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        elif compressed_path.endswith('.bz2'):
            import bz2
            decompressed_path = compressed_path[:-4]
            with bz2.open(compressed_path, 'rb') as f_in:
                with open(decompressed_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        elif compressed_path.endswith('.zip'):
            import zipfile
            decompressed_path = compressed_path[:-4]
            with zipfile.ZipFile(compressed_path, 'r') as f:
                f.extractall(os.path.dirname(decompressed_path))
                # 获取zip中的文件名
                zip_names = f.namelist()
                if zip_names:
                    extracted_file = os.path.join(os.path.dirname(decompressed_path), zip_names[0])
                    if extracted_file != decompressed_path:
                        shutil.move(extracted_file, decompressed_path)
        else:
            raise ValueError(f"不支持的压缩格式: {compressed_path}")
        
        # 删除压缩文件
        os.remove(compressed_path)
        return decompressed_path


class EncryptionManager:
    """加密管理器"""
    
    def __init__(self, key: Optional[str] = None):
        self.key = key or Fernet.generate_key() if HAS_CRYPTO else None
        self.cipher = Fernet(self.key) if HAS_CRYPTO and self.key else None
    
    def encrypt_file(self, file_path: str) -> str:
        """加密文件"""
        if not HAS_CRYPTO or not self.cipher:
            raise RuntimeError("加密功能不可用，请安装cryptography库")
        
        encrypted_path = f"{file_path}.encrypted"
        
        with open(file_path, 'rb') as f:
            data = f.read()
        
        encrypted_data = self.cipher.encrypt(data)
        
        with open(encrypted_path, 'wb') as f:
            f.write(encrypted_data)
        
        # 删除原文件
        os.remove(file_path)
        return encrypted_path
    
    def decrypt_file(self, encrypted_path: str) -> str:
        """解密文件"""
        if not HAS_CRYPTO or not self.cipher:
            raise RuntimeError("解密功能不可用，请安装cryptography库")
        
        decrypted_path = encrypted_path.replace('.encrypted', '')
        
        with open(encrypted_path, 'rb') as f:
            encrypted_data = f.read()
        
        decrypted_data = self.cipher.decrypt(encrypted_data)
        
        with open(decrypted_path, 'wb') as f:
            f.write(decrypted_data)
        
        # 删除加密文件
        os.remove(encrypted_path)
        return decrypted_path


class DataExporter:
    """T6数据导出器主类"""
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """
        初始化数据导出器
        
        Args:
            config: 导出配置，如果为None则使用默认配置
        """
        self.config = config or ExportConfig()
        self.logger = self._setup_logging()
        self.security_manager = SecurityManager(self.config)
        self.performance_monitor = PerformanceMonitor()
        self.compression_manager = CompressionManager()
        self.encryption_manager = EncryptionManager(self.config.encryption_key)
        
        # 初始化格式处理器
        self.format_handlers = self._initialize_format_handlers()
        
        # 调度器
        self.scheduler_thread = None
        self.scheduler_running = False
        
        self.logger.info("数据导出器初始化完成")
    
    def _setup_logging(self) -> logging.Logger:
        """设置日志"""
        logger = logging.getLogger("DataExporter")
        logger.setLevel(getattr(logging, self.config.log_level.upper()))
        
        # 清除已存在的处理器
        logger.handlers.clear()
        
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(getattr(logging, self.config.log_level.upper()))
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        
        # 文件处理器
        if self.config.log_file:
            file_handler = logging.FileHandler(self.config.log_file, encoding='utf-8')
            file_handler.setLevel(getattr(logging, self.config.log_level.upper()))
            file_handler.setFormatter(formatter)
            logger.addHandler(file_handler)
        
        return logger
    
    def _initialize_format_handlers(self) -> Dict[str, FormatHandler]:
        """初始化格式处理器"""
        handlers = {}
        
        # 注册各种格式处理器
        csv_handler = CSVHandler()
        for fmt in csv_handler.get_supported_formats():
            handlers[fmt] = csv_handler
        
        if HAS_OPENPYXL:
            excel_handler = ExcelHandler()
            for fmt in excel_handler.get_supported_formats():
                handlers[fmt] = excel_handler
        
        json_handler = JSONHandler()
        for fmt in json_handler.get_supported_formats():
            handlers[fmt] = json_handler
        
        xml_handler = XMLHandler()
        for fmt in xml_handler.get_supported_formats():
            handlers[fmt] = xml_handler
        
        if HAS_PYARROW:
            parquet_handler = ParquetHandler()
            for fmt in parquet_handler.get_supported_formats():
                handlers[fmt] = parquet_handler
        
        if HAS_H5PY:
            hdf5_handler = HDF5Handler()
            for fmt in hdf5_handler.get_supported_formats():
                handlers[fmt] = hdf5_handler
        
        return handlers
    
    def export_data(
        self, 
        data: Any, 
        file_path: str, 
        format: Optional[str] = None,
        user_id: str = "system"
    ) -> ExportResult:
        """
        导出数据到文件
        
        Args:
            data: 要导出的数据
            file_path: 输出文件路径
            format: 导出格式，如果为None则从文件扩展名推断
            user_id: 用户ID，用于安全检查和审计
        
        Returns:
            ExportResult: 导出结果
        """
        # 安全检查
        if not self.security_manager.check_access(user_id):
            return ExportResult(
                success=False,
                error_message=f"用户 {user_id} 没有权限执行导出操作"
            )
        
        # 数据验证
        is_valid, validation_msg = self.validate_data(data)
        if not is_valid:
            return ExportResult(
                success=False,
                error_message=f"数据验证失败: {validation_msg}"
            )
        
        # 确定格式
        export_format = format or self._infer_format_from_path(file_path)
        if export_format not in self.format_handlers:
            return ExportResult(
                success=False,
                error_message=f"不支持的导出格式: {export_format}"
            )
        
        self.logger.info(f"开始导出数据到 {file_path}，格式: {export_format}")
        
        # 审计日志
        self.security_manager.audit_export("data_export", {
            "user_id": user_id,
            "file_path": file_path,
            "format": export_format,
            "data_type": type(data).__name__
        })
        
        try:
            # 创建输出目录
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 获取格式处理器
            handler = self.format_handlers[export_format]
            
            # 执行导出
            result = handler.export(data, file_path, self.config)
            
            # 压缩处理
            if result.success and self.config.enable_compression and self.config.compression:
                try:
                    compressed_path = self.compression_manager.compress_file(
                        result.file_path, self.config.compression
                    )
                    result.file_path = compressed_path
                    result.file_size = os.path.getsize(compressed_path)
                    self.logger.info(f"文件已压缩: {compressed_path}")
                except Exception as e:
                    self.logger.warning(f"压缩失败: {e}")
            
            # 加密处理
            if result.success and self.config.encryption_key:
                try:
                    if HAS_CRYPTO:
                        encrypted_path = self.encryption_manager.encrypt_file(result.file_path)
                        result.file_path = encrypted_path
                        result.file_size = os.path.getsize(encrypted_path)
                        self.logger.info(f"文件已加密: {encrypted_path}")
                    else:
                        self.logger.warning("加密功能不可用，请安装cryptography库")
                except Exception as e:
                    self.logger.warning(f"加密失败: {e}")
            
            # 记录性能指标
            if result.success:
                self.performance_monitor.record_export(
                    result.export_time or 0,
                    result.records_exported or 0,
                    result.file_size or 0
                )
            
            if result.success:
                self.logger.info(f"数据导出成功: {file_path}")
            else:
                self.logger.error(f"数据导出失败: {result.error_message}")
            
            return result
            
        except Exception as e:
            error_msg = f"导出过程中发生错误: {str(e)}"
            self.logger.error(error_msg)
            self.logger.error(traceback.format_exc())
            
            return ExportResult(
                success=False,
                error_message=error_msg
            )
    
    def export_data_batch(
        self, 
        data_list: List[Any], 
        output_dir: str, 
        format: Optional[str] = None,
        user_id: str = "system"
    ) -> List[ExportResult]:
        """
        批量导出数据
        
        Args:
            data_list: 数据列表
            output_dir: 输出目录
            format: 导出格式
            user_id: 用户ID
        
        Returns:
            List[ExportResult]: 导出结果列表
        """
        self.logger.info(f"开始批量导出 {len(data_list)} 个数据集")
        
        # 创建输出目录
        os.makedirs(output_dir, exist_ok=True)
        
        results = []
        
        # 多线程批量导出
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            # 提交任务
            future_to_index = {}
            for i, data in enumerate(data_list):
                file_path = os.path.join(output_dir, f"export_{i:04d}.{format or 'csv'}")
                future = executor.submit(self.export_data, data, file_path, format, user_id)
                future_to_index[future] = i
            
            # 收集结果
            for future in as_completed(future_to_index):
                index = future_to_index[future]
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    results.append(ExportResult(
                        success=False,
                        error_message=f"批量导出第 {index} 项时发生错误: {str(e)}"
                    ))
        
        success_count = sum(1 for r in results if r.success)
        self.logger.info(f"批量导出完成: {success_count}/{len(results)} 成功")
        
        return results
    
    def export_data_streaming(
        self, 
        data_stream: Iterator[Any], 
        file_path: str, 
        format: Optional[str] = None,
        user_id: str = "system"
    ) -> ExportResult:
        """
        流式导出数据
        
        Args:
            data_stream: 数据流
            file_path: 输出文件路径
            format: 导出格式
            user_id: 用户ID
        
        Returns:
            ExportResult: 导出结果
        """
        self.logger.info(f"开始流式导出到 {file_path}")
        
        # 安全检查
        if not self.security_manager.check_access(user_id):
            return ExportResult(
                success=False,
                error_message=f"用户 {user_id} 没有权限执行导出操作"
            )
        
        # 确定格式
        export_format = format or self._infer_format_from_path(file_path)
        
        try:
            # 创建输出目录
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            start_time = time.time()
            records_count = 0
            
            if export_format == "csv":
                # CSV流式导出
                with open(file_path, 'w', newline='', encoding=self.config.encoding) as csvfile:
                    writer = csv.writer(csvfile)
                    
                    for chunk in self._chunk_data_stream(data_stream, self.config.chunk_size):
                        if isinstance(chunk, list) and chunk:
                            for row in chunk:
                                writer.writerow(row)
                                records_count += 1
                                
                                # 进度监控
                                if self.config.enable_progress_monitor:
                                    # 创建临时进度监控器
                                    if not hasattr(self, '_temp_progress_monitor'):
                                        self._temp_progress_monitor = ProgressMonitor(1000, self.config.progress_callback)
                                    self._temp_progress_monitor.update(1)
                        
                        elif isinstance(chunk, dict):
                            writer.writerow(chunk.values())
                            records_count += 1
                            
                            if self.config.enable_progress_monitor:
                                # 创建临时进度监控器
                                if not hasattr(self, '_temp_progress_monitor'):
                                    self._temp_progress_monitor = ProgressMonitor(1000, self.config.progress_callback)
                                self._temp_progress_monitor.update(1)
            
            elif export_format == "json":
                # JSON流式导出
                with open(file_path, 'w', encoding=self.config.encoding) as f:
                    f.write('[')
                    first = True
                    
                    for chunk in self._chunk_data_stream(data_stream, self.config.chunk_size):
                        if isinstance(chunk, list):
                            for item in chunk:
                                if not first:
                                    f.write(',')
                                json.dump(item, f, ensure_ascii=False, default=str)
                                first = False
                                records_count += 1
                                
                                if self.config.enable_progress_monitor:
                                    self.progress_monitor.update(1)
                        
                        elif isinstance(chunk, dict):
                            if not first:
                                f.write(',')
                            json.dump(chunk, f, ensure_ascii=False, default=str)
                            first = False
                            records_count += 1
                            
                            if self.config.enable_progress_monitor:
                                # 创建临时进度监控器
                                if not hasattr(self, '_temp_progress_monitor'):
                                    self._temp_progress_monitor = ProgressMonitor(1000, self.config.progress_callback)
                                self._temp_progress_monitor.update(1)
                    
                    f.write(']')
            
            else:
                # 其他格式不支持流式导出
                return ExportResult(
                    success=False,
                    error_message=f"格式 {export_format} 不支持流式导出"
                )
            
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            # 计算校验和
            checksum = self._calculate_file_checksum(file_path)
            
            result = ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                export_time=export_time,
                records_exported=records_count,
                checksum=checksum
            )
            
            # 记录性能指标
            self.performance_monitor.record_export(export_time, records_count, file_size)
            
            # 审计日志
            self.security_manager.audit_export("streaming_export", {
                "user_id": user_id,
                "file_path": file_path,
                "format": export_format,
                "records_count": records_count
            })
            
            self.logger.info(f"流式导出成功: {file_path}")
            return result
            
        except Exception as e:
            error_msg = f"流式导出过程中发生错误: {str(e)}"
            self.logger.error(error_msg)
            self.logger.error(traceback.format_exc())
            
            return ExportResult(
                success=False,
                error_message=error_msg
            )
    
    def _chunk_data_stream(self, data_stream: Iterator[Any], chunk_size: int) -> Iterator[Any]:
        """将数据流分块"""
        chunk = []
        for item in data_stream:
            chunk.append(item)
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        
        # 处理剩余数据
        if chunk:
            yield chunk
    
    def _infer_format_from_path(self, file_path: str) -> str:
        """从文件路径推断格式"""
        extension = Path(file_path).suffix.lower()
        format_map = {
            '.csv': 'csv',
            '.tsv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.json': 'json',
            '.xml': 'xml',
            '.parquet': 'parquet',
            '.hdf5': 'hdf5',
            '.h5': 'hdf5'
        }
        return format_map.get(extension, 'csv')
    
    def _calculate_file_checksum(self, file_path: str) -> str:
        """计算文件校验和"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def schedule_export(
        self, 
        data_source: Callable, 
        file_path: str, 
        format: Optional[str] = None,
        user_id: str = "system"
    ) -> bool:
        """
        调度自动导出
        
        Args:
            data_source: 数据源函数
            file_path: 输出文件路径
            format: 导出格式
            user_id: 用户ID
        
        Returns:
            bool: 是否成功设置调度
        """
        if not self.config.schedule_type:
            self.logger.error("未设置调度类型")
            return False
        
        try:
            if self.config.schedule_type == "daily":
                if not self.config.schedule_time:
                    self.logger.error("未设置调度时间")
                    return False
                schedule.every().day.at(self.config.schedule_time).do(
                    self._scheduled_export, data_source, file_path, format, user_id
                )
            elif self.config.schedule_type == "weekly":
                if not self.config.schedule_time:
                    self.logger.error("未设置调度时间")
                    return False
                schedule.every().week.at(self.config.schedule_time).do(
                    self._scheduled_export, data_source, file_path, format, user_id
                )
            elif self.config.schedule_type == "monthly":
                schedule.every().month.do(
                    self._scheduled_export, data_source, file_path, format, user_id
                )
            elif self.config.schedule_type == "cron":
                if not self.config.schedule_cron:
                    self.logger.error("未设置cron表达式")
                    return False
                schedule.every().cron(self.config.schedule_cron).do(
                    self._scheduled_export, data_source, file_path, format, user_id
                )
            else:
                self.logger.error(f"不支持的调度类型: {self.config.schedule_type}")
                return False
            
            # 启动调度器线程
            if not self.scheduler_running:
                self._start_scheduler()
            
            self.logger.info(f"已设置调度导出: {self.config.schedule_type}")
            return True
            
        except Exception as e:
            self.logger.error(f"设置调度失败: {str(e)}")
            return False
    
    def _scheduled_export(
        self, 
        data_source: Callable, 
        file_path: str, 
        format: Optional[str] = None,
        user_id: str = "system"
    ):
        """执行调度的导出任务"""
        try:
            self.logger.info("开始执行调度的导出任务")
            data = data_source()
            result = self.export_data(data, file_path, format, user_id)
            
            if result.success:
                self.logger.info("调度导出任务执行成功")
            else:
                self.logger.error(f"调度导出任务执行失败: {result.error_message}")
                
        except Exception as e:
            self.logger.error(f"调度导出任务执行异常: {str(e)}")
    
    def _start_scheduler(self):
        """启动调度器线程"""
        def run_scheduler():
            self.scheduler_running = True
            while self.scheduler_running:
                schedule.run_pending()
                time.sleep(1)
        
        self.scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        self.scheduler_thread.start()
        self.logger.info("调度器已启动")
    
    def stop_scheduler(self):
        """停止调度器"""
        self.scheduler_running = False
        schedule.clear()
        if self.scheduler_thread:
            self.scheduler_thread.join(timeout=5)
        self.logger.info("调度器已停止")
    
    def get_performance_metrics(self) -> Dict[str, Any]:
        """获取性能指标"""
        return self.performance_monitor.get_metrics()
    
    def reset_performance_metrics(self):
        """重置性能指标"""
        self.performance_monitor.reset_metrics()
    
    def get_audit_logs(self) -> List[Dict[str, Any]]:
        """获取审计日志"""
        return self.security_manager.audit_log.copy()
    
    def get_access_logs(self) -> List[Dict[str, Any]]:
        """获取访问日志"""
        return self.security_manager.access_log.copy()
    
    def set_progress_callback(self, callback: Callable):
        """设置进度回调函数"""
        self.config.progress_callback = callback
    
    def validate_data(self, data: Any) -> Tuple[bool, str]:
        """
        验证数据
        
        Args:
            data: 要验证的数据
        
        Returns:
            Tuple[bool, str]: (是否有效, 错误信息)
        """
        if data is None:
            return False, "数据为空"
        
        # 检查空容器
        if isinstance(data, (list, dict, pd.DataFrame, pd.Series)):
            if len(data) == 0:
                return False, "数据为空"
        
        # pandas DataFrame和Series特殊检查
        if HAS_PANDAS:
            if isinstance(data, pd.DataFrame) and data.empty:
                return False, "DataFrame为空"
            if isinstance(data, pd.Series) and data.empty:
                return False, "Series为空"
        
        # 检查字符串类型数据的空值
        if isinstance(data, str) and not data.strip():
            return False, "字符串数据为空"
        
        return True, "数据有效"
    
    def cleanup_temp_files(self):
        """清理临时文件"""
        temp_dir = tempfile.gettempdir()
        temp_pattern = "data_export_*"
        
        try:
            import glob
            temp_files = glob.glob(os.path.join(temp_dir, temp_pattern))
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                    self.logger.info(f"已清理临时文件: {temp_file}")
                except Exception as e:
                    self.logger.warning(f"清理临时文件失败 {temp_file}: {e}")
        except Exception as e:
            self.logger.error(f"清理临时文件时发生错误: {e}")
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.stop_scheduler()
        self.cleanup_temp_files()


# 装饰器：自动重试
def retry_on_failure(max_retries: int = 3, delay: float = 1.0):
    """重试装饰器"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    if attempt < max_retries:
                        time.sleep(delay * (2 ** attempt))  # 指数退避
                        continue
                    raise last_exception
        return wrapper
    return decorator


# 示例用法和测试函数
def create_sample_data() -> List[Dict[str, Any]]:
    """创建示例数据"""
    return [
        {"id": 1, "name": "张三", "age": 25, "city": "北京"},
        {"id": 2, "name": "李四", "age": 30, "city": "上海"},
        {"id": 3, "name": "王五", "age": 28, "city": "广州"},
        {"id": 4, "name": "赵六", "age": 35, "city": "深圳"},
        {"id": 5, "name": "钱七", "age": 22, "city": "杭州"}
    ]


def data_generator() -> Iterator[Dict[str, Any]]:
    """数据生成器（用于流式导出测试）"""
    for i in range(1000):
        yield {
            "id": i,
            "value": f"数据项_{i}",
            "timestamp": datetime.now().isoformat(),
            "category": f"类别_{i % 10}"
        }


def progress_callback(progress: float, processed: int, total: int, 
                     elapsed_time: float, remaining_time: float):
    """进度回调函数示例"""
    print(f"进度: {progress:.2%} ({processed}/{total}) - "
          f"已用时: {elapsed_time:.2f}s - "
          f"剩余时间: {remaining_time:.2f}s")


def test_basic_export():
    """测试基本导出功能"""
    print("=== 测试基本导出功能 ===")
    
    # 创建配置
    config = ExportConfig(
        format="csv",
        enable_compression=True,
        compression="gzip",
        enable_progress_monitor=True,
        progress_callback=progress_callback,
        log_level="INFO"
    )
    
    # 创建导出器
    with DataExporter(config) as exporter:
        # 创建示例数据
        data = create_sample_data()
        
        # 测试CSV导出
        result = exporter.export_data(data, "/tmp/test_export.csv")
        print(f"CSV导出结果: {result.success}")
        if result.success:
            print(f"文件路径: {result.file_path}")
            print(f"文件大小: {result.file_size} bytes")
            print(f"导出时间: {result.export_time:.2f}s")
            print(f"记录数: {result.records_exported}")
            print(f"校验和: {result.checksum}")
        
        # 测试JSON导出
        result = exporter.export_data(data, "/tmp/test_export.json")
        print(f"JSON导出结果: {result.success}")
        
        # 测试Excel导出（如果可用）
        if HAS_OPENPYXL:
            result = exporter.export_data(data, "/tmp/test_export.xlsx")
            print(f"Excel导出结果: {result.success}")
        
        # 获取性能指标
        metrics = exporter.get_performance_metrics()
        print(f"性能指标: {metrics}")


def test_batch_export():
    """测试批量导出功能"""
    print("\n=== 测试批量导出功能 ===")
    
    config = ExportConfig(
        max_workers=2,
        log_level="INFO"
    )
    
    with DataExporter(config) as exporter:
        # 创建多个数据集
        data_list = [create_sample_data() for _ in range(5)]
        
        # 批量导出
        results = exporter.export_data_batch(data_list, "/tmp/batch_export")
        
        success_count = sum(1 for r in results if r.success)
        print(f"批量导出结果: {success_count}/{len(results)} 成功")


def test_streaming_export():
    """测试流式导出功能"""
    print("\n=== 测试流式导出功能 ===")
    
    config = ExportConfig(
        chunk_size=100,
        enable_progress_monitor=True,
        progress_callback=progress_callback
    )
    
    with DataExporter(config) as exporter:
        # 流式导出
        result = exporter.export_data_streaming(
            data_generator(), 
            "/tmp/streaming_export.csv"
        )
        print(f"流式导出结果: {result.success}")
        if result.success:
            print(f"文件路径: {result.file_path}")
            print(f"记录数: {result.records_exported}")


def test_scheduled_export():
    """测试调度导出功能"""
    print("\n=== 测试调度导出功能 ===")
    
    config = ExportConfig(
        schedule_type="daily",
        schedule_time="14:30",
        log_level="INFO"
    )
    
    with DataExporter(config) as exporter:
        # 设置调度导出
        success = exporter.schedule_export(
            data_source=create_sample_data,
            file_path="/tmp/scheduled_export.csv"
        )
        print(f"调度设置结果: {success}")
        
        if success:
            print("调度器已启动，30秒后停止...")
            time.sleep(30)
            exporter.stop_scheduler()


def test_security_features():
    """测试安全功能"""
    print("\n=== 测试安全功能 ===")
    
    config = ExportConfig(
        enable_access_control=True,
        allowed_users=["admin", "user1"],
        enable_audit=True,
        enable_audit_log=True
    )
    
    with DataExporter(config) as exporter:
        # 测试有权限用户
        result = exporter.export_data(create_sample_data(), "/tmp/secure_export.csv", user_id="admin")
        print(f"有权限用户导出结果: {result.success}")
        
        # 测试无权限用户
        result = exporter.export_data(create_sample_data(), "/tmp/secure_export2.csv", user_id="hacker")
        print(f"无权限用户导出结果: {result.success}")
        
        # 查看审计日志
        audit_logs = exporter.get_audit_logs()
        print(f"审计日志数量: {len(audit_logs)}")
        for log in audit_logs[-3:]:  # 显示最近3条
            print(f"审计记录: {log}")


def main():
    """主测试函数"""
    print("T6数据导出器测试")
    print("=" * 50)
    
    try:
        # 运行各种测试
        test_basic_export()
        test_batch_export()
        test_streaming_export()
        test_scheduled_export()
        test_security_features()
        
        print("\n所有测试完成!")
        
    except Exception as e:
        print(f"测试过程中发生错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()